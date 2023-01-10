from openpyxl import load_workbook
import datetime
import locale
import os 
import logging
import xml.etree.ElementTree as ET

def sibxdw():
    try:
        tree = ET.parse("C:\\Users\\User\\Desktop\\206-BS-571\\diretório\\diretorio.xml")
        root2 = tree.getroot()
        for child2 in root2:
                for x2 in root2.findall(child2.tag):
                        ansop = x2.find('ansop').text
                        moniop = x2.find('moniop').text
        moniopmeditoramento = moniop
        lista_arquivo = os.listdir(moniopmeditoramento)
        lista_data = []
        for arquivo in lista_arquivo:#pegando arquivo mais recente
            if '.xlsm' in arquivo:
                data = os.path.getmtime(f"{moniopmeditoramento}/{arquivo}")
                lista_data.append((arquivo))
        # lista_data.sort(reverse=True)
        ultimo_arquivo = lista_data[0]#variavel com o arquivo recente
        # print(ultimo_arquivo)
        for arquivo in lista_arquivo:
            # print(arquivo)
            if arquivo.endswith('.xlsm'):
                wb1 = load_workbook(moniopmeditoramento +  ultimo_arquivo)#abrindo arquivo mais recente 
                # print(moniopmeditoramento + ultimo_arquivo) 
                ws1 = wb1['Base']
        locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
        wb = load_workbook(ansop)
        ws = wb['Planilha1']


        lista_cabecalho = []
        for coluna in ws1.iter_rows(max_row= 14 , min_row=14):#pegando linha 14 para adicionar colunas

            for a in coluna:
                dt = str(a.value).strip()
                if (('PRODUTO' in dt) or ('Situação do produto' in dt) or ('Contratação' in dt) or ('Formação de Preços' in dt )):
                    continue
                else:
                    dt = dt.replace('dw bs','').replace('Variação','').replace('- Quantidade de vidas','').replace(' ','')
                dat = datetime.datetime.strptime(dt, '%b/%y')
                if dat in lista_cabecalho:
                    indice = len(lista_cabecalho)
                    coluna_ref = (ws1.cell(row = 1, column= indice).column_letter)#POSIÇÃO DA COLUNA A SER TRATADA
                    coluna_ref1 = (ws1.cell(row = 1, column= (indice * 2) + 2).column_letter)#POSIÇÃO DA COLUNA A SER TRATADA

                    break
                else:
                    lista_cabecalho.append(dat)
        contador_final = 0
        for colunas in ws1.iter_rows(min_row=15):#PEGANDO primeira linha para colocar NOVAS INSFORMAÇÕES
            colunaB = str(colunas[1].value)
            if 'None' not in colunaB:
                contador_final = contador_final + 1
        contador = contador_final + 15


        inicio = (indice * 2) + 1
        # print(inicio)
        for colunas in ws1.iter_rows(min_row=contador):
            if colunas[0].value != None:
                for row in range(contador,ws1.max_row + 1,1):#adicionando 0 
                    for column in range(2,ws1.max_column,1):
                        if (ws1.cell(row= row, column= column).value) == None:
                            letra = (ws1.cell(row = row, column= column, ).column_letter)
                            ws1[f'{letra}{row}'] = '0'
                        else:
                            break

        ############################### ADICIONANDO NOVAS INFORMAÇÕES NAS ULTIMAS COLUNAS PLANILHA moniopmediTORAMENTO##########################

        contador_final = 0
        for colunas in ws1.iter_rows(min_row=15):#PEGANDO primeira linha para colocar NOVAS INSFORMAÇÕES
            coluna = str(colunas[inicio].value)
            # print(coluna)
            if 'None' not in coluna:
                contador_final = contador_final + 1
        contador2 = contador_final + 15



        col_a = []
        for colunas in ws1.iter_rows(min_row=contador2):#PEGANDO ULTIMAS INFORMAÇÕES INSERIADAS NA COLUNA A PLANILHA moniopmediTORAMENTO
            coluna_A = str(colunas[0].value)
            col_a.append(coluna_A)

        con1 = []
        seg1 = []
        sit1 = []

        for colunas in ws.iter_rows(min_row=11):#PEGANDO INFORMAÇÕES DAS COLUNA A,D ,E H  PLANILHA moniopmediTORAMENTO
            pro = str(colunas[0].value).strip()
            con = str(colunas[3].value).strip()
            seg = str(colunas[4].value).strip()
            sit = str(colunas[7].value).strip()
            if pro in col_a:
                con1.append(con)
                seg1.append(seg)
                sit1.append(sit)


        for colunas in ws1.iter_rows(min_row=15):#PEGANDO primeira linha para colocar NOVAS INSFORMAÇÕES
            tamanho = len(colunas)
        seg = tamanho #posição precos
        contratacao = tamanho - 1#posicao contratacao
        situacao = tamanho - 2  #posicao situação



        contador = contador2
        for item in con1:#ADICIONANDO NOVAS INFORMAÇÕES CONTRATAÇÃO PLANILHA moniopmediTORAMENTO
            coluna_ref = (ws1.cell(row = 1, column= contratacao).column_letter)#POSIÇÃO DA COLUNA A SER TRATADA
            ws1[f'{coluna_ref}{contador}'] = item
            contador += 1

        # contador = contador2
        # for item in seg1:#ADICIONANDO NOVAS INFORMAÇÕES PREÇO PLANILHA moniopmediTORAMENTO
        #     coluna_ref1 = (ws1.cell(row = 1, column=seg).column_letter)#POSIÇÃO DA COLUNA A SER TRATADA
        #     ws1[f'{coluna_ref1}{contador}'] = item
        #     contador += 1

        contador = contador2
        for item in sit1:#ADICIONANDO NOVAS INFORMAÇÕES SITUAÇÃO PLANILHA moniopmediTORAMENTO
            coluna_ref2 = (ws1.cell(row = 1, column= situacao).column_letter)#POSIÇÃO DA COLUNA A SER TRATADA
            ws1[f'{coluna_ref2}{contador}'] = item
            contador += 1

        ##################### ADICIONANDO -100% NAS COLUNAS VARIAÇÕES MESES ##################### 

        contador_final = 0
        for colunas in ws1.iter_rows(min_row=15):#PEGANDO POSIÇÃO ONDE SERÁ INSERIDA NOVAS INSFORMAÇÕES
            col = str(colunas[inicio].value)
            if 'None' not in col:
                contador_final = contador_final + 1
        contador3 = contador_final + 15
        # print(contador3)
        for colunas in ws1.iter_rows(min_row=contador3):
            if colunas[0].value != None:
                for row in range(contador3,ws1.max_row + 1,1):#ADICONANDO -100% NAS COLUNAS
                    for column in range(inicio,ws1.max_column,1):
                        if (ws1.cell(row= row, column= column).value) == None:
                            letra = (ws1.cell(row = row, column= column, ).column_letter)
                            ws1[f'{letra}{row}'] = '-100%'
                        else:
                            break

        wb1.save(moniopmeditoramento + ultimo_arquivo)
    except Exception as e:
        logging.error('| Ocorreu um erro: | 3')
        logging.exception(str(e))

