from openpyxl import load_workbook
import datetime
import locale
import os 
import logging
import xml.etree.ElementTree as ET


def sibxdw101():
    try:
        tree = ET.parse("C:\\Users\\User\\Desktop\\206-BS-571\\diretório\\diretorio.xml")
        root2 = tree.getroot()
        for child2 in root2:
                for x2 in root2.findall(child2.tag):
                    moniop = x2.find('moniop').text
        data_atual = datetime.datetime.now()
        data = data_atual.date()
        locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
        ano = data.strftime("%Y")
        mes = data.strftime("%m")
        anoMes = data.strftime("%Y%m") 
        locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')

        moniopmeditoramento = moniop
        lista_arquivo = os.listdir(moniopmeditoramento)
        lista_data = []
        for arquivo in lista_arquivo:#pegando arquivo mais recente
            if '.xlsm' in arquivo:
                data = os.path.getmtime(f"{moniopmeditoramento}/{arquivo}")
                lista_data.append((arquivo))
        # lista_data.sort(reverse=True)
        ultimo_arquivo = lista_data[0]#variavel com o arquivo recente
        # print(ultimo_arquivo)
        for arquivo in lista_arquivo:
            # print(arquivo)
            if arquivo.endswith('.xlsm'):
                wb = load_workbook(moniopmeditoramento +  ultimo_arquivo)#abrindo arquivo mais recente  
                # print(moniopmeditoramento + ultimo_arquivo)
                ws1 = wb['Base']


        data_atual = datetime.datetime.now()
        data = data_atual.date()
        data_br = data.strftime("%b/%y")

        contador = 3
        lista_cabecalho = []
        for s in ws1.iter_rows(max_row= 14 , min_row=14):#pegando linha 14 para adicionar colunas

            for a in s:
                #inicio da tratativa dos nomes de cada coluna
                dt = str(a.value).strip()
                if (('PRODUTO' in dt) or ('Situação do produto' in dt) or ('Contratação' in dt) or ('Formação de Preços' in dt )):
                    continue
                else:
                    dt = dt.replace('dw bs','').replace('Variação','').replace('- Quantidade de vidas','').replace(' ','')
                dat = datetime.datetime.strptime(dt, '%b/%y')
                if dat in lista_cabecalho:
                    indice = len(lista_cabecalho)
                    coluna_ref3 = (ws1.cell(row = 1, column= (indice * 3 ) + 1).column_letter)
                    break
                else:
                    lista_cabecalho.append(dat)


        coluna_2 = (indice * 3)
        # print((coluna_2))



        contador = 15
        for colunas in ws1.iter_rows(min_row=15):
            coluna = str(colunas[coluna_2].value).upper().strip()
            # print(coluna)
            if not "NONE" in coluna:
                contador += 1
                continue
            else:
                ws1[f'{coluna_ref3}{contador}'] = 1
                contador += 1
            
        wb.save(moniopmeditoramento +  ultimo_arquivo)
    except Exception as e:
        logging.error('| Ocorreu um erro: | 3')
        logging.exception(str(e))