from openpyxl import load_workbook
import os 
import logging
import xml.etree.ElementTree as ET
import locale
import datetime
def sibxdw3():
    try:
        tree = ET.parse("C:\\Users\\User\\Desktop\\206-BS-571\\diretório\\diretorio.xml")
        root2 = tree.getroot()
        for child2 in root2:
                for x2 in root2.findall(child2.tag):
                        proativosop = x2.find('proativosop').text
                        moniop = x2.find('moniop').text
        data_atual = datetime.datetime.now()
        data = data_atual.date()
        locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
        ano = data.strftime("%Y")
        mes = data.strftime("%m")
        anoMes = data.strftime("%Y%m") 
        locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')

        moniopmeditoramento = proativosop #PRODUTOS ATIVOS
        arquivo=load_workbook(moniopmeditoramento)
        planilha = arquivo['Situação do Plano']

        diretorio = moniop

        lista_arquivo = os.listdir(diretorio)

        lista_data = []
        for arquivo1 in lista_arquivo:#pegando arquivo mais recente
            if '.xlsm' in arquivo1:
                data = os.path.getmtime(f"{diretorio}/{arquivo1}")
                lista_data.append((arquivo1))
        # lista_data.sort(reverse=True)
        ultimo_arquivo = lista_data[0]#variavel com o arquivo recente
        # print(ultimo_arquivo)
        for arquivo in lista_arquivo:
            # print(arquivo)
            if arquivo.endswith('.xlsm'):
                arquivo2 = load_workbook(diretorio + ultimo_arquivo)#abrindo arquivo mais recente
                # print(diretorio + ultimo_arquivo)
                aba2 = arquivo2['planilha2']



        coluna_a = []
        coluna_b = []
        coluna_c = []

        for linhas in planilha.iter_rows(min_row=2):
            col_a = str(linhas[0].value)
            if col_a == 'None':
                continue
            else:
                coluna_a.append(col_a)
            col_b = str(linhas[1].value)
            if col_b == 'NÃO LOCALIZADO':
                continue
            else:
                coluna_b.append(col_b)
            col_c = str(linhas[2].value)
            if col_c == 'None':
                continue
            else:
                coluna_c.append(col_c)


        contador = 2
        for item in coluna_a:
            aba2[f'J{contador}'] = item
            contador +=1

        contador = 2
        for item in coluna_b:
            aba2[f'K{contador}'] = item
            contador +=1

        contador = 2
        for item in coluna_c:
            aba2[f'L{contador}'] = item
            contador +=1
        aba2['J1'] = 'PRODUTO'
        aba2['K1'] = 'SITUAÇÃO DO PLANO'
        aba2['L1'] = 'VIDAS'
        arquivo2.save(diretorio + ultimo_arquivo)
    except Exception as e:
        logging.error('| Ocorreu um erro: | 3')
        logging.exception(str(e))

