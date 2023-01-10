from openpyxl import load_workbook
import datetime
import logging
import xml.etree.ElementTree as ET
import locale
import os
def sibxdw():
    try:    
        tree = ET.parse("C:\\Users\\User\\Desktop\\206-BS-571\\diretório\\diretorio.xml")
        root2 = tree.getroot()
        for child2 in root2:
                for x2 in root2.findall(child2.tag):
                        moniop = x2.find('moniop').text
        data_atual = datetime.datetime.now()
        data = data_atual.date()
        locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
        ano = data.strftime("%Y")
        mes = data.strftime("%m")
        anoMes = data.strftime("%Y%m") 
        locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
        # print(anoMes)
        moniopmeditoramento =moniop
        lista_arquivo = os.listdir(moniopmeditoramento)

        lista_data = []
        for arquivo in lista_arquivo:#pegando arquivo mais recente
            # if '.xlsm' in arquivo:
            data2 = os.path.getmtime(f"{moniopmeditoramento}/{arquivo}")
            lista_data.append((arquivo))
        # lista_data.sort(reverse=False)
        ultimo_arquivo = lista_data[0]#variavel com o arquivo recente
        # print(ultimo_arquivo)
        for arquivo2 in lista_arquivo:
            # print(arquivo)
            if arquivo2.endswith('.xlsm'):
                wb = load_workbook(moniopmeditoramento + ultimo_arquivo)#abrindo arquivo mais recente
                # print(moniopmeditoramento + ultimo_arquivo)
                planilha = wb['Base']
                wb.create_sheet('planilha2')


        # RENOMEANDO COM MES E ANO ATUAL
        data_atual = datetime.datetime.now()
        data = data_atual.date()
        data_br = data.strftime("%b/%y")#data atual mes e ano, formato PT-BR
        locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
        # print(data_br)
        lista_cabecalho = []
        for s in planilha.iter_rows(max_row= 14 , min_row=14):#pegando linha 14 para adicionar colunas
            for a in s:
                #inicio da tratativa dos nomes de cada coluna
                dt = str(a.value).strip()
                if (('PRODUTO' in dt) or ('Situação do produto' in dt) or ('Contratação' in dt) or ('Formação de Preços' in dt )):
                    continue
                else:
                    dt = dt.replace('dw bs','').replace('Variação','').replace('- Quantidade de vidas','').replace(' ','')
                dat = datetime.datetime.strptime(dt, "%b/%y")
                print(dat)
                if dat in lista_cabecalho:
                    indice = len(lista_cabecalho)
                    planilha.insert_cols(indice + 2)#adicionando coluna apos mes anterior ja processado
                    coluna_ref = (planilha.cell(row = 1, column= indice + 2).column_letter)
                    planilha[f'{coluna_ref}14'] = data_br #renomeando coluna com mes a ano atual

                    planilha.insert_cols((indice * 2) + 3) #adicionando coluna apos mes anterior ja processado
                    coluna_ref = (planilha.cell(row = 1, column= (indice * 2) + 3).column_letter)
                    planilha[f'{coluna_ref}14'] = data_br + " " + 'dw bs' #renomeando coluna com mes a ano atual

                    planilha.insert_cols((indice * 3 ) + 4) #adicionando coluna apos mes anterior ja processado
                    coluna_ref = (planilha.cell(row = 1, column= (indice * 3 ) + 4).column_letter)
                    planilha[f'{coluna_ref}14'] = 'Variação' + " " + data_br #renomeando coluna com mes a ano atual
                    break
                else:
                    lista_cabecalho.append(dat)

        wb.save(moniopmeditoramento + ultimo_arquivo)
    except Exception as e:
        logging.error('| Ocorreu um erro: | 3')
        logging.exception(str(e))
