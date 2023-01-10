from openpyxl import load_workbook
import os 
import logging
import xml.etree.ElementTree as ET
import locale
import datetime
def sibxdw2():
    try:
        tree = ET.parse("C:\\Users\\User\\Desktop\\206-BS-571\\diretório\\diretorio.xml")
        root2 = tree.getroot()
        for child2 in root2:
                for x2 in root2.findall(child2.tag):
                        dwop = x2.find('dwop').text
                        moniop = x2.find('moniop').text
        data_atual = datetime.datetime.now()
        data = data_atual.date()
        locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
        ano = data.strftime("%Y")
        mes = data.strftime("%m")
        anoMes = data.strftime("%Y%m") 

        dw = dwop
        arquivo=load_workbook(dw)
        planilha = arquivo['05 - Qtd de Segurados Ativos po']

        moniopmeditoramento = moniop

        lista_arquivo = os.listdir(moniopmeditoramento)
        lista_data = []
        for arquivo2 in lista_arquivo:#pegando arquivo mais recente
            if '.xlsm' in arquivo2:
                data = os.path.getmtime(f"{moniopmeditoramento}/{arquivo2}")
                lista_data.append((arquivo2))
        # lista_data.sort(reverse=True)
        ultimo_arquivo = lista_data[0]#variavel com o arquivo recente
        # print(ultimo_arquivo)
        for arquivo3 in lista_arquivo:
            # print(arquivo)
            if arquivo3.endswith('.xlsm'):
                arquivo_moniopmeditoramento = load_workbook(moniopmeditoramento + ultimo_arquivo)#abrindo arquivo mais recente
                # print(moniopmeditoramento + ultimo_arquivo)
                aba2 = arquivo_moniopmeditoramento['planilha2']




        coluna_a = []
        coluna_c = []
        contador = 7
        for linha in planilha.iter_rows(min_row=6):
            col_A = str(linha[0].value).strip()
            coluna_a.append(col_A)
            col_C = str(linha[2].value).strip()
            coluna_c.append(col_C)
            primeiro_num = col_A[:1]
            if '0' in primeiro_num:
                planilha[f'A{contador}'] = col_A[1:]
                contador +=1

        contador = 1
        for item in coluna_a:
            aba2[f'B{contador}'] = item
            contador +=1

        contador = 1
        for item in coluna_c:
            aba2[f'D{contador}'] = item
            contador +=1

        aba2['C1'] = 'Métrica'
        arquivo.save(dw)
        arquivo_moniopmeditoramento.save(moniopmeditoramento +  ultimo_arquivo)
    except Exception as e:
        logging.error('| Ocorreu um erro: | 3')
        logging.exception(str(e))


