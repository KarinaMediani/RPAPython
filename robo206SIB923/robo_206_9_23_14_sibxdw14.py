from openpyxl import load_workbook
import locale
import datetime
import os 
import logging
import xml.etree.ElementTree as ET
def sibxdw14():
    try:
        tree = ET.parse("C:\\Users\\User\\Desktop\\206-BS-571\\diretório\\diretorio.xml")
        root2 = tree.getroot()
        for child2 in root2:
                for x2 in root2.findall(child2.tag):
                        confeop = x2.find('confeop').text
                        moniop = x2.find('moniop').text
        data_atual = datetime.datetime.now()
        data = data_atual.date()
        locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
        ano = data.strftime("%Y")
        mes = data.strftime("%m")
        anoMes = data.strftime("%Y%m") 
        locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')

        moniopmeditoramento = moniop
        lista_arquivo = os.listdir(moniopmeditoramento)
        lista_data = []
        for arquivo in lista_arquivo:#pegando arquivo mais recente
            if '.xlsm' in arquivo:
                data = os.path.getmtime(f"{moniopmeditoramento}/{arquivo}")
                lista_data.append((arquivo))

        # lista_data.sort(reverse=True)
        ultimo_arquivo = lista_data[0]#variavel com o arquivo recente
        # print(ultimo_arquivo)
        locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
        for arquivo in lista_arquivo:
            # print(arquivo)
            if arquivo.endswith('.xlsm'):
                wb = load_workbook(moniopmeditoramento +  ultimo_arquivo)#abrindo arquivo mais recente  
                # print(moniopmeditoramento + ultimo_arquivo)
                ws = wb['planilha2']
                ws1 = wb['Base']


        contador_final = 0
        produtoPlanilha1 = []
        for colunas in ws1.iter_rows(min_row=15):#PEGANDO VALORES DA COLUNA A ABA BASE
            colunaA = str(colunas[0].value)
            if colunaA == 'None':
                continue
            else:
                produtoPlanilha1.append(colunaA)

        novos = []
        for colunas in ws.iter_rows(min_row=2):#PEGANDO VALORES EXISTENTES NA COLUNA J ABA PLANILHA2 NAO EXISTENTES NA COLUNA A DA ABA BASE
            colunaj = str(colunas[9].value)
            if ((colunaj not in produtoPlanilha1) and (colunaj != "None")):
                novos.append(colunaj)
        # print(novos)

        vidas = []
        for colunas in ws.iter_rows(min_row=2):#PEGANDO VALORES DE VIDAS NAO CONTIDAS NA ABA BASE
            colunaj = str(colunas[9].value)
            colunal = str(colunas[11].value)
            if colunaj in novos:
                vidas.append(colunal)


        for colunas in ws1.iter_rows(min_row=15, max_col=0, min_col=0):#PEGANDO POSIÇÃO ONDE SERÁ INSERIDA NOVAS INSFORMAÇÕES
            colunaA = str(colunas[0].value)
            if 'None' not in colunaA:
                contador_final = contador_final + 1
        tamanho = contador_final


        lista_cabecalho = []
        for s in ws1.iter_rows(max_row= 14 , min_row=14):#pegando linha 14 para adicionar colunas

            for a in s:
                #inicio da tratativa dos nomes de cada coluna
                dt = str(a.value).strip()
                if (('PRODUTO' in dt) or ('Situação do produto' in dt) or ('Contratação' in dt) or ('Formação de Preços' in dt )):
                    continue
                else:
                    dt = dt.replace('dw bs','').replace('Variação','').replace('- Quantidade de vidas','').replace(' ','')
                dat = datetime.datetime.strptime(dt, '%b/%y')
                if dat in lista_cabecalho:
                    indice = len(lista_cabecalho)
                    coluna_ref1 = (ws1.cell(row = 1, column= indice + 1).column_letter)#POSIÇÃO DA COLUNA A SER TRATADA
                    break
                else:
                    lista_cabecalho.append(dat)

        coluna_ = indice

        contador_final2 = 0
        for coluna in ws1.iter_rows(min_row=15):#PEGANDO POSIÇÃO ONDE SERÁ INSERIDA NOVAS INSFORMAÇÕES
            colunaS = str(coluna[coluna_].value)
            if 'None' not in colunaS:
                contador_final2 = contador_final2 + 1
        posicao = contador_final2


        contador = tamanho + 15
        for item in novos:#ADICIONANDO NOVAS INFORMAÇÕES PRODUTO
            ws1[f'A{contador}'] = item
            contador += 1

        contador = posicao + 15
        for item in vidas:#ADICIONANDO NOVAS INFORMAÇÕES VIDAS
            ws1[f'{coluna_ref1}{contador}'] = item
            contador += 1


        wb.save(moniopmeditoramento + ultimo_arquivo)
    except Exception as e:
        logging.error('| Ocorreu um erro: | 3')
        logging.exception(str(e))
