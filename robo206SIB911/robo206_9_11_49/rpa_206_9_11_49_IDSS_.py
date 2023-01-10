import logging
import xml.etree.ElementTree as ET
import PyPDF2
from openpyxl import load_workbook


def extração_pdf_cpf_repetido():
    try:
        tree = ET.parse("C:\\Users\\User\\Desktop\\206-BS-571\\diretório\\diretorio.xml")
        root2 = tree.getroot()
        for child2 in root2:
                for x2 in root2.findall(child2.tag):
                        pdfidssop = x2.find('pdfidssop').text
                        apcia = x2.find('apcia').text
            
        diretorio = (pdfidssop)
        dir_2 = apcia

                
        wb = load_workbook(dir_2) #diretorio apuração 
        ws = wb.active
        


        #Abertura do arquivo PDF 
        arquivo_pdf = open(diretorio, 'rb')
        ler_arquivo = PyPDF2.PdfFileReader(arquivo_pdf)
        n_paginas = ler_arquivo.getNumPages()
        paginas = ler_arquivo.getPage(0)
        conteudo = paginas.extract_text()
        conteudo.split()
        conteudo.splitlines()

        l=[]
        l.append(conteudo)
        for i in range(0,len(l)):
            l[i] = l[i].split(None)
        IDSS = l[0][-2]
        # print(IDSS)
        IDSS_ = IDSS[-5:]
        # print(IDSS_)
        ws['J82'] = 'IDSS ' + IDSS_ + "%" #insere no campo solicitado 
        wb.save(dir_2)
    except Exception as e:
                logging.error('| Ocorreu um erro: | 3')
                logging.exception(str(e))



# import logging
# import xml.etree.ElementTree as ET
# import PyPDF2
# from openpyxl import load_workbook


# def extração_pdf_cpf_repetido():
#     try:
#         tree = ET.parse("C:\\Users\\User\\Desktop\\206-BS-571\\diretório\\diretorio.xml")
#         root2 = tree.getroot()
#         for child2 in root2:
#                 for x2 in root2.findall(child2.tag):
#                         pdfidssop = x2.find('pdfidssop').text
#                         apcia = x2.find('apcia').text
            
#         diretorio = (pdfidssop)
#         dir_2 = apcia

                
#         wb = load_workbook(dir_2) #diretorio apuração 
#         ws = wb.active

#         #Abertura do arquivo PDF 
#         arquivo_pdf = open(diretorio, 'rb')
#         ler_arquivo = PyPDF2.PdfFileReader(arquivo_pdf)
#         n_paginas = ler_arquivo.getNumPages()
#         paginas = ler_arquivo.getPage(0)
#         conteudo = paginas.extract_text()
#         conteudo.split()
#         conteudo.splitlines()

#         l=[]
#         l.append(conteudo)
#         for i in range(0,len(l)):
#             l[i] = l[i].split(None)
#         IDSS = l[0][-2]
#         # print(IDSS)
#         ws['J82'] = 'IDSS ' + IDSS + "%" #insere no campo solicitado 
#         wb.save(dir_2)
#     except Exception as e:
#                 logging.error('| Ocorreu um erro: | 3')
#                 logging.exception(str(e))    