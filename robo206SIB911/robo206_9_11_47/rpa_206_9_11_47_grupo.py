import logging
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
log_format = '%(asctime)s:%(levelname)s:%(filename)s:%(message)s'

def grupo():
    try:
        tree = ET.parse("C:\\Users\\User\\Desktop\\206-BS-571\\diretório\\diretorio.xml")
        root2 = tree.getroot()
        for child2 in root2:
                for x2 in root2.findall(child2.tag):
                        confeop = x2.find('confeop').text
                        apcia = x2.find('apcia').text
        dir_2 = apcia
        wb1 = load_workbook(dir_2)
        ws1 = wb1.active

        ##VALORES SOLICITADOS 
        B14 = ws1['B14'].value
        B15 = ws1['B15'].value
        C14 = ws1['C14'].value
        C15 = ws1['C15'].value
        D14 = ws1['D14'].value
        D15 = ws1['D15'].value
        E14 = ws1['E14'].value
        E15 = ws1['E15'].value
        F14 = ws1['F14'].value
        F15 = ws1['F15'].value
        I14 = ws1['I14'].value
        J14 = ws1['J14'].value
        K14 = ws1['K14'].value
        L14 = ws1['L14'].value
        M14 = ws1['M14'].value
        I15 = ws1['I15'].value
        J15 = ws1['J15'].value
        K15 = ws1['K15'].value
        L15 = ws1['L15'].value
        M15 = ws1['M15'].value
        
        valor1 = ws1['G80'].value
        valor2 = ws1['G81'].value
        valor3 = ws1['G82'].value
        valor4 = ws1['G83'].value
        valor5 = ws1['G84'].value
        valor6 = ws1['G85'].value
        valor7 = ws1['G86'].value
        valor8 = ws1['G87'].value
        valor9 = ws1['G88'].value
        valor10 = ws1['G89'].value
        # print(type(valor10))
        if ((valor1 == 0) or (valor5 == 0)):
            ws1['H80'] = '0'
        else:
            divisao1 = (valor1/valor5) #H80 

        if ((valor2 ==0) or (valor5 ==0)):
            ws1['H81'] = '0'
        else:
            divisao2 = (valor2/valor5) #h81

        if ((valor3 ==0) or (valor5 ==0)):
            ws1['H82'] = '0'
        else:    
            divisao3 = (valor3/valor5) #H82

        if ((valor4 ==0) or (valor5 ==0)):
            ws1['H83'] = '0'
        else:
            divisao4= (valor4/valor5) #H83

        if ((valor6 ==0) or (valor10 ==0)):
            ws1['H85'] = '0'
        else:
            divisao5 = (valor6/valor10 )#h85

        if ((valor7 ==0) or (valor10 ==0)):
            ws1['H86'] = '0'
        else:
            divisao6 = (valor7/valor10) #H86

        if ((valor8 ==0) or (valor10 ==0)):
            ws1['H87'] = '0'
        else:
            divisao7 = (valor8/valor10) #h87

        if ((valor9 ==0) or (valor10 ==0)):
            ws1['H80'] = '0'
        else:
            divisao8 = (valor9/valor10) #h88
            
        SomaB16 = (B14 + B15)
        SomaC16 = (C14 +  C15)
        SomaD16 =(D14 + D15)
        SomaE16 = (E14 + E15)
        SomaF16 = (F14 + F15)
        soma_TotalG14 = (B14 + C14 + D14 + E14 + F14)
        soma_TotalG15 = (B15 + C15 + D15 + E15 + F15)
        SomaN14 = (I14 + J14 + K14 + L14 + M14)
        SomaN15 = (I15 + J15 + K15 + L15 + M15)
        
        # SomaG16 = (soma_TotalG14 + soma_TotalG15)#TODO - PROXIMO CÓDIGO 
        
        ws1['N14']=SomaN14
        ws1['N15']=SomaN15
        ws1['B16'] = SomaB16
        ws1['C16'] = SomaC16
        ws1['D16'] = SomaD16
        ws1['E16'] = SomaE16
        ws1['F16'] = SomaF16 
        ws1['G14'] = soma_TotalG14
        ws1['G15'] = soma_TotalG15    
        ws1['H80'] = divisao1
        ws1['H81'] = divisao2
        ws1['H82'] = divisao3
        ws1['H83'] = divisao4
        ws1['H85'] = divisao5
        ws1['H86'] = divisao6
        ws1['H87'] = divisao7
        ws1['H88'] = divisao8

        wb1.save(dir_2)
    except Exception as e:
                logging.error('| Ocorreu um erro: | 3')
                logging.exception(str(e))



# import logging
# import xml.etree.ElementTree as ET
# from openpyxl import load_workbook
# log_format = '%(asctime)s:%(levelname)s:%(filename)s:%(message)s'

# def grupo():
#     try:
#         tree = ET.parse("C:\\Users\\User\\Desktop\\206-BS-571\\diretório\\diretorio.xml")
#         root2 = tree.getroot()
#         for child2 in root2:
#                 for x2 in root2.findall(child2.tag):
#                         confeop = x2.find('confeop').text
#                         apcia = x2.find('apcia').text
#         dir_2 = apcia
#         wb1 = load_workbook(dir_2)
#         ws1 = wb1.active

#         ##VALORES SOLICITADOS 
#         valor1 = ws1['G80'].value
#         valor2 = ws1['G81'].value
#         valor3 = ws1['G82'].value
#         valor4 = ws1['G83'].value
#         valor5 = ws1['G84'].value
#         valor6 = ws1['G85'].value
#         valor7 = ws1['G86'].value
#         valor8 = ws1['G87'].value
#         valor9 = ws1['G88'].value
#         valor10 = ws1['G89'].value
#         # print(type(valor10))
#         if ((valor1 == 0) or (valor5 == 0)):
#             ws1['H80'] = '0'
#         else:
#             divisao1 = (valor1/valor5) #H80 

#         if ((valor2 ==0) or (valor5 ==0)):
#             ws1['H81'] = '0'
#         else:
#             divisao2 = (valor2/valor5) #h81

#         if ((valor3 ==0) or (valor5 ==0)):
#             ws1['H82'] = '0'
#         else:    
#             divisao3 = (valor3/valor5) #H82

#         if ((valor4 ==0) or (valor5 ==0)):
#             ws1['H83'] = '0'
#         else:
#             divisao4= (valor4/valor5) #H83

#         if ((valor6 ==0) or (valor10 ==0)):
#             ws1['H85'] = '0'
#         else:
#             divisao5 = (valor6/valor10 )#h85

#         if ((valor7 ==0) or (valor10 ==0)):
#             ws1['H86'] = '0'
#         else:
#             divisao6 = (valor7/valor10) #H86

#         if ((valor8 ==0) or (valor10 ==0)):
#             ws1['H87'] = '0'
#         else:
#             divisao7 = (valor8/valor10) #h87

#         if ((valor9 ==0) or (valor10 ==0)):
#             ws1['H80'] = '0'
#         else:
#             divisao8 = (valor9/valor10) #h88

#             ws1['H80'] = divisao1
#             ws1['H81'] = divisao2
#             ws1['H82'] = divisao3
#             ws1['H83'] = divisao4
#             ws1['H85'] = divisao5
#             ws1['H86'] = divisao6
#             ws1['H87'] = divisao7
#             ws1['H88'] = divisao8

#         wb1.save(dir_2)
#     except Exception as e:
#                 logging.error('| Ocorreu um erro: | 3')
#                 logging.exception(str(e))    