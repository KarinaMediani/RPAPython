import logging
import xml.etree.ElementTree as ET
import PyPDF2
from openpyxl import load_workbook


def extração_pdf_cpf_repetido():
    try:
        tree = ET.parse("C:\\Users\\User\\Desktop\\206-BS-571\\diretório\\diretorio.xml")
        root2 = tree.getroot()
        for child2 in root2:
                for x2 in root2.findall(child2.tag):
                        pdfop = x2.find('pdfop').text
                        apcia = x2.find('apcia').text

        diretorio = pdfop
        dir_2 = apcia

                
        wb = load_workbook(dir_2) #diretorio apuração 
        ws = wb.active

        G14 = ws['G14'].value
        G15 = ws['G15'].value
        N14 = ws['N14'].value
        N15 = ws['N15'].value 
        SomaN16 = (N14 + N15)
        SomaG16 = (G14 + G15)

        #Abertura do arquivo PDF 
        arquivo_pdf = open(diretorio, 'rb')
        ler_arquivo = PyPDF2.PdfFileReader(arquivo_pdf)
        n_paginas = ler_arquivo.getNumPages()
        paginas = ler_arquivo.getPage(0)
        conteudo = paginas.extract_text()
        conteudo.split()
        cpf_repetido = conteudo[980:986] #captura no pdf a quanti de cpf repetido 
        # print(conteudo)
        conteudo.split()
        conteudo.splitlines()

        l=[]
        l.append(conteudo)
        for i in range(0,len(l)):
            l[i] = l[i].split(None)
        cpf_repetido = l[0][-34].split("Nº")
        repetidos = cpf_repetido[0]

        ws['N16'] = SomaN16
        ws['G16'] = SomaG16
        ws['J79'] = 'CPF REPETIDO ' + repetidos #insere no campo solicitado 
        wb.save(dir_2)
    except Exception as e:
                logging.error('| Ocorreu um erro: | 3')
                logging.exception(str(e))





# import logging
# import xml.etree.ElementTree as ET
# import PyPDF2
# from openpyxl import load_workbook


# def extração_pdf_cpf_repetido():
#     try:
#         tree = ET.parse("C:\\Users\\User\\Desktop\\206-BS-571\\diretório\\diretorio.xml")
#         root2 = tree.getroot()
#         for child2 in root2:
#                 for x2 in root2.findall(child2.tag):
#                         pdfop = x2.find('pdfop').text
#                         apcia = x2.find('apcia').text

#         diretorio = pdfop
#         dir_2 = apcia

                
#         wb = load_workbook(dir_2) #diretorio apuração 
#         ws = wb.active


#         #Abertura do arquivo PDF 
#         arquivo_pdf = open(diretorio, 'rb')
#         ler_arquivo = PyPDF2.PdfFileReader(arquivo_pdf)
#         n_paginas = ler_arquivo.getNumPages()
#         paginas = ler_arquivo.getPage(0)
#         conteudo = paginas.extract_text()
#         conteudo.split()
#         cpf_repetido = conteudo[980:986] #captura no pdf a quanti de cpf repetido 
#         # print(conteudo)
#         conteudo.split()
#         conteudo.splitlines()

#         l=[]
#         l.append(conteudo)
#         for i in range(0,len(l)):
#             l[i] = l[i].split(None)
#         cpf_repetido = l[0][-38]
#         # print(l)
#         # print(cpf_repetido)


#         ws['J79'] = 'CPF REPETIDO ' + cpf_repetido #insere no campo solicitado 
#         wb.save(dir_2)
#     except Exception as e:
#                 logging.error('| Ocorreu um erro: | 3')
#                 logging.exception(str(e))    
