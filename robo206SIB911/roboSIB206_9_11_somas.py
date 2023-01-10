import os
from openpyxl import load_workbook
import logging
import xml.etree.ElementTree as ET
import time
from validate_docbr import CPF

def confx():
    try:
        tree = ET.parse("C:\\Users\\User\\Desktop\\206-BS-571\\diretório\\diretorio.xml")
        root2 = tree.getroot()
        for child2 in root2:
                for x2 in root2.findall(child2.tag):
                        apcia = x2.find('apcia').text
                        confeop = x2.find('confeop').text
        dir_2 = apcia
        diretorio = (confeop)
        lista_arquivo = os.listdir(diretorio)
        contador_final1 = 0
        contador_final2 = 0
        contador_final3 = 0 
        contador_final4 = 0
        contador_final5 = 0
        contador_final6 = 0
        contador_final7 = 0
        contador_final8 = 0
        contador_final9 = 0
        contador_final10 = 0
        contador_final11 = 0
        contador_final12 = 0
        contador_final13 = 0
        contador_final14 = 0
        contador_final15 = 0
        contador_final16 = 0
        contador_final17 = 0
        contador_final18 = 0
        contador_final19 = 0
        contador_final20 = 0
        contador_final21 = 0
        contador_final22 = 0
        contador_final22 = 0
        contador_final23 = 0
        contador_final24 = 0
        contador_final25 = 0
        contador_final26 = 0
        contador_final27 = 0
        contador_final28 = 0
        contador_final29 = 0
        contador_final30 = 0
        contador_final31 = 0
        contador_final32 = 0
        contador_final33 = 0
        contador_final34 = 0
        contador_final35 = 0
        contador_final36 = 0
        contador_final37 = 0
        contador_final38 = 0
        contador_final39 = 0
        contador_final40 = 0
        contador_final41 = 0
        contador_final42 = 0
        contador_final43 = 0
        contador_final44 = 0
        contador_final45 = 0
        contador_final46 = 0
        contador_final47 = 0
        contador_final48 = 0
        contador_final49 = 0
        contador_final50 = 0
        contador_final51 = 0
        contador_final52 = 0
        contador_final53 = 0
        contador_final54 = 0
        contador_final55 = 0
        contador_final56 = 0
        contador_final57 = 0
        contador_final58 = 0
        contador_final59 = 0
        contador_final60 = 0
        contador_final61 = 0
        contador_final62 = 0
        contador_final63 = 0
        contador_final64 = 0
        contador_final65 = 0
        contador_final66 = 0
        contador_final67 = 0
        contador_final68 = 0
        contador_final69 = 0
        contador_final70 = 0
        contador_final71 = 0
        contador_final72 = 0
        contador_final73 = 0
        contador_final74 = 0
        contador_final75 = 0
        contador_final76 = 0
        contador_final77 = 0
        contador_final78 = 0
        contador_final79 = 0
        contador_final80 = 0
        contador_final81 = 0
        contador_final82 = 0
        contador_final83 = 0
        contador_final84 = 0
        contador_final85 = 0
        contador_final86 = 0
        contador_final87 = 0
        contador_final88 = 0
        contador_final89 = 0
        contador_final90 = 0
        contador_final91 = 0
        contador_final92 = 0
        contador_final93 = 0
        contador_final94 = 0
        contador_final95 = 0
        contador_final96 = 0
        contador_final97 = 0
        contador_final98 = 0
        contador_final99 = 0
        contador_final100 = 0
        contador1 = 0
        contador2 = 0
        contador3 = 0
        contador4 = 0
        contador5 = 0
        contador6 = 0
        contador7 = 0
        contador8 = 0
        contador9 = 0
        contador10 = 0
        contador11 = 0
        contador12 = 0
        contador13 = 0
        contador14 = 0
        contador15 = 0
        contador16 = 0
        contador17 = 0
        contador18 = 0
        contador19 = 0
        contador20 = 0
        contador21 = 0
        contador22 = 0
        contador23 = 0
        contador24 = 0
        contador25 = 0
        contador26 = 0
        contador27 = 0
        contador28 = 0
        contador29 = 0
        contador30 = 0
        contador31 = 0
        contador32 = 0
        contador33 = 0
        contador34 = 0
        contador35 = 0
        contador36 = 0
        contador37 = 0
        contador38 = 0
        contador39 = 0
        contador40 = 0
        contador41 = 0
        contador42 = 0
        contador43 = 0
        contador44 = 0
        contador45 = 0
        contador46 = 0
        contador47 = 0
        contador48 = 0
        contador49 = 0
        contador50 = 0
        contador51 = 0
        contador52 = 0
        contador53 = 0
        contador54 = 0
        contador55 = 0
        contador56 = 0
        contador57 = 0
        contador58 = 0
        contador59 = 0
        contador60 = 0
        contador61 = 0
        contador62 = 0
        contador63 = 0
        contador64 = 0
        contador65 = 0
        contador66 = 0
        contador67 = 0
        contador68 = 0
        contador69 = 0
        contador70 = 0
        contador71 = 0
        contador72 = 0
        contador73 = 0
        contador74 = 0
        contador75 = 0
        contador76 = 0
        contador77 = 0
        contador78 = 0
        contador79 = 0
        contador80 = 0
        contador81 = 0
        contador82 = 0
        contador83 = 0
        contador84 = 0
        contador85 = 0
        contador86 = 0
        contador87 = 0
        contador88 = 0
        contador89 = 0
        contador90 = 0
        contador91 = 0
        contador92 = 0
        contador93 = 0
        contador94 = 0
        contador95 = 0
        contador96 = 0
        contador97 = 0
        contador98 = 0
        contador99 = 0
        contador100 = 0
        final = 0
        final1  = 0
        final2 = 0
        final3 = 0
        final4 = 0
        final5 = 0
        final6 = 0
        final7 = 0
        final8 = 0
        final9 = 0
        final10 = 0
        final11 = 0
        final12 = 0
        final13 = 0
        final14 = 0
        final15 = 0
        final16 = 0
        final17 = 0
        final18 = 0
        final19 = 0
        final20 = 0
        final21 = 0
        final22 = 0
        final23 = 0
        final24 = 0
        final25 = 0
        final26 = 0
        final27 = 0
        final28 = 0
        final29 = 0
        final30 = 0
        final31 = 0
        final32 = 0
        final33 = 0
        final34 = 0
        final35 = 0
        final36 = 0
        final37 = 0
        final38 = 0
        final39 = 0
        final40 = 0
        final41 = 0
        final42 = 0
        final43 = 0
        final44 = 0
        final45 = 0
        final46 = 0
        final47 = 0
        final48 = 0
        final49 = 0
        final50 = 0
        final51 = 0
        final52 = 0
        final53 = 0
        final54 = 0
        final55 = 0
        final56 = 0
        final57 = 0
        final58 = 0
        final59 = 0
        final60 = 0
        final61 = 0
        final62 = 0
        final63 = 0
        final64 = 0
        final65 = 0
        final66 = 0
        final67 = 0
        final68 = 0
        final69 = 0
        final70 = 0
        final71 = 0
        final72 = 0
        final73 = 0
        final74 = 0
        final75 = 0
        final76 = 0
        final77 = 0
        final78 = 0
        final79 = 0
        final80 = 0
        final81 = 0
        final82 = 0
        final83 = 0
        final84 = 0
        final85 = 0
        final86 = 0
        final87 = 0
        final88 = 0
        final89 = 0
        final90 = 0
        final91 = 0
        final92 = 0
        final93 = 0
        final94 = 0
        final95 = 0
        final96 = 0
        final97 = 0
        final98 = 0
        final99 = 0
        final100 = 0
        final_1 = 0
        final_2 = 0
        final_3 = 0
        final_4 = 0
        final_5 = 0
        final_6 = 0
        final_7 = 0
        final_8 = 0
        final_9 = 0
        final_10 = 0
        final_11 = 0
        final_12 = 0
        final_13 = 0
        final_14 = 0
        contadorNovo = 0
        contadorNovo1 = 0
        contadorNovo2 = 0
        contadorNovo3 = 0
        contadorNovo4 = 0
        contadorNovo5 = 0
        contadorNovo6 = 0
        contadorNovo7 = 0
        contadorNovo8 = 0
        contadorNovo9 = 0
        contadorNovo10 = 0
        contadorNovo11 = 0
        contadorNovo12 = 0
        contadorNovo13 = 0
        contadorNovo14 = 0
        contadorNovo15 = 0
        contadorNovo16 = 0
        contadorNovo17 = 0
        contadorNovo18 = 0
        contadorNovo19 = 0
        contadorNovo20= 0
        contadorNovo21= 0
        contadorNovo22= 0
        contadorNovo23= 0
        contadorNovo24= 0
        contadorNovo25= 0
        contadorNovo26= 0
        contadorNovo27= 0
        contadorNovo28= 0
        contadorNovo29= 0
        contadorNovo30= 0
        contadorNovo31= 0
        contadorNovo32= 0
        contadorNovo33= 0
        contadorNovo34= 0
        contadorNovo35= 0
        contadorNovo36= 0
        contadorNovo37= 0
        contadorNovo38= 0
        contadorNovo39= 0
        contadorNovo40= 0
        contadorNovo41= 0
        contadorNovo42= 0
        contadorNovo43= 0
        contadorNovo44= 0
        contadorNovo45= 0
        contadorNovo46= 0
        contadorNovo47= 0
        contadorNovo48= 0
        contadorNovo49= 0
        contadorNovo50= 0
        contadorNovo51= 0
        contadorNovo52= 0
        contadorNovo53= 0
        contadorNovo54= 0
        contadorNovo55= 0
        contadorNovo56= 0
        contadorNovo57= 0
        contadorNovo58= 0
        contadorNovo59= 0
        contadorNovo60= 0
        contadorNovo61= 0
        contadorNovo62= 0
        contadorNovo63= 0
        contadorNovo64= 0
        contadorNovo65= 0
        contadorNovo66= 0
        contadorNovo67= 0
        contadorNovo68= 0
        contadorNovo69= 0
        contadorNovo70= 0
        contadorNovo71= 0
        contadorNovo72= 0
        contadorNovo73= 0
        contadorNovo74= 0
        contadorNovo75= 0
        contadorNovo76= 0
        contadorNovo77= 0
        contadorNovo78= 0
        contadorNovo79= 0
        contadorNovo80= 0
        contadorNovo81= 0
        contadorNovo82= 0
        contadorNovo83= 0
        contadorNovo84= 0
        contadorNovo85= 0
        contadorNovo86= 0
        contadorNovo87= 0
        contadorNovo88= 0
        contadorNovo89= 0
        contadorNovo90= 0
        contadorNovo91= 0
        contadorNovo92= 0
        contadorNovo93= 0
        contadorNovo94= 0
        contadorNovo95= 0
        contadorNovo96= 0
        contadorNovo97= 0
        contadorNovo98= 0
        contadorNovo99= 0
        contadorNovo100= 0
        contadorF0= 0
        contadorF1= 0
        contadorF2= 0
        contadorF3= 0
        contadorF4= 0
        contadorF5= 0
        contadorF6= 0
        contadorF7= 0
        contadorF8= 0
        contadorF9= 0
        contadorF10= 0
        contadorF11= 0
        contadorF12= 0
        contadorF13= 0
        contadorF14= 0
        contadorF15= 0
        contadorF16= 0
        contadorF17= 0
        contadorF18= 0
        contadorF19= 0
        contadorF20= 0
        contadorF21= 0
        contadorF22= 0
        contadorF23= 0
        contadorF24= 0
        contadorF25= 0
        contadorF26= 0
        contadorF27= 0
        contadorF28= 0
        contadorF29= 0
        contadorF30= 0
        contadorF31= 0
        contadorF32= 0
        contadorF33= 0
        contadorF34= 0
        contadorF35= 0
        contadorF36= 0
        contadorF37= 0
        contadorF38= 0
        contadorF39= 0
        contadorF40= 0
        contadorF41= 0
        contadorF42= 0
        contadorF43= 0
        contadorF44= 0
        contadorF45= 0
        contadorF46= 0
        contadorF47= 0
        contadorF48= 0
        contadorF49= 0
        contadorF50= 0
        contadorF51= 0
        contadorF52= 0
        contadorF53= 0
        contadorF54= 0
        contadorF55= 0
        contadorF56= 0
        contadorF57= 0
        contadorF58= 0
        contadorF59= 0
        contadorF60= 0
        contadorF61= 0
        contadorF62= 0
        contadorF63= 0
        contadorF64= 0
        contadorF65= 0
        contadorF66= 0
        contadorF67= 0
        contadorF68= 0
        contadorF69= 0
        contadorF70= 0
        contadorF71= 0
        contadorF72= 0
        contadorF73= 0
        contadorF74= 0
        contadorF75= 0
        contadorF76= 0
        contadorF77= 0
        contadorF78= 0
        contadorF79= 0
        contadorF80= 0
        contadorF81= 0
        contadorF82= 0
        contadorF83= 0
        contadorF84= 0
        contadorF85= 0
        contadorF86= 0
        contadorF87= 0
        contadorF88= 0
        contadorF89= 0
        contadorF90= 0
        contadorF91= 0
        contadorF92= 0
        contadorF93= 0
        contadorF94= 0
        contadorF95= 0
        contadorF96= 0
        contadorF97= 0
        contadorF98= 0
        contadorF99= 0
        contadorF100= 0

        cpf = CPF()
        for arquivo in lista_arquivo:
            if arquivo.startswith('ArqConf') and arquivo.endswith('.xlsx'):
                contador = 0
            
                wb = load_workbook(diretorio + '\\' + arquivo)
                ws = wb['ORIGEM']
                for linha in ws.iter_rows(min_row= 2):
                    contador += 1
            
                    valorJ = str(linha[9].value).upper().strip() #IDADE
                    valorAC = str(linha[28].value).upper().strip() #RAMO
                    valorAA = str(linha[26].value).upper().strip() #TITULARIDADE
                    valorAC = str(linha[28].value).upper().strip() #ramo
                    valorE = str(linha[4].value).upper().strip() #num_cpf
                    valorAS = str(linha[44].value).upper().strip()#dep maior n validado
                    valorAT = str(linha[45].value).upper().strip()#dep maior n encontrado
                    valorAU = str(linha[46].value).upper().strip()#dep maior n identif.
                    valorD = str(linha[3].value).upper().strip()#CPF
                    validacao = cpf.validate(valorD)
                    valorG = str(linha[6].value).upper().strip()#Abrev_nome
                    valorL = str(linha[11].value).upper().strip()##Abrev_mae
                    valorM = str(linha[12].value).upper().strip()#Nome_mae
                    valorI = str(linha[8].value).upper().strip()#Dt_nascim.
                    valorBC = str(linha[54].value).upper().strip() #titular não identificado
                    valorAX = str(linha[49].value).upper().strip() #cns nao preenchido
                    valorAW = str(linha[48].value).upper().strip() #sem cpf
                    valorAY = str(linha[50].value).upper().strip() #sem cpf
                    valorBB = str(linha[53].value).upper().strip() #sem cpf
                    valorAZ = str(linha[53].value).upper().strip() #sem cpf  
                    valorP = str(linha[15].value).upper().strip() #tipo cns
                    valorBD = str(linha[55].value).upper().strip() #sem cpf     
                        

                    ##DEPENDENTE MAIOR##
                    if ((valorJ == 'MAIOR')and (valorAC == '870')and(valorAA == 'DEPENDENTE')):
                        contador_final1 = contador_final1 + 1
            

                    if ((valorJ == 'MAIOR')and (valorAC == '874')and(valorAA == 'DEPENDENTE')):
                        contador_final2 = contador_final2 + 1
            
                    if ((valorJ == 'MAIOR')and (valorAC == '875')and(valorAA == 'DEPENDENTE')):
                        contador_final3 = contador_final3 + 1
        
                    if ((valorJ == 'MAIOR')and (valorAC == '876')and(valorAA == 'DEPENDENTE')):
                        contador_final4 = contador_final4 + 1
            
                    if ((valorJ == 'MAIOR')and (valorAC == '878')and(valorAA == 'DEPENDENTE')):
                        contador_final5 = contador_final5 + 1
                    ###DEPENDENTE MENOR###         
                    if ((valorJ == 'MENOR')and (valorAC == '870')and(valorAA == 'DEPENDENTE')):
                        contador_final6 = contador_final6 + 1
                
                    if ((valorJ == 'MENOR')and (valorAC == '874')and(valorAA == 'DEPENDENTE')):
                        contador_final7 = contador_final7 + 1
            
                    if ((valorJ == 'MENOR')and (valorAC == '875')and(valorAA == 'DEPENDENTE')):
                        contador_final8 = contador_final8 + 1
                        
                    if ((valorJ == 'MENOR')and (valorAC == '876')and(valorAA == 'DEPENDENTE')):
                        contador_final9 = contador_final9 + 1
        
                    if ((valorJ == 'MENOR')and (valorAC == '878')and(valorAA == 'DEPENDENTE')):
                        contador_final10 = contador_final10 + 1
                
                    if ((valorAC == '870') and (valorE == 'VAZIO')and (valorJ == 'MENOR')and(valorAA == 'DEPENDENTE') and (valorAT == 'VALIDO') 
                            and (valorAS == 'VALIDO') and (valorAU == 'VALIDO')):
                        contador_final11 = contador_final11 + 1
                    if ((valorAC == '874') and (valorE == 'VAZIO')and (valorJ == 'MENOR')and(valorAA == 'DEPENDENTE') and (valorAT == 'VALIDO') 
                            and (valorAS == 'VALIDO') and (valorAU == 'VALIDO')):
                        contador_final12 = contador_final12 + 1
            
                    if ((valorAC == '875') and (valorE == 'VAZIO')and (valorJ == 'MENOR')and(valorAA == 'DEPENDENTE') and (valorAT == 'VALIDO') 
                            and (valorAS == 'VALIDO') and (valorAU == 'VALIDO')):
                        contador_final13 = contador_final13 + 1
                    
                    if ((valorAC == '876') and (valorE == 'VAZIO')and (valorJ == 'MENOR')and(valorAA == 'DEPENDENTE') and (valorAT == 'VALIDO') 
                            and (valorAS == 'VALIDO') and (valorAU == 'VALIDO')):
                        contador_final14 = contador_final14 + 1
                    
                    if ((valorAC == '878') and (valorE == 'VAZIO')and (valorJ == 'MENOR')and(valorAA == 'DEPENDENTE') and (valorAT == 'VALIDO') 
                            and (valorAS == 'VALIDO') and (valorAU == 'VALIDO')):
                        contador_final15 = contador_final15 + 1
                    
                        ##CPF INVALIDO TIT MAIOR##
                    if ((valorAC == '870') and (validacao == False) and (valorJ =='MAIOR') and (valorAA == 'TITULAR')):
                        contador_final16 = contador_final16 + 1
                
                    if ((valorAC == '874') and (validacao == False) and (valorJ =='MAIOR') and (valorAA == 'TITULAR')):
                        contador_final17 = contador_final17 + 1
                
                    if ((valorAC == '875') and (validacao == False) and (valorJ =='MAIOR') and (valorAA == 'TITULAR')):
                        contador_final18 = contador_final18 + 1
                
                    if ((valorAC == '876') and (validacao == False) and (valorJ =='MAIOR') and (valorAA == 'TITULAR')):
                        contador_final19 = contador_final19 + 1
                            
                    if ((valorAC == '878') and (validacao == False) and (valorJ =='MAIOR')  and (valorAA == 'TITULAR')):
                        contador_final20 = contador_final20 + 1
                            
                    ##CPF EM BRANCO TIT MAIOR##
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (linha[3].value == None) and (valorAC == '870')):
                        contador_final21 = contador_final21 + 1 
                
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (linha[3].value == None) and (valorAC == '874')):
                        contador_final22 = contador_final22 + 1
                            
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (linha[3].value == None) and (valorAC == '875')):
                        contador_final23 = contador_final23 + 1
                
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (linha[3].value == None) and (valorAC == '876')):
                        contador_final24 = contador_final24 + 1
                
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (linha[3].value == None) and (valorAC == '878')):
                        contador_final25 = contador_final25 + 1
                
                    ##SEGURADO_ABRE_TIT MAIOR##
                
                    if ((valorAC == '870') and (valorJ == 'MAIOR') and(valorG == 'SIM') and (valorAA == 'TITULAR')):
                        contador_final26 = contador_final26 + 1
                
                    if ((valorAC == '874') and (valorJ == 'MAIOR') and(valorG == 'SIM') and (valorAA == 'TITULAR')):
                        contador_final27 = contador_final27 + 1
                            
                    if ((valorAC == '875') and (valorJ == 'MAIOR') and(valorG == 'SIM') and (valorAA == 'TITULAR')):
                        contador_final28 = contador_final28 + 1
                
                    if ((valorAC == '876') and (valorJ == 'MAIOR') and(valorG == 'SIM') and (valorAA == 'TITULAR')):
                        contador_final29 = contador_final29 + 1
                            
                    if ((valorAC == '878') and (valorJ == 'MAIOR') and(valorG == 'SIM') and (valorAA == 'TITULAR')):
                        contador_final30 = contador_final30 + 1
                
                    ##NOME MÃE ABREVIADO TITULAR MAIOR###
                    if ((valorAC == '870') and (valorJ == 'MAIOR') and(valorL == 'SIM') and (valorAA == 'TITULAR')):
                        contador_final31 = contador_final31 + 1
                    
                    if ((valorAC == '874') and (valorJ == 'MAIOR') and(valorL == 'SIM') and (valorAA == 'TITULAR')):
                        contador_final32 = contador_final32 + 1
                
                    if ((valorAC == '875') and (valorJ == 'MAIOR') and(valorL == 'SIM') and (valorAA == 'TITULAR')):
                        contador_final33 = contador_final33 + 1

                    if ((valorAC == '876') and (valorJ == 'MAIOR') and(valorL == 'SIM') and (valorAA == 'TITULAR')):
                        contador_final34 = contador_final34 + 1
                
                    if ((valorAC == '878') and (valorJ == 'MAIOR') and(valorL == 'SIM') and (valorAA == 'TITULAR')):
                        contador_final35 = contador_final35 + 1
                    
                    ##NOME DA MÃE EM BRANCO TITULAR MAIOR 
                
                    if ((valorAC == '870') and (valorJ == 'MAIOR') and(valorM == 'VAZIO') and (valorAA == 'TITULAR')):
                        contador_final36 = contador_final36 + 1
                    
                    if ((valorAC == '874') and (valorJ == 'MAIOR') and(valorM == 'VAZIO') and (valorAA == 'TITULAR')):
                        contador_final37 = contador_final37 + 1
                    
                    if ((valorAC == '875') and (valorJ == 'MAIOR') and(valorM == 'VAZIO') and (valorAA == 'TITULAR')):
                        contador_final38 = contador_final38 + 1
                
                    if ((valorAC == '876') and (valorJ == 'MAIOR') and(valorM == 'VAZIO') and (valorAA == 'TITULAR')):
                        contador_final39 = contador_final39 + 1
                            
                    if ((valorAC == '878') and (valorJ == 'MAIOR') and(valorM == 'VAZIO') and (valorAA == 'TITULAR')):
                        contador_final40 = contador_final40 + 1
                    ### NOME MAE INCORRETO TITULAR MAIOR###
                
                    if ((valorAC == '870') and (valorJ == 'MAIOR') and(valorM == 'INVALIDO') and (valorAA == 'TITULAR')):
                        contador_final41 = contador_final41 + 1
                            
                    if ((valorAC == '874') and (valorJ == 'MAIOR') and(valorM == 'INVALIDO') and (valorAA == 'TITULAR')):
                        contador_final42 = contador_final42 + 1
                
                    if ((valorAC == '876') and (valorJ == 'MAIOR') and(valorM == 'INVALIDO') and (valorAA == 'TITULAR')):
                        contador_final43 = contador_final43 + 1
                
                    if ((valorAC == '878') and (valorJ == 'MAIOR') and(valorM == 'INVALIDO') and (valorAA == 'TITULAR')):
                        contador_final44 = contador_final44 + 1
                            
                
                    ##DATA DE NASCIMENTO NAO DECLARADA TITULAR MAIOR###
                    if ((valorAC == '870') and (valorJ == 'MAIOR') and(linha[8].value == None) and (valorAA == 'TITULAR')):
                        contador_final45 = contador_final45 + 1
                
                    if ((valorAC == '874') and (valorJ == 'MAIOR') and(linha[8].value == None) and (valorAA == 'TITULAR')):
                        contador_final46 = contador_final46 + 1
                
                    if ((valorAC == '875') and (valorJ == 'MAIOR') and(linha[8].value == None) and (valorAA == 'TITULAR')):
                        contador_final47 = contador_final47 + 1
                
                    if ((valorAC == '876') and (valorJ == 'MAIOR') and(linha[8].value == None) and (valorAA == 'TITULAR')):
                        contador_final48 = contador_final48 + 1
                
                    if ((valorAC == '878') and (valorJ == 'MAIOR') and(linha[8].value == None) and (valorAA == 'TITULAR')):
                        contador_final49 = contador_final49 + 1
                
                    ## CPF EM BRANCO DEPENDENTE MENOR###
                    if ((valorJ == 'MAIOR') and(valorAA == 'DEPENDENTE') and (linha[3].value == None) and (valorAC == '870')):
                        contador_final50 = contador_final50 + 1
                
                    if ((valorJ == 'MAIOR') and(valorAA == 'DEPENDENTE') and (linha[3].value == None) and (valorAC == '874')):
                        contador_final51 = contador_final51 + 1
                
                    if ((valorJ == 'MAIOR') and(valorAA == 'DEPENDENTE') and (linha[3].value == None) and (valorAC == '875')):
                        contador_final52 = contador_final52 + 1
                
                    if ((valorJ == 'MAIOR') and(valorAA == 'DEPENDENTE') and (linha[3].value == None) and (valorAC == '876')):
                        contador_final53 = contador_final53 + 1
                
                    if ((valorJ == 'MAIOR') and(valorAA == 'DEPENDENTE') and (linha[3].value == None) and (valorAC == '878')):
                        contador_final54 = contador_final54 + 1
                
                    ##CPF INVALIDO DEPENDENTE MAIOR ##
                    if ((valorAC == '870') and (validacao == False) and (valorJ =='MAIOR') and (valorAA == 'DEPENDENTE')):
                        contador_final55 = contador_final55 + 1
                
                    if ((valorAC == '874') and (validacao == False) and (valorJ =='MAIOR') and (valorAA == 'DEPENDENTE')):
                        contador_final56 = contador_final56 + 1
                
                    if ((valorAC == '875') and (validacao == False) and (valorJ =='MAIOR') and (valorAA == 'DEPENDENTE')):
                        contador_final57 = contador_final57 + 1
                
                    if ((valorAC == '876') and (validacao == False) and (valorJ =='MAIOR') and (valorAA == 'DEPENDENTE')):
                        contador_final58 = contador_final58 + 1

                    if ((valorAC == '878') and (validacao == False) and (valorJ =='MAIOR') and (valorAA == 'DEPENDENTE')):
                        contador_final59 = contador_final59 + 1
                    ###DT NAO DECLARADA DEP MAIOR ###
                    if ((valorAC == '870') and (valorJ == 'MAIOR') and(linha[8].value == None) and (valorAA == 'DEPENDENTE')):
                        contador_final60 = contador_final60 + 1  
                
                    if ((valorAC == '874') and (valorJ == 'MAIOR') and(linha[8].value == None) and (valorAA == 'DEPENDENTE')):
                        contador_final61 = contador_final61 + 1
                
                    if ((valorAC == '875') and (valorJ == 'MAIOR') and(linha[8].value == None) and (valorAA == 'DEPENDENTE')):
                        contador_final62 = contador_final62 + 1
                
                    if ((valorAC == '876') and (valorJ == 'MAIOR') and(linha[8].value == None) and (valorAA == 'DEPENDENTE')):
                        contador_final63 = contador_final63 + 1
                
                    if ((valorAC == '878') and (valorJ == 'MAIOR') and(linha[8].value == None) and (valorAA == 'DEPENDENTE')):
                        contador_final64 = contador_final64 + 1
                    ##NOME DA MAE EM BRANCO DEPENDENTE MAIOR ###
                    if ((valorAC == '870') and (valorJ == 'MAIOR') and(valorM == 'VAZIO') and (valorAA == 'DEPENDENTE')):
                        contador_final65 = contador_final65 + 1
                    
                    if ((valorAC == '874') and (valorJ == 'MAIOR') and(valorM == 'VAZIO') and (valorAA == 'DEPENDENTE')):
                        contador_final66 = contador_final66 + 1
                        
                    if ((valorAC == '875') and (valorJ == 'MAIOR') and(valorM == 'VAZIO') and (valorAA == 'DEPENDENTE')):
                        contador_final67 = contador_final67 + 1
                    if ((valorAC == '876') and (valorJ == 'MAIOR') and(valorM == 'VAZIO') and (valorAA == 'DEPENDENTE')):
                        contador_final68 = contador_final68 + 1
                
                    if ((valorAC == '878') and (valorJ == 'MAIOR') and(valorM == 'VAZIO') and (valorAA == 'DEPENDENTE')):
                        contador_final69 = contador_final69 + 1
                
                    if ((valorAC == '870') and (valorJ == 'MAIOR') and(valorM == 'INVALIDO') and (valorAA == 'DEPENDENTE')):
                        contador_final70 = contador_final70 + 1
                
                    if ((valorAC == '874') and (valorJ == 'MAIOR') and(valorM == 'INVALIDO') and (valorAA == 'DEPENDENTE')):
                        contador_final71 = contador_final71 + 1
                
                    if ((valorAC == '875') and (valorJ == 'MAIOR') and(valorM == 'INVALIDO') and (valorAA == 'DEPENDENTE')):
                        contador_final72 = contador_final72 + 1
                
                    if ((valorAC == '876') and (valorJ == 'MAIOR') and(valorM == 'INVALIDO') and (valorAA == 'DEPENDENTE')):
                        contador_final73 = contador_final73 + 1
                
                    if ((valorAC == '878') and (valorJ == 'MAIOR') and(valorM == 'INVALIDO') and (valorAA == 'DEPENDENTE')):
                        contador_final74 = contador_final74 + 1
                    ##NOME MÃE ABREVIADO DEPENDENTE MAIOR###
                
                    if ((valorAC == '870') and (valorJ == 'MAIOR') and(valorL == 'SIM') and (valorAA == 'DEPENDENTE')):
                        contador_final75 = contador_final75 + 1
                
                    if ((valorAC == '874') and (valorJ == 'MAIOR') and(valorL == 'SIM') and (valorAA == 'DEPENDENTE')):
                        contador_final76 = contador_final76 + 1
                    
                    if ((valorAC == '875') and (valorJ == 'MAIOR') and(valorL == 'SIM') and (valorAA == 'DEPENDENTE')):
                        contador_final77 = contador_final77 + 1
                    
                    if ((valorAC == '876') and (valorJ == 'MAIOR') and(valorL == 'SIM') and (valorAA == 'DEPENDENTE')):
                        contador_final78 = contador_final78 + 1
                
                    if ((valorAC == '878') and (valorJ == 'MAIOR') and(valorL == 'SIM') and (valorAA == 'DEPENDENTE')):
                        contador_final79 = contador_final79 + 1
                
                    if ((valorAC == '870') and (valorJ == 'MAIOR') and(valorG == 'SIM') and (valorAA == 'DEPENDENTE')):
                        contador_final80 = contador_final80 + 1
            
                    if ((valorAC == '874') and (valorJ == 'MAIOR') and(valorG == 'SIM') and (valorAA == 'DEPENDENTE')):
                        contador_final81 = contador_final81 + 1
                
                    if ((valorAC == '875') and (valorJ == 'MAIOR') and(valorG == 'SIM') and (valorAA == 'DEPENDENTE')):
                        contador_final82 = contador_final82 + 1
                
                    if ((valorAC == '876') and (valorJ == 'MAIOR') and(valorG == 'SIM') and (valorAA == 'DEPENDENTE')):
                        contador_final83 = contador_final83 + 1
                
                    if ((valorAC == '878') and (valorJ == 'MAIOR') and(valorG == 'SIM') and (valorAA == 'DEPENDENTE')):
                        contador_final84 = contador_final84 + 1
                
                    ##CPF EM BRANCO TITULAR MENOR ###
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (linha[3].value == None) and (valorAC == '870')):
                        contador_final85 = contador_final85 + 1
                
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (linha[3].value == None) and (valorAC == '874')):
                        contador_final86 = contador_final86 + 1
                
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (linha[3].value == None) and (valorAC == '875')):
                        contador_final87 = contador_final87 + 1
                
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (linha[3].value == None) and (valorAC == '876')):
                        contador_final88 = contador_final88 + 1
                
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (linha[3].value == None) and (valorAC == '878')):
                        contador_final89 = contador_final89 + 1
                    ###CPF INVALIDO TITULAR MENOR ##

                    if ((valorAC == '870') and (validacao == False) and (valorJ =='MENOR') and (valorAA == 'TITULAR')):
                        contador_final90 = contador_final90 + 1
                
                    if ((valorAC == '874') and (validacao == False) and (valorJ =='MENOR') and (valorAA == 'TITULAR')):
                        contador_final91 = contador_final91 + 1

                    if ((valorAC == '875') and (validacao == False) and (valorJ =='MENOR') and (valorAA == 'TITULAR')):
                        contador_final92 = contador_final92 + 1
                
                    if ((valorAC == '876') and (validacao == False) and (valorJ =='MENOR') and (valorAA == 'TITULAR')):
                        contador_final93 = contador_final93 + 1
                
                    if ((valorAC == '878') and (validacao == False) and (valorJ =='MENOR') and (valorAA == 'TITULAR')):
                        contador_final94 = contador_final94 + 1
                    ##DT NAO DECLARADA TITULAR MENOR ##
                    if ((valorAC == '870') and (valorJ == 'MENOR') and(linha[8].value == None) and (valorAA == 'TITULAR ')):
                        contador_final95 = contador_final95 + 1
                
                    if ((valorAC == '874') and (valorJ == 'MENOR') and(linha[8].value == None) and (valorAA == 'TITULAR ')):
                        contador_final96 = contador_final96 + 1
                
                    if ((valorAC == '875') and (valorJ == 'MENOR') and(linha[8].value == None) and (valorAA == 'TITULAR ')):
                        contador_final97 = contador_final97 + 1
                
                    if ((valorAC == '876') and (valorJ == 'MENOR') and(linha[8].value == None) and (valorAA == 'TITULAR ')):
                        contador_final98 = contador_final98 + 1
                
                    if ((valorAC == '878') and (valorJ == 'MENOR') and(linha[8].value == None) and (valorAA == 'TITULAR ')):
                        contador_final99 = contador_final99 + 1
                    ###NOME MAE EM BRANCO TITULAR MENOR###
                    if ((valorAC == '870') and (valorJ == 'MENOR') and(valorM == 'VAZIO') and (valorAA == 'TITULAR')):
                        contador_final100 = contador_final100 + 1
                
                    if ((valorAC == '874') and (valorJ == 'MENOR') and(valorM == 'VAZIO') and (valorAA == 'TITULAR')):
                        contador1 = contador1 + 1
                
                    if ((valorAC == '875') and (valorJ == 'MENOR') and(valorM == 'VAZIO') and (valorAA == 'TITULAR')):
                        contador2 = contador2 + 1
                
                    if ((valorAC == '876') and (valorJ == 'MENOR') and(valorM == 'VAZIO') and (valorAA == 'TITULAR')):
                        contador3 = contador3 + 1
                
                    if ((valorAC == '878') and (valorJ == 'MENOR') and(valorM == 'VAZIO') and (valorAA == 'TITULAR')):
                        contador4 = contador4 + 1

                    ##NOME MAE INCORRETO TITULAR MENOR ##
                    if ((valorAC == '870') and (valorJ == 'MENOR') and(valorM == 'INVALIDO') and (valorAA == 'TITULAR')):
                        contador5 = contador5 + 1
                
                    if ((valorAC == '874') and (valorJ == 'MENOR') and(valorM == 'INVALIDO') and (valorAA == 'TITULAR')):
                        contador6 = contador6 + 1

                    if ((valorAC == '875') and (valorJ == 'MENOR') and(valorM == 'INVALIDO') and (valorAA == 'TITULAR')):
                        contador7 = contador7 + 1
                    if ((valorAC == '876') and (valorJ == 'MENOR') and(valorM == 'INVALIDO') and (valorAA == 'TITULAR')):
                        contador8 = contador8 + 1

                    if ((valorAC == '878') and (valorJ == 'MENOR') and(valorM == 'INVALIDO') and (valorAA == 'TITULAR')):
                        contador9 = contador9 + 1
                            
                    ###NOME MÃE ABREV TIT MENOR###
                    if ((valorAC == '870') and (valorJ == 'MENOR') and(valorL == 'SIM') and (valorAA == 'TITULAR')):
                        contador10 = contador10 + 1
                
                    if ((valorAC == '874') and (valorJ == 'MENOR') and(valorL == 'SIM') and (valorAA == 'TITULAR')):
                        contador11 = contador11 + 1
                
                    if ((valorAC == '875') and (valorJ == 'MENOR') and(valorL == 'SIM') and (valorAA == 'TITULAR')):
                        contador12 = contador12 + 1
                
                    if ((valorAC == '876') and (valorJ == 'MENOR') and(valorL == 'SIM') and (valorAA == 'TITULAR')):
                        contador13 = contador13 + 1
                
                    if ((valorAC == '878') and (valorJ == 'MENOR') and(valorL == 'SIM') and (valorAA == 'TITULAR')):
                        contador14 = contador14 + 1
                            
                
                    ##NOME ABREV TIT MENOR ##
                
                    if ((valorAC == '870') and (valorJ == 'MENOR') and(valorG == 'SIM') and (valorAA == 'TITULAR')):
                        contador15 = contador15 + 1
                
                    if ((valorAC == '874') and (valorJ == 'MENOR') and(valorG == 'SIM') and (valorAA == 'TITULAR')):
                        contador16 = contador16 + 1
                            
                    if ((valorAC == '875') and (valorJ == 'MENOR') and(valorG == 'SIM') and (valorAA == 'TITULAR')):
                        contador17 = contador17 + 1
                            
                    if ((valorAC == '876') and (valorJ == 'MENOR') and(valorG == 'SIM') and (valorAA == 'TITULAR')):
                        contador18 = contador18 + 1
                
                    if ((valorAC == '878') and (valorJ == 'MENOR') and(valorG == 'SIM') and (valorAA == 'TITULAR')):
                        contador19 = contador19 + 1
                
                    ###CPF EM BRANCO DEP MENOR###
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (linha[3].value == None) and (valorAC == '870')):
                        contador20 = contador20 + 1
                
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (linha[3].value == None) and (valorAC == '874')):
                        contador21 = contador21 + 1
                
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (linha[3].value == None) and (valorAC == '875')):
                        contador22 = contador22 + 1
                
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (linha[3].value == None) and (valorAC == '876')):
                        contador23 = contador23 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (linha[3].value == None) and (valorAC == '878')):
                        contador24 = contador24 + 1
                
                    ##CPF INVALIDO DEP MENOR 
                    if ((valorAC == '870') and (validacao == False) and (valorJ =='MENOR') and (valorAA == 'DEPENDENTE')):
                        contador25 = contador25 + 1
                
                    if ((valorAC == '874') and (validacao == False) and (valorJ =='MENOR') and (valorAA == 'DEPENDENTE')):
                        contador26 = contador26 + 1
                
                    if ((valorAC == '875') and (validacao == False) and (valorJ =='MENOR') and (valorAA == 'DEPENDENTE')):
                        contador27 = contador27 + 1
                
                    if ((valorAC == '876') and (validacao == False) and (valorJ =='MENOR') and (valorAA == 'DEPENDENTE')):
                        contador28 = contador28 + 1
                
                    if ((valorAC == '878') and (validacao == False) and (valorJ =='MENOR') and (valorAA == 'DEPENDENTE')):
                        contador29 = contador29 + 1

                    ## DT NAO DECLARADA DEP MENOR ##
                    if ((valorAC == '870') and (valorJ == 'MENOR') and(linha[8].value == None) and (valorAA == 'DEPENDENTE')):
                        contador30 = contador30 + 1
                
                    if ((valorAC == '874') and (valorJ == 'MENOR') and(linha[8].value == None) and (valorAA == 'DEPENDENTE')):
                        contador31 = contador31 + 1
                
                    if ((valorAC == '875') and (valorJ == 'MENOR') and(linha[8].value == None) and (valorAA == 'DEPENDENTE')):
                        contador31 = contador31 + 1
                
                    if ((valorAC == '876') and (valorJ == 'MENOR') and(linha[8].value == None) and (valorAA == 'DEPENDENTE')):
                        contador32 = contador32 + 1
                
                    if ((valorAC == '878') and (valorJ == 'MENOR') and(linha[8].value == None) and (valorAA == 'DEPENDENTE')):
                        contador33 = contador33 + 1
                
                    ##NOME MAE ABREV DEP MENOR ###
                    if ((valorAC == '870') and (valorJ == 'MENOR') and(valorL == 'SIM') and (valorAA == 'DEPENDENTE')):
                        contador34 = contador34 + 1
                            
                    if ((valorAC == '874') and (valorJ == 'MENOR') and(valorL == 'SIM') and (valorAA == 'DEPENDENTE')):
                        contador35 = contador35 + 1
                
                    if ((valorAC == '875') and (valorJ == 'MENOR') and(valorL == 'SIM') and (valorAA == 'DEPENDENTE')):
                        contador36 = contador36 + 1
                
                    if ((valorAC == '876') and (valorJ == 'MENOR') and(valorL == 'SIM') and (valorAA == 'DEPENDENTE')):
                        contador37 = contador37 + 1
                
                    if ((valorAC == '878') and (valorJ == 'MENOR') and(valorL == 'SIM') and (valorAA == 'DEPENDENTE')):
                        contador38 = contador38 + 1
                
                    ##NOME MAE BCO DEP MENOR ##
                
                    if ((valorAC == '870') and (valorJ == 'MENOR') and(valorM == 'VAZIO') and (valorAA == 'DEPENDENTE')):
                        contador39 = contador39 + 1
                
                    if ((valorAC == '874') and (valorJ == 'MENOR') and(valorM == 'VAZIO') and (valorAA == 'DEPENDENTE')):
                        contador40 = contador40 + 1
                
                    if ((valorAC == '875') and (valorJ == 'MENOR') and(valorM == 'VAZIO') and (valorAA == 'DEPENDENTE')):
                        contador41 = contador41 + 1
                
                    if ((valorAC == '875') and (valorJ == 'MENOR') and(valorM == 'VAZIO') and (valorAA == 'DEPENDENTE')):
                        contador42 = contador42 + 1
                    
                    if ((valorAC == '878') and (valorJ == 'MENOR') and(valorM == 'VAZIO') and (valorAA == 'DEPENDENTE')):
                        contador43 = contador43 + 1
                
                    ##NOME MAE INCORR. DEP MENOR ###
                    if ((valorAC == '870') and (valorJ == 'MENOR') and(valorM == 'INVALIDO') and (valorAA == 'DEPENDENTE')):
                        contador45 = contador45 + 1
                            
                    if ((valorAC == '874') and (valorJ == 'MENOR') and(valorM == 'INVALIDO') and (valorAA == 'DEPENDENTE')):
                        contador46 = contador46 + 1
                    
                    if ((valorAC == '875') and (valorJ == 'MENOR') and(valorM == 'INVALIDO') and (valorAA == 'DEPENDENTE')):
                        contador47 = contador47 + 1
                    
                    if ((valorAC == '876') and (valorJ == 'MENOR') and(valorM == 'INVALIDO') and (valorAA == 'DEPENDENTE')):
                        contador48 = contador48 + 1
                    
                    if ((valorAC == '878') and (valorJ == 'MENOR') and(valorM == 'INVALIDO') and (valorAA == 'DEPENDENTE')):
                        contador49 = contador49 + 1
                    
                    ###NOME SEGURADO ABREVIDO ##
                    if ((valorAC == '870') and (valorJ == 'MENOR') and(valorG == 'SIM') and (valorAA == 'DEPENDENTE')):
                        contador50 = contador50 + 1 
                    
                    if ((valorAC == '874') and (valorJ == 'MENOR') and(valorG == 'SIM') and (valorAA == 'DEPENDENTE')):
                        contador50 = contador50 + 1
                    
                    if ((valorAC == '875') and (valorJ == 'MENOR') and(valorG == 'SIM') and (valorAA == 'DEPENDENTE')):
                        contador51 = contador51 + 1
                    
                    if ((valorAC == '876') and (valorJ == 'MENOR') and(valorG == 'SIM') and (valorAA == 'DEPENDENTE')):
                        contador52 = contador52 + 1
                    
                    if ((valorAC == '878') and (valorJ == 'MENOR') and(valorG == 'SIM') and (valorAA == 'DEPENDENTE')):
                        contador53 = contador53 + 1
                        ###DEPENDENTE NAO VALIDADO MAIOR ###
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorAC == '870')):
                        contador54 = contador54 + 1
                    
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorAC == '874')):
                        contador55 = contador55 + 1
                    
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorAC == '875')):
                        contador56 = contador56 + 1
                    
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorAC == '876')):
                        contador57 = contador57 + 1
                    
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorAC == '878')):
                        contador58 = contador58 + 1
                    
                    ## DEPENDENTE NAO VALIDADO MENOR ###
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorAC == '870')):
                        contador59 = contador59 + 1
                    
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorAC == '874')):
                        contador60 = contador60 + 1
                    
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorAC == '875')):
                        contador61 = contador61 + 1
                    
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorAC == '876')):
                        contador62 = contador62 + 1
                    
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorAC == '878')):
                        contador63 = contador63 + 1
                    ###DEPENDENTE MAIOR NAO INDENTIFICADO ###
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorBC == '1') and (valorAC == '870')):
                        contador64 = contador64 + 1
                    
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorBC == '1') and (valorAC == '874')):
                        contador65 = contador65 + 1
                    
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorBC == '1') and (valorAC == '875')):
                        contador66 = contador66 + 1
                    
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorBC == '1') and (valorAC == '876')):
                        contador67 = contador67 + 1
                    
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorBC == '1') and (valorAC == '878')):
                        contador68 = contador68 + 1
                    
                    ###DEPENDENTE MENOR NÃO IDENTIFICADO ###
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorBC == '1') and (valorAC == '870')):
                        contador69 = contador69 + 1
                    
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorBC == '1') and (valorAC == '874')):
                        contador70 = contador70 + 1
                    
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorBC == '1') and (valorAC == '875')):
                        contador71 = contador71 + 1
                    
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorBC == '1') and (valorAC == '876')):
                        contador72 = contador72 + 1
                    
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorBC == '1') and (valorAC == '878')):
                        contador73 = contador73 + 1
                    ## TIT MAIOR NAO IDENTIFICADO ###
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorBC == '1') and (valorAC == '870')):
                        contador74 = contador74 + 1
                    
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorBC == '1') and (valorAC == '874')):
                        contador75 = contador75 + 1
                    
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorBC == '1') and (valorAC == '875')):
                        contador76 = contador76 + 1
                    
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorBC == '1') and (valorAC == '876')):
                        contador77 = contador77 + 1
                    
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorBC == '1') and (valorAC == '878')):
                        contador78 = contador78 + 1
                    ##TIT MENOR NAO IDENTIFICADO ###
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorBC == '1') and (valorAC == '870')):
                        contador79 = contador79 + 1
                    
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorBC == '1') and (valorAC == '874')):
                        contador80 = contador80 + 1
                    
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorBC == '1') and (valorAC == '875')):
                        contador81 = contador81 + 1
                    
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorBC == '1') and (valorAC == '876')):
                        contador82 = contador82 + 1
                    
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorBC == '1') and (valorAC == '878')):
                        contador83 = contador83 + 1
                    ###TITULAR NAO VALIDADO MAIOR ###
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorAC == '870')):
                        contador84 = contador84 + 1
                    
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorAC == '874')):
                        contador85 = contador85 + 1
                    
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorAC == '875')):
                        contador86 = contador86 + 1
                    
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorAC == '876')):
                        contador87 = contador87 + 1
                    
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorAC == '878')):
                        contador88 = contador88 + 1              
                    ###TITULAR NAO VALIDADO MENOR ###
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorAC == '870')):
                        contador89 = contador89 + 1
                    
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorAC == '874')):
                        contador90 = contador90 + 1
                    
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorAC == '875')):
                        contador91 = contador91 + 1
                    
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorAC == '876')):
                        contador92 = contador92 + 1
                    
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorAC == '878')):
                        contador93 = contador93 + 1
                    ### CNS N PREENCHIDA DEP MAIOR ###
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorAX == '1') and (valorAC == '870')):
                        contador94 = contador94 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorAX == '1') and (valorAC == '874')):
                        contador95 = contador95 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorAX == '1') and (valorAC == '875')):
                        contador96 = contador96 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorAX == '1') and (valorAC == '876')):
                        contador97 = contador97 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorAX == '1') and (valorAC == '878')):
                        contador98 = contador98 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorAX == '1') and (valorAC == '870')):
                        contador99 = contador99 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorAX == '1') and (valorAC == '874')):
                        contador100 = contador100 + 1
                    ##CNS N PREENCHIDA DEP MENOR 
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorAX == '1') and (valorAC == '875')):
                        final = final + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorAX == '1') and (valorAC == '876')):
                        final1 = final1 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorAX == '1') and (valorAC == '878')):
                        final2 = final2 + 1
                    ###CNS N PREENCHIDA TIT MAIOR ###
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorAX == '1') and (valorAC == '870')):
                        final3 = final3 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorAX == '1') and (valorAC == '874')):
                        final4 = final4 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorAX == '1') and (valorAC == '875')):
                        final5 = final5 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorAX == '1') and (valorAC == '876')):
                        final6 = final6 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorAX == '1') and (valorAC == '878')):
                        final7 = final7 + 1
                    ##CNS NAO PREENCHIDO TITULAR MENOR ###
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorAX == '1') and (valorAC == '870')):
                        final8 = final8 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorAX == '1') and (valorAC == '874')):
                        final9 = final9 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorAX == '1') and (valorAC == '875')):
                        final10 = final10 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorAX == '1') and (valorAC == '876')):
                        final11 = final11 + 1
                
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorAX == '1') and (valorAC == '878')):
                        final12 = final12 + 1
                            
                    ### CPF N PREENCHIDO DEP MAIOR ###
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorAW == '1') and (valorAC == '870')):
                        final13 = final13 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorAW == '1') and (valorAC == '874')):
                        final14 = final14 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorAW == '1') and (valorAC == '875')):
                        final15 = final15 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorAW == '1') and (valorAC == '876')):
                        final16 = final16 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorAW == '1') and (valorAC == '878')):
                        final17 = final17 + 1
                    ### CPF N PREENCHIDO DEP MENOR ###
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorAW == '1') and (valorAC == '870')):
                        final18 = final18 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorAW == '1') and (valorAC == '874')):
                        final19 = final19 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorAW == '1') and (valorAC == '875')):
                        final20 = final20 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorAW == '1') and (valorAC == '876')):
                                final21 = final21 + 1
            
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorAW == '1') and (valorAC == '878')):
                        final22 = final22 + 1
                    ### CPF N PREENCHIDO TIT MAIOR ###
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorAW == '1') and (valorAC == '870')):
                        final23 = final23 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorAW == '1') and (valorAC == '874')):
                        final24 = final24 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorAW == '1') and (valorAC == '875')):
                        final25 = final25 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorAW == '1') and (valorAC == '876')):
                        final26 = final26 + 1
                
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorAW == '1') and (valorAC == '878')):
                        final27 = final27 + 1
                    ## CPF NAO PREENCHIDO TIT MENOR ##
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorAW == '1') and (valorAC == '870')):
                        final28 = final28 + 1 
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorAW == '1') and (valorAC == '874')):
                        final29 = final29 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorAW == '1') and (valorAC == '875')):
                        final30 = final30 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorAW == '1') and (valorAC == '876')):
                        final31 = final31 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorAW == '1') and (valorAC == '878')):
                        final32 = final32 + 1
                
                    ##SEM CPF DEP MAIOR ##
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorAY == '1') and (valorAC == '870')):
                        final33 = final33 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorAY == '1') and (valorAC == '874')):
                        final34 = final34 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorAY == '1') and (valorAC == '875')):
                        final35 = final35 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorAY == '1') and (valorAC == '876')):
                        final36 = final36 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorAY == '1') and (valorAC == '878')):
                        final37 = final37 + 1
                    ## SEM CPF DEP MENOR ##
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorAY == '1') and (valorAC == '870')):
                        final38 = final38 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorAY == '1') and (valorAC == '874')):
                        final39 = final39 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorAY == '1') and (valorAC == '875')):
                        final40 = final40 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorAY == '1') and (valorAC == '876')):
                        final41 = final41 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorAY == '1') and (valorAC == '878')):
                        final42 = final42 + 1
                        ##SEM CPF TIT MAIOR ##
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorAY == '1') and (valorAC == '870')):
                        final43 = final43 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorAY == '1') and (valorAC == '874')):
                        final44 = final44 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorAY == '1') and (valorAC == '875')):
                        final45 = final45 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorAY == '1') and (valorAC == '876')):
                        final46 = final46 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorAY == '1') and (valorAC == '878')):
                        final47 = final47 + 1
                    ## sem cpf tit menor ##
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorAY == '1') and (valorAC == '870')):
                        final48 = final48 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorAY == '1') and (valorAC == '874')):
                        final49 = final49 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorAY == '1') and (valorAC == '875')):
                                final50 = final50 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorAY == '1') and (valorAC == '876')):
                        final51 = final51 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorAY == '1') and (valorAC == '878')):
                        final52 = final52 + 1
                    ## REG ENV REP CPF MAIOR 
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorBC == '1') and (valorAC == '870')):
                        final53 = final53 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorBC == '1') and (valorAC == '874')):
                        final54 = final54 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorBC == '1') and (valorAC == '875')):
                        final55 = final55 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorBC == '1') and (valorAC == '876')):
                        final56 = final56 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorBC == '1') and (valorAC == '878')):
                        final57 = final57 + 1
                    ## REG ENV REP CPF DEP MENOR ##
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorBC == '1') and (valorAC == '870')):
                        final58 = final58 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorBC == '1') and (valorAC == '874')):
                        final59 = final59 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorBC == '1') and (valorAC == '875')):
                        final60 = final60 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorBC == '1') and (valorAC == '876')):
                        final61 = final61 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorBC == '1') and (valorAC == '878')):
                        final62 = final62 + 1
                    ## REG ENV REP CPF TIT MAIOR ####
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorBC == '1') and (valorAC == '870')):
                        final63 = final63 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorBC == '1') and (valorAC == '874')):
                        final64 = final64 + 1
                
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorBC == '1') and (valorAC == '875')):
                        final65 = final65 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorBC == '1') and (valorAC == '876')):
                        final66 = final66 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorBC == '1') and (valorAC == '878')):
                        final67 = final67 + 1
                    ## REG ENV REP CPF TIT MENOR###
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorBC == '1') and (valorAC == '870')):
                        final68 = final68 + 1 
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorBC == '1') and (valorAC == '874')):
                        final69 = final69 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorBC == '1') and (valorAC == '875')):
                        final70 = final70 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorBC == '1') and (valorAC == '876')):
                        final71 = final71 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorBC == '1') and (valorAC == '878')):
                        final72 = final72 + 1
                    ## REG RESP CNS DEP MAIOR ###
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorBB == '1') and (valorAC == '870')):
                        final73 = final73 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorBB == '1') and (valorAC == '874')):
                        final74 = final74 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorBB == '1') and (valorAC == '875')):
                        final75 = final75 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorBB == '1') and (valorAC == '876')):
                        final76 = final76 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorBB == '1') and (valorAC == '878')):
                        final77 = final77 + 1            
                
                    ###REG RESP CNS DEP MENOR ###         
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorBB == '1') and (valorAC == '870')):
                        final78 = final78 + 1   
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorBB == '1') and (valorAC == '874')):
                        final79 = final79 + 1                 
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorBB == '1') and (valorAC == '875')):
                        final80 = final80 + 1  
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorBB == '1') and (valorAC == '876')):
                        final81 = final81 + 1  
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorBB == '1') and (valorAC == '878')):
                        final82 = final82 + 1            
                    ### REG RESP CND TIT MAIOR ###
                
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorBB == '1') and (valorAC == '870')):
                        final83 = final83 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorBB == '1') and (valorAC == '874')):
                        final84 = final84 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorBB == '1') and (valorAC == '875')):
                        final85 = final85 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorBB == '1') and (valorAC == '876')):
                        final86 = final86 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorBB == '1') and (valorAC == '878')):
                        final87 = final87 + 1
                    ### reg resp cnd tit menor ###
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorBB == '1') and (valorAC == '870')):
                        final88 = final88 + 1           
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorBB == '1') and (valorAC == '874')):
                        final89 = final89 + 1  
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorBB == '1') and (valorAC == '875')):
                        final90 = final90 + 1  
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorBB == '1') and (valorAC == '876')):
                        final91 = final91 + 1   
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorBB == '1') and (valorAC == '878')):
                        final92 = final92 + 1    
                
                    ### DT NASC DIV DEP MAIOR ### 
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorAZ == '1') and (valorAC == '870')):
                        final93 = final93 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorAZ == '1') and (valorAC == '874')):
                        final94 = final94 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorAZ == '1') and (valorAC == '875')):
                        final95 = final95 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorAZ == '1') and (valorAC == '876')):
                        final96 = final96 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorAZ == '1') and (valorAC == '878')):
                        final97 = final97 + 1
                    ### DT NASC DIV DEP MENOR ###
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorAZ == '1') and (valorAC == '870')):
                        final98 = final98 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorAZ == '1') and (valorAC == '874')):
                        final99 = final99 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorAZ == '1') and (valorAC == '874')):
                        final_1 = final_1 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorAZ == '1') and (valorAC == '875')):
                        final_2 = final_2 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorAZ == '1') and (valorAC == '876')):
                        final_3 = final_3 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorAZ == '1') and (valorAC == '878')):
                        final_4 = final_4 + 1
                    ### DT NASC DIV TIT MAIOR ###
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorAZ == '1') and (valorAC == '870')):
                        final_5 = final_5 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorAZ == '1') and (valorAC == '874')):
                        final_6 = final_6 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorAZ == '1') and (valorAC == '875')):
                        final_7 = final_7 + 1
                
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorAZ == '1') and (valorAC == '876')):
                        final_8 = final_8 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorAZ == '1') and (valorAC == '878')):
                        final_9 = final_9 + 1
                    ### DT NASC TIT MENOR ###
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorAZ == '1') and (valorAC == '870')):
                        final_10 = final_10 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorAZ == '1') and (valorAC == '874')):
                        final_11 = final_11 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorAZ == '1') and (valorAC == '875')):
                        final_12 = final_12 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorAZ == '1') and (valorAC == '876')):
                        final_13 = final_13 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorAZ == '1') and (valorAC == '878')):
                        final_14 = final_14 + 1
                                # SEM CPF SEM CNS MAIOR 9.11.45 E 46
                    if ( (valorAC == '870') and (valorJ == 'MAIOR')  and (valorE == 'VAZIO')
                            and (valorP== 'DEFINITIVO' or 'CNS TEMPORARIO')):
                        contadorNovo1 = contadorNovo1 + 1

                    if ( (valorAC == '874') and (valorJ == 'MAIOR')  and (valorE == 'VAZIO')
                            and (valorP== 'DEFINITIVO' or 'CNS TEMPORARIO')):
                        contadorNovo2 = contadorNovo2 + 1

                    if ( (valorAC == '875') and (valorJ == 'MAIOR')  and (valorE == 'VAZIO')
                            and (valorP== 'DEFINITIVO' or 'CNS TEMPORARIO')):
                        contadorNovo3 = contadorNovo3 + 1
                        
                    if ( (valorAC == '876') and (valorJ == 'MAIOR')  and (valorE == 'VAZIO')
                            and (valorP== 'DEFINITIVO' or 'CNS TEMPORARIO')):
                        contadorNovo4 = contadorNovo4 + 1
                    if ( (valorAC == '878') and (valorJ == 'MAIOR')  and (valorE == 'VAZIO')
                            and (valorP== 'DEFINITIVO' or 'CNS TEMPORARIO')):
                        contadorNovo5 = contadorNovo5 + 1
                    #SEM CPF SEM CNS MENOR 9.11.45 E 46
                    if ( (valorAC == '870') and (valorJ == 'MENOR')  and (valorE == 'VAZIO')
                            and (valorP== 'DEFINITIVO' or 'CNS TEMPORARIO')):
                        contadorNovo6 = contadorNovo6 + 1
                    if ( (valorAC == '874') and (valorJ == 'MENOR')  and (valorE == 'VAZIO')
                            and (valorP== 'DEFINITIVO' or 'CNS TEMPORARIO')):
                        contadorNovo7 = contadorNovo7 + 1
                    if ( (valorAC == '875') and (valorJ == 'MENOR')  and (valorE == 'VAZIO')
                            and (valorP== 'DEFINITIVO' or 'CNS TEMPORARIO')):
                        contadorNovo8 = contadorNovo8 + 1
                    if ( (valorAC == '876') and (valorJ == 'MENOR')  and (valorE == 'VAZIO')
                            and (valorP== 'DEFINITIVO' or 'CNS TEMPORARIO')):
                        contadorNovo9 = contadorNovo9 + 1
                    if ( (valorAC == '878') and (valorJ == 'MENOR')  and (valorE == 'VAZIO')
                            and (valorP== 'DEFINITIVO' or 'CNS TEMPORARIO')):
                        contadorNovo10 = contadorNovo10 + 1
                    #SEM CPF SEM CNS MAIOR 9.11.44                            
                    if ( (valorAC == '870') and (valorJ == 'MAIOR')  and (valorE == 'VAZIO')
                            and (valorP== 'SEM CNS')):
                        contadorNovo11 = contadorNovo11 + 1
                    if ( (valorAC == '874') and (valorJ == 'MAIOR')  and (valorE == 'VAZIO')
                            and (valorP== 'SEM CNS')):
                        contadorNovo12 = contadorNovo12 + 1
                    if ( (valorAC == '875') and (valorJ == 'MAIOR')  and (valorE == 'VAZIO')
                            and (valorP== 'SEM CNS')):
                        contadorNovo13 = contadorNovo13 + 1
                    if ( (valorAC == '876') and (valorJ == 'MAIOR')  and (valorE == 'VAZIO')
                            and (valorP== 'SEM CNS')):
                        contadorNovo14 = contadorNovo14 + 1
                    if ( (valorAC == '878') and (valorJ == 'MAIOR')  and (valorE == 'VAZIO')
                            and (valorP== 'SEM CNS')):
                        contadorNovo15 = contadorNovo15 + 1
                    #SEM CPF SEM CNS MENOR 9.11.44 
                    if ( (valorAC == '870') and (valorJ == 'MENOR')  and (valorE == 'VAZIO')
                            and (valorP== 'SEM CNS')):
                        contadorNovo16 = contadorNovo16 + 1
                    if ( (valorAC == '874') and (valorJ == 'MENOR')  and (valorE == 'VAZIO')
                            and (valorP== 'SEM CNS')):
                        contadorNovo17 = contadorNovo17 + 1
                    if ( (valorAC == '875') and (valorJ == 'MENOR')  and (valorE == 'VAZIO')
                            and (valorP== 'SEM CNS')):
                        contadorNovo18 = contadorNovo18 + 1
                    if ( (valorAC == '876') and (valorJ == 'MENOR')  and (valorE == 'VAZIO')
                            and (valorP== 'SEM CNS')):
                        contadorNovo19 = contadorNovo19 + 1
                    if ( (valorAC == '878') and (valorJ == 'MENOR')  and (valorE == 'VAZIO')
                            and (valorP== 'SEM CNS')):
                        contadorNovo20 = contadorNovo20 + 1
                    #COM CPF SCOM CNS MAIOR 9.11.43 
                    if ( (valorAC == '870') and (valorJ == 'MAIOR')  and (valorE == 'VALIDO')
                            and (valorP== 'CNS TEMPORARIO' or 'DEFINITIVO')):
                        contadorNovo21 = contadorNovo21 + 1
                    if ( (valorAC == '874') and (valorJ == 'MAIOR')  and (valorE == 'VALIDO')
                            and (valorP== 'CNS TEMPORARIO' or 'DEFINITIVO')):
                        contadorNovo22 = contadorNovo22 + 1
                    if ( (valorAC == '875') and (valorJ == 'MAIOR')  and (valorE == 'VALIDO')
                            and (valorP== 'CNS TEMPORARIO' or 'DEFINITIVO')):
                        contadorNovo23 = contadorNovo23 + 1
                    if ( (valorAC == '876') and (valorJ == 'MAIOR')  and (valorE == 'VALIDO')
                            and (valorP== 'CNS TEMPORARIO' or 'DEFINITIVO')):
                        contadorNovo24 = contadorNovo24 + 1
                    if ( (valorAC == '878') and (valorJ == 'MAIOR')  and (valorE == 'VALIDO')
                            and (valorP== 'CNS TEMPORARIO' or 'DEFINITIVO')):
                        contadorNovo25 = contadorNovo25 + 1
                    #COM CPF SCOM CNS MENOR 9.11.43      
                    if ( (valorAC == '870') and (valorJ == 'MENOR')  and (valorE == 'VALIDO')
                            and (valorP== 'CNS TEMPORARIO' or 'DEFINITIVO')):
                        contadorNovo26 = contadorNovo26 + 1
                    if ( (valorAC == '874') and (valorJ == 'MENOR')  and (valorE == 'VALIDO')
                            and (valorP== 'CNS TEMPORARIO' or 'DEFINITIVO')):
                        contadorNovo27 = contadorNovo27 + 1
                    if ( (valorAC == '875') and (valorJ == 'MENOR')  and (valorE == 'VALIDO')
                            and (valorP== 'CNS TEMPORARIO' or 'DEFINITIVO')):
                        contadorNovo28 = contadorNovo28 + 1
                    if ( (valorAC == '876') and (valorJ == 'MENOR')  and (valorE == 'VALIDO')
                            and (valorP== 'CNS TEMPORARIO' or 'DEFINITIVO')):
                        contadorNovo29 = contadorNovo29 + 1
                    if ( (valorAC == '878') and (valorJ == 'MENOR')  and (valorE == 'VALIDO')
                            and (valorP== 'CNS TEMPORARIO' or 'DEFINITIVO')):
                        contadorNovo30 = contadorNovo30 + 1
                    #COM CPF SEM CNS MAIOR 9.11.42               
                    if ( (valorAC == '870') and (valorJ == 'MAIOR')  and (valorE == 'VALIDO')
                            and (valorP== 'SEM CNS')):
                        contadorNovo31 = contadorNovo31 + 1
                    if ( (valorAC == '874') and (valorJ == 'MAIOR')  and (valorE == 'VALIDO')
                            and (valorP== 'SEM CNS')):
                        contadorNovo32 = contadorNovo32 + 1
                    if ( (valorAC == '875') and (valorJ == 'MAIOR')  and (valorE == 'VALIDO')
                            and (valorP== 'SEM CNS')):
                        contadorNovo33 = contadorNovo33 + 1
                    if ( (valorAC == '876') and (valorJ == 'MAIOR')  and (valorE == 'VALIDO')
                            and (valorP== 'SEM CNS')):
                        contadorNovo34 = contadorNovo34 + 1
                    if ( (valorAC == '878') and (valorJ == 'MAIOR')  and (valorE == 'VALIDO')
                            and (valorP== 'SEM CNS')):
                        contadorNovo35 = contadorNovo35 + 1
                    #COM CPF SEM CNS MENOR 9.11.42  
                    if ( (valorAC == '870') and (valorJ == 'MENOR')  and (valorE == 'VALIDO')
                            and (valorP== 'SEM CNS')):
                        contadorNovo36 = contadorNovo36 + 1
                    if ( (valorAC == '874') and (valorJ == 'MENOR')  and (valorE == 'VALIDO')
                            and (valorP== 'SEM CNS')):
                        contadorNovo37 = contadorNovo37 + 1
                    if ( (valorAC == '875') and (valorJ == 'MENOR')  and (valorE == 'VALIDO')
                            and (valorP== 'SEM CNS')):
                        contadorNovo38 = contadorNovo38 + 1
                    if ( (valorAC == '876') and (valorJ == 'MENOR')  and (valorE == 'VALIDO')
                            and (valorP== 'SEM CNS')):
                        contadorNovo39 = contadorNovo39 + 1
                    if ( (valorAC == '878') and (valorJ == 'MENOR')  and (valorE == 'VALIDO')
                            and (valorP== 'SEM CNS')):
                        contadorNovo40 = contadorNovo40 + 1
                    #DEP MAIOR 9.11.39 E 40  
                    if ( (valorAC == '870') and (valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorE == 'VALIDO')
                            and (valorG == 'NAO' or 'CONECTIVO')  and (valorL == 'NAO' or 'CONECTIVO') and (valorM == 'VALIDO')):
                        contadorNovo41 = contadorNovo41 + 1
                    if ( (valorAC == '874') and (valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorE == 'VALIDO')
                            and (valorG == 'NAO' or 'CONECTIVO')  and (valorL == 'NAO' or 'CONECTIVO') and (valorM == 'VALIDO')):
                        contadorNovo42 = contadorNovo42 + 1
                    if ( (valorAC == '875') and (valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorE == 'VALIDO')
                            and (valorG == 'NAO' or 'CONECTIVO')  and (valorL == 'NAO' or 'CONECTIVO') and (valorM == 'VALIDO')):
                        contadorNovo43 = contadorNovo43 + 1
                    if ( (valorAC == '876') and (valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorE == 'VALIDO')
                            and (valorG == 'NAO' or 'CONECTIVO')  and (valorL == 'NAO' or 'CONECTIVO') and (valorM == 'VALIDO')):
                        contadorNovo44 = contadorNovo44 + 1
                    if ( (valorAC == '878') and (valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorE == 'VALIDO')
                            and (valorG == 'NAO' or 'CONECTIVO')  and (valorL == 'NAO' or 'CONECTIVO') and (valorM == 'VALIDO')):
                        contadorNovo45 = contadorNovo45 + 1
                    ##DEP MENOR 9.11.39 E 40
                    if ( (valorAC == '870') and (valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorE == 'VALIDO')
                            and (valorG == 'NAO' or 'CONECTIVO')  and (valorL == 'NAO' or 'CONECTIVO') and (valorM == 'VALIDO')):
                        contadorNovo46 = contadorNovo46 + 1
                    if ( (valorAC == '874') and (valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorE == 'VALIDO')
                            and (valorG == 'NAO' or 'CONECTIVO')  and (valorL == 'NAO' or 'CONECTIVO') and (valorM == 'VALIDO')):
                        contadorNovo47 = contadorNovo47 + 1
                    if ( (valorAC == '875') and (valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorE == 'VALIDO')
                            and (valorG == 'NAO' or 'CONECTIVO')  and (valorL == 'NAO' or 'CONECTIVO') and (valorM == 'VALIDO')):
                        contadorNovo48 = contadorNovo48 + 1
                    if ( (valorAC == '876') and (valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorE == 'VALIDO')
                            and (valorG == 'NAO' or 'CONECTIVO')  and (valorL == 'NAO' or 'CONECTIVO') and (valorM == 'VALIDO')):
                        contadorNovo49 = contadorNovo49 + 1
                    if ( (valorAC == '878') and (valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorE == 'VALIDO')
                            and (valorG == 'NAO' or 'CONECTIVO')  and (valorL == 'NAO' or 'CONECTIVO') and (valorM == 'VALIDO')):
                        contadorNovo50 = contadorNovo50 + 1
                    ##TIT MAIOR 9.11.39 E 40
                    if ( (valorAC == '870') and (valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorE == 'VALIDO')
                            and (valorG == 'NAO' or 'CONECTIVO')  and (valorL == 'NAO' or 'CONECTIVO') and (valorM == 'VALIDO')):
                        contadorNovo51 = contadorNovo51 + 1
                    if ( (valorAC == '874') and (valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorE == 'VALIDO')
                            and (valorG == 'NAO' or 'CONECTIVO')  and (valorL == 'NAO' or 'CONECTIVO') and (valorM == 'VALIDO')):
                        contadorNovo52 = contadorNovo52 + 1
                    if ( (valorAC == '875') and (valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorE == 'VALIDO')
                            and (valorG == 'NAO' or 'CONECTIVO')  and (valorL == 'NAO' or 'CONECTIVO') and (valorM == 'VALIDO')):
                        contadorNovo53 = contadorNovo53 + 1
                    if ( (valorAC == '876') and (valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorE == 'VALIDO')
                            and (valorG == 'NAO' or 'CONECTIVO')  and (valorL == 'NAO' or 'CONECTIVO') and (valorM == 'VALIDO')):
                        contadorNovo54 = contadorNovo54 + 1
                    if ( (valorAC == '878') and (valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorE == 'VALIDO')
                            and (valorG == 'NAO' or 'CONECTIVO')  and (valorL == 'NAO' or 'CONECTIVO') and (valorM == 'VALIDO')):
                        contadorNovo55 = contadorNovo55 + 1
                        ###TIT MMENOR 9.11.39 E 40
                    if ( (valorAC == '870') and (valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorE == 'VALIDO')
                            and (valorG == 'NAO' or 'CONECTIVO')  and (valorL == 'NAO' or 'CONECTIVO') and (valorM == 'VALIDO')):
                        contadorNovo56 = contadorNovo56 + 1
                    if ( (valorAC == '870') and (valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorE == 'VALIDO')
                            and (valorG == 'NAO' or 'CONECTIVO')  and (valorL == 'NAO' or 'CONECTIVO') and (valorM == 'VALIDO')):
                        contadorNovo57 = contadorNovo57 + 1
                    if ( (valorAC == '875') and (valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorE == 'VALIDO')
                            and (valorG == 'NAO' or 'CONECTIVO')  and (valorL == 'NAO' or 'CONECTIVO') and (valorM == 'VALIDO')):
                        contadorNovo58 = contadorNovo58 + 1
                    if ( (valorAC == '876') and (valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorE == 'VALIDO')
                            and (valorG == 'NAO' or 'CONECTIVO')  and (valorL == 'NAO' or 'CONECTIVO') and (valorM == 'VALIDO')):
                        contadorNovo59 = contadorNovo59 + 1
                    if ( (valorAC == '878') and (valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorE == 'VALIDO')
                            and (valorG == 'NAO' or 'CONECTIVO')  and (valorL == 'NAO' or 'CONECTIVO') and (valorM == 'VALIDO')):
                        contadorNovo60 = contadorNovo60 + 1
                    #RAMO TOTAL DEP MAIOR 11.38
                    if ( (valorAC == '870') and (valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE')):
                        contadorNovo61 = contadorNovo61 + 1
                    if ( (valorAC == '874') and (valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE')):
                        contadorNovo62 = contadorNovo62 + 1
                    if ( (valorAC == '875') and (valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE')):
                        contadorNovo63 = contadorNovo63 + 1
                    if ( (valorAC == '876') and (valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE')):
                        contadorNovo64 = contadorNovo64 + 1
                    if ( (valorAC == '878') and (valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE')):
                        contadorNovo65 = contadorNovo65 + 1
                        #RAMO TOTAL DEP menor
                    if ( (valorAC == '870') and (valorJ == 'MENOR') and (valorAA == 'DEPENDENTE')):
                        contadorNovo66 = contadorNovo66 + 1
                    if ( (valorAC == '874') and (valorJ == 'MENOR') and (valorAA == 'DEPENDENTE')):
                        contadorNovo67 = contadorNovo67 + 1
                    if ( (valorAC == '875') and (valorJ == 'MENOR') and (valorAA == 'DEPENDENTE')):
                        contadorNovo68 = contadorNovo68 + 1
                    if ( (valorAC == '876') and (valorJ == 'MENOR') and (valorAA == 'DEPENDENTE')):
                        contadorNovo69 = contadorNovo69 + 1
                    if ( (valorAC == '878') and (valorJ == 'MENOR') and (valorAA == 'DEPENDENTE')):
                        contadorNovo70 = contadorNovo70 + 1
                        #RAMO TOTAL tit MAIOR
                    if ( (valorAC == '870') and (valorJ == 'MAIOR') and (valorAA == 'TITULAR')):
                        contadorNovo71 = contadorNovo71 + 1
                    if ( (valorAC == '874') and (valorJ == 'MAIOR') and (valorAA == 'TITULAR')):
                        contadorNovo72 = contadorNovo72 + 1
                    if ( (valorAC == '875') and (valorJ == 'MAIOR') and (valorAA == 'TITULAR')):
                        contadorNovo73 = contadorNovo73 + 1
                    if ( (valorAC == '876') and (valorJ == 'MAIOR') and (valorAA == 'TITULAR')):
                        contadorNovo74 = contadorNovo74 + 1
                    if ( (valorAC == '878') and (valorJ == 'MAIOR') and (valorAA == 'TITULAR')):
                        contadorNovo75 = contadorNovo75 + 1
                        #RAMO TOTAL tit menor
                    if ( (valorAC == '870') and (valorJ == 'MENOR') and (valorAA == 'TITULAR')):
                        contadorNovo76 = contadorNovo76 + 1
                    if ( (valorAC == '874') and (valorJ == 'MENOR') and (valorAA == 'TITULAR')):
                        contadorNovo77 = contadorNovo77 + 1
                    if ( (valorAC == '875') and (valorJ == 'MENOR') and (valorAA == 'TITULAR')):
                        contadorNovo78 = contadorNovo78 + 1
                    if ( (valorAC == '876') and (valorJ == 'MENOR') and (valorAA == 'TITULAR')):
                        contadorNovo79 = contadorNovo79 + 1
                    if ( (valorAC == '878') and (valorJ == 'MENOR') and (valorAA == 'TITULAR')):
                        contadorNovo80 = contadorNovo80 + 1
                                #9.11.37 cns definitivo def dep manior
                    if ( (valorAC == '870') and (valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorP == 'DEFINITIVO')):
                        contadorNovo81 = contadorNovo81 + 1
                    if ( (valorAC == '874') and (valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorP == 'DEFINITIVO')):
                        contadorNovo82 = contadorNovo82 + 1
                    if ( (valorAC == '875') and (valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorP == 'DEFINITIVO')):
                        contadorNovo83 = contadorNovo83 + 1
                    if ( (valorAC == '876') and (valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorP == 'DEFINITIVO')):
                        contadorNovo84 = contadorNovo84 + 1
                    if ( (valorAC == '878') and (valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorP == 'DEFINITIVO')):
                        contadorNovo85 = contadorNovo85 + 1
                            #9.11.37 cns def dep menor
                    if ( (valorAC == '870') and (valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorP == 'DEFINITIVO')):
                        contadorNovo86 = contadorNovo86 + 1
                    if ( (valorAC == '874') and (valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorP == 'DEFINITIVO')):
                        contadorNovo87 = contadorNovo87 + 1
                    if ( (valorAC == '875') and (valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorP == 'DEFINITIVO')):
                        contadorNovo88 = contadorNovo88 + 1
                    if ( (valorAC == '876') and (valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorP == 'DEFINITIVO')):
                        contadorNovo89 = contadorNovo89 + 1
                    if ( (valorAC == '878') and (valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorP == 'DEFINITIVO')):
                        contadorNovo90 = contadorNovo90 + 1
                            #9.11.37 cns def tit maior
                    if ( (valorAC == '870') and (valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorP == 'DEFINITIVO')):
                        contadorNovo91 = contadorNovo91 + 1
                    if ( (valorAC == '874') and (valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorP == 'DEFINITIVO')):
                        contadorNovo92 = contadorNovo92 + 1
                    if ( (valorAC == '875') and (valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorP == 'DEFINITIVO')):
                        contadorNovo93 = contadorNovo93 + 1
                    if ( (valorAC == '876') and (valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorP == 'DEFINITIVO')):
                        contadorNovo94 = contadorNovo94 + 1
                    if ( (valorAC == '878') and (valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorP == 'DEFINITIVO')):
                        contadorNovo95 = contadorNovo95 + 1
                        #9.11.37 cns def tit MENOR
                    if ( (valorAC == '870') and (valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorP == 'DEFINITIVO')):
                        contadorNovo96 = contadorNovo96 + 1
                    if ( (valorAC == '874') and (valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorP == 'DEFINITIVO')):
                        contadorNovo97 = contadorNovo97 + 1
                    if ( (valorAC == '875') and (valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorP == 'DEFINITIVO')):
                        contadorNovo98 = contadorNovo98 + 1
                    if ( (valorAC == '876') and (valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorP == 'DEFINITIVO')):
                        contadorNovo99 = contadorNovo99 + 1
                    if ( (valorAC == '878') and (valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorP == 'DEFINITIVO')):
                        contadorNovo100 = contadorNovo100 + 1
                                #9.11.37 CNS TEMPORARIO 
                                #CNS TEMP DEP MAIOR
                    if ( (valorAC == '870') and (valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorP == 'CNS TEMPORARIO')):
                        contadorF1 = contadorF1 + 1
                    if ( (valorAC == '874') and (valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorP == 'CNS TEMPORARIO')):
                        contadorF2 = contadorF2 + 1
                    if ( (valorAC == '875') and (valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorP == 'CNS TEMPORARIO')):
                        contadorF3 = contadorF3 + 1
                    if ( (valorAC == '876') and (valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorP == 'CNS TEMPORARIO')):
                        contadorF4 = contadorF4 + 1
                    if ( (valorAC == '878') and (valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorP == 'CNS TEMPORARIO')):
                        contadorF5 = contadorF5 + 1
                            #CNS TEMP DEP MENOR
                    if ( (valorAC == '870') and (valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorP == 'CNS TEMPORARIO')):
                        contadorF6 = contadorF6 + 1
                    if ( (valorAC == '874') and (valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorP == 'CNS TEMPORARIO')):
                        contadorF7 = contadorF7 + 1
                    if ( (valorAC == '875') and (valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorP == 'CNS TEMPORARIO')):
                        contadorF8 = contadorF8 + 1
                    if ( (valorAC == '878') and (valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorP == 'CNS TEMPORARIO')):
                        contadorF9 = contadorF9 + 1
                    if ( (valorAC == '878') and (valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorP == 'CNS TEMPORARIO')):
                        contadorF10 = contadorF10 + 1
                        #CNS TEMP tit maior
                    if ( (valorAC == '870') and (valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorP == 'CNS TEMPORARIO')):
                        contadorF11 = contadorF11 + 1
                    if ( (valorAC == '874') and (valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorP == 'CNS TEMPORARIO')):
                        contadorF12 = contadorF12 + 1
                    if ( (valorAC == '875') and (valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorP == 'CNS TEMPORARIO')):
                        contadorF13 = contadorF13 + 1
                    if ( (valorAC == '876') and (valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorP == 'CNS TEMPORARIO')):
                        contadorF14 = contadorF14 + 1
                    if ( (valorAC == '878') and (valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorP == 'CNS TEMPORARIO')):
                        contadorF15= contadorF15+ 1
                        #CNS TEMP tit MENOR
                    if ( (valorAC == '870') and (valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorP == 'CNS TEMPORARIO')):
                        contadorF16 = contadorF16 + 1
                    if ( (valorAC == '874') and (valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorP == 'CNS TEMPORARIO')):
                        contadorF17 = contadorF17 + 1
                    if ( (valorAC == '875') and (valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorP == 'CNS TEMPORARIO')):
                        contadorF18 = contadorF18 + 1
                    if ( (valorAC == '876') and (valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorP == 'CNS TEMPORARIO')):
                        contadorF19 = contadorF19 + 1
                    if ( (valorAC == '878') and (valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorP == 'CNS TEMPORARIO')):
                        contadorF20 = contadorF20 + 1
                            #SEM CNS
                            #DEP MAIOR
                    if ( (valorAC == '870') and (valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorP == 'SEM CNS')):
                        contadorF21 = contadorF21 + 1
                    if ( (valorAC == '874') and (valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorP == 'SEM CNS')):
                        contadorF22 = contadorF22 + 1
                    if ( (valorAC == '875') and (valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorP == 'SEM CNS')):
                        contadorF23 = contadorF23 + 1
                    if ( (valorAC == '875') and (valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorP == 'SEM CNS')):
                        contadorF24 = contadorF24 + 1
                    if ( (valorAC == '878') and (valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorP == 'SEM CNS')):
                        contadorF25 = contadorF25 + 1
                    #DEP MENOR
                    if ( (valorAC == '870') and (valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorP == 'SEM CNS')):
                        contadorF26 = contadorF26 + 1
                    if ( (valorAC == '874') and (valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorP == 'SEM CNS')):
                        contadorF27 = contadorF27 + 1
                    if ( (valorAC == '875') and (valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorP == 'SEM CNS')):
                        contadorF28 = contadorF28 + 1
                    if ( (valorAC == '876') and (valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorP == 'SEM CNS')):
                        contadorF29 = contadorF29 + 1
                    if ( (valorAC == '878') and (valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorP == 'SEM CNS')):
                        contadorF30 = contadorF30 + 1
                    #tit maior
                    if ( (valorAC == '870') and (valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorP == 'SEM CNS')):
                        contadorF31 = contadorF31 + 1
                    if ( (valorAC == '874') and (valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorP == 'SEM CNS')):
                        contadorF32 = contadorF32 + 1
                    if ( (valorAC == '875') and (valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorP == 'SEM CNS')):
                        contadorF33 = contadorF33 + 1
                    if ( (valorAC == '876') and (valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorP == 'SEM CNS')):
                        contadorF34 = contadorF34 + 1
                    if ( (valorAC == '878') and (valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorP == 'SEM CNS')):
                        contadorF35 = contadorF35 + 1
                        #tit MENOR
                    if ( (valorAC == '870') and (valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorP == 'SEM CNS')):
                        contadorF36 = contadorF36 + 1
                    if ( (valorAC == '874') and (valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorP == 'SEM CNS')):
                        contadorF37 = contadorF37 + 1
                    if ( (valorAC == '875') and (valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorP == 'SEM CNS')):
                        contadorF38 = contadorF38 + 1
                    if ( (valorAC == '876') and (valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorP == 'SEM CNS')):
                        contadorF39 = contadorF39 + 1
                    if ( (valorAC == '878') and (valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorP == 'SEM CNS')):
                        contadorF40 = contadorF40 + 1
                        #DEP N ENCONTRADO MAIOR 9.11.36
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorBD == '1') and (valorAC == '870')):
                        contadorF41 = contadorF41 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorBD == '1') and (valorAC == '874')):
                        contadorF42 = contadorF42 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorBD == '1') and (valorAC == '875')):
                        contadorF43 = contadorF43 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorBD == '1') and (valorAC == '876')):
                        contadorF44 = contadorF44 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'DEPENDENTE') and (valorBD == '1') and (valorAC == '878')):
                        contadorF45 = contadorF45 + 1
                        #DEP N ENCONTRADO MENOR 9.11.36
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorBD == '1') and (valorAC == '870')):
                        contadorF46 = contadorF46 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorBD == '1') and (valorAC == '874')):
                        contadorF47 = contadorF47 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorBD == '1') and (valorAC == '875')):
                        contadorF48 = contadorF48 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorBD == '1') and (valorAC == '876')):
                        contadorF49 = contadorF49 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'DEPENDENTE') and (valorBD == '1') and (valorAC == '878')):
                        contadorF50 = contadorF50 + 1
                        #TIT N ENCONTRADO MAIOR 9.11.36
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorBD == '1') and (valorAC == '870')):
                        contadorF51 = contadorF51 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorBD == '1') and (valorAC == '874')):
                        contadorF52 = contadorF52 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorBD == '1') and (valorAC == '875')):
                        contadorF53 = contadorF53 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorBD == '1') and (valorAC == '876')):
                        contadorF54 = contadorF54 + 1
                    if ((valorJ == 'MAIOR') and (valorAA == 'TITULAR') and (valorBD == '1') and (valorAC == '878')):
                        contadorF55 = contadorF55 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorBD == '1') and (valorAC == '870')):
                        contadorF56 = contadorF56 + 1
                        #TIT N ENCONTRADO MENOR 9.11.36
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorBD == '1') and (valorAC == '874')):
                        contadorF57 = contadorF57 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorBD == '1') and (valorAC == '875')):
                        contadorF58 = contadorF58 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorBD == '1') and (valorAC == '876')):
                        contadorF59 = contadorF59 + 1
                    if ((valorJ == 'MENOR') and (valorAA == 'TITULAR') and (valorBD == '1') and (valorAC == '878')):
                        contadorF60 = contadorF60 + 1



            resultado_linhasI10 = contador_final1 #I10
            resultado_linhasJ10 = contador_final2 #J10
            resultado_linhasK10 = contador_final3 #K10
            resultado_linhasL10 = contador_final4 #L10
            resultado_linhasM10 = contador_final5 #M10
            resultado_linhasI11 = contador_final6 #I11
            resultado_linhasJ11 = contador_final7 #J11
            resultado_linhasK11 = contador_final8 #K11
            resultado_linhasL11 = contador_final9 #L11
            resultado_linhasM11 = contador_final10 #M11
            resultado_linhasI19 = contador_final11 #I19
            resultado_linhasJ19 = contador_final12 #J19
            resultado_linhasK19 = contador_final13 #K19
            resultado_linhasL19 = contador_final14 #L19
            resultado_linhasM19 = contador_final15 #M19
            resultado_linhasB22 = contador_final16 #B22
            resultado_linhasC22 = contador_final17 #C22
            resultado_linhasD22 = contador_final18 #D22
            resultado_linhasE22 = contador_final19 #E22
            resultado_linhasF22 = contador_final20 #F22
            resultado_linhasB23 = contador_final21 #B23
            resultado_linhasC23 = contador_final22 #C23
            resultado_linhasD23 = contador_final23 #D23
            resultado_linhasE23 = contador_final24 #E23
            resultado_linhasF23 = contador_final25 #F23
            resultado_linhasB24 = contador_final26 #B24
            resultado_linhasC24 = contador_final27 #C24
            resultado_linhasD24 = contador_final28 #D24
            resultado_linhasE24 = contador_final29 #E24
            resultado_linhasF24 = contador_final30 #F24
            resultado_linhasB25 = contador_final31 #B25
            resultado_linhasC25 = contador_final32 #C25
            resultado_linhasD25 = contador_final33 #D25
            resultado_linhasE25 = contador_final34 #E25
            resultado_linhasF25 = contador_final35 #F25
            resultado_linhasB26 = contador_final36 #B26
            resultado_linhasC26 = contador_final37 #C26
            resultado_linhasD26 = contador_final38 #D26
            resultado_linhasE26 = contador_final39 #E26
            resultado_linhasF26 = contador_final40 #F26
            resultado_linhasB27 = contador_final41 #B27
            resultado_linhasC27 = contador_final42 #C27
            resultado_linhasD27 = contador_final43 #D27
            resultado_linhasE27 = contador_final43 #E27
            resultado_linhasF27 = contador_final44 #F27
            resultado_linhasB28 = contador_final45 #B28
            resultado_linhasC28 = contador_final46 #C28
            resultado_linhasD28 = contador_final47 #D28
            resultado_linhasE28 = contador_final48 #E28
            resultado_linhasF28 = contador_final49 #F28
            resultado_linhasI23 = contador_final50 #I23
            resultado_linhasJ23 = contador_final51 #J23
            resultado_linhasK23 = contador_final52 #K23
            resultado_linhasL23 = contador_final53 #L23
            resultado_linhasM23 = contador_final54 #M23
            resultado_linhasI22 = contador_final55 #I22
            resultado_linhasJ22 = contador_final56 #J22
            resultado_linhasK22 = contador_final57 #K22
            resultado_linhasL22 = contador_final58 #L22
            resultado_linhasM22 = contador_final59 #M22
            resultado_linhasI28 = contador_final60 #I28
            resultado_linhasJ28 = contador_final61 #J28
            resultado_linhasK28 = contador_final62 #K28
            resultado_linhasL28 = contador_final63 #l28
            resultado_linhasM28 = contador_final64 #M28
            resultado_linhasI26 = contador_final65 #I26
            resultado_linhasJ26 = contador_final66 #J26
            resultado_linhasK26 = contador_final67 #K26
            resultado_linhasL26 = contador_final68 #L26
            resultado_linhasM26 = contador_final69 #M26
            resultado_linhasI27 = contador_final70 #I27
            resultado_linhasJ27 = contador_final71 #J27
            resultado_linhasK27 = contador_final72 #K27
            resultado_linhasL27 = contador_final73 #L27
            resultado_linhasM27 = contador_final74 #M27
            resultado_linhasI25 = contador_final75 #I25
            resultado_linhasJ25 = contador_final76 #J25
            resultado_linhasK25 = contador_final77 #K25
            resultado_linhasL25 = contador_final78 #L25
            resultado_linhasM25 = contador_final79 #M25
            resultado_linhasI24 = contador_final80 #I24
            resultado_linhasJ24 = contador_final81 #J24
            resultado_linhasK24 = contador_final82 #K24
            resultado_linhasL24 = contador_final83 #L24
            resultado_linhasM24 = contador_final84 #M24
            resultado_linhasB32 = contador_final85 #B32
            resultado_linhasC32 = contador_final86 #C32
            resultado_linhasD32 = contador_final87 #D32
            resultado_linhasE32 = contador_final88 #E32
            resultado_linhasF32 = contador_final89 #F32
            resultado_linhasB31 = contador_final90 #B31
            resultado_linhasC31 = contador_final91 #C31
            resultado_linhasD31 = contador_final92 #D31
            resultado_linhasE31 = contador_final93 #E31
            resultado_linhasF31 = contador_final94 #F31
            resultado_linhasB37 = contador_final95 #B37
            resultado_linhasC37 = contador_final96 #C37
            resultado_linhasD37 = contador_final97 #D37
            resultado_linhasE37 = contador_final98 #E37
            resultado_linhasF37 = contador_final99 #F37
            resultado_linhasB35 = contador_final100 #B35
            resultado_linhasC35 = contador1 #C35
            resultado_linhasD35 = contador2 #D35
            resultado_linhasE35 = contador3 #E35
            resultado_linhasF35 = contador4 #F35
            resultado_linhasB36 = contador5 #B36
            resultado_linhasC36 = contador6 #C36
            resultado_linhasD36 = contador7 #D36
            resultado_linhasE36 = contador8 #E36
            resultado_linhasF36 = contador9 #F36
            resultado_linhasB34 = contador10 #B34
            resultado_linhasC34 = contador11 #C34
            resultado_linhasD34 = contador12 #D34
            resultado_linhasE34 = contador13 #e34
            resultado_linhasF34 = contador14 #F34
            resultado_linhasB33 = contador15
            resultado_linhasC33 = contador16
            resultado_linhasD33 = contador17
            resultado_linhasE33 = contador18
            resultado_linhasF33 = contador19
            resultado_linhasI32 = contador20
            resultado_linhasJ32 = contador21
            resultado_linhasK32 = contador22
            resultado_linhasL32 = contador23
            resultado_linhasM32 = contador24
            resultado_linhasI31 = contador25
            resultado_linhasJ31 = contador26
            resultado_linhasK31 = contador27
            resultado_linhasL31 = contador28
            resultado_linhasM31 = contador29
            resultado_linhasI37 = contador30
            resultado_linhasJ37 = contador31
            resultado_linhasK37 = contador31
            resultado_linhasL37 = contador32
            resultado_linhasM37 = contador33
            resultado_linhasI34 = contador34
            resultado_linhasJ34 = contador35
            resultado_linhasK34 = contador36
            resultado_linhasL34 = contador37
            resultado_linhasM34 = contador38
            resultado_linhasI35 = contador39
            resultado_linhasJ35 = contador40
            resultado_linhasK35 = contador41
            resultado_linhasL35 = contador42
            resultado_linhasM35 = contador43
            resultado_linhasI36 = contador45
            resultado_linhasJ36 = contador46
            resultado_linhasK36 = contador47
            resultado_linhasL36 = contador48
            resultado_linhasM36 = contador49
            resultado_linhasI33 = contador50
            resultado_linhasJ33 = contador50
            resultado_linhasK33 = contador51
            resultado_linhasL33 = contador52
            resultado_linhasM33 = contador53
            resultado_linhasI40 = contador54
            resultado_linhasJ40 = contador55
            resultado_linhasK40 = contador56
            resultado_linhasL40 = contador57
            resultado_linhasM40 = contador58
            resultado_linhasI51 = contador59
            resultado_linhasJ51 = contador60
            resultado_linhasK51 = contador61
            resultado_linhasL51 = contador62
            resultado_linhasM51 = contador63
            resultado_linhasI41 = contador64
            resultado_linhasJ41 = contador65
            resultado_linhasK41 = contador66
            resultado_linhasL41 = contador67
            resultado_linhasM41 = contador68
            resultado_linhasI52 = contador69
            resultado_linhasJ52 = contador70
            resultado_linhasK52 = contador71
            resultado_linhasL52 = contador72
            resultado_linhasM52 = contador73
            resultado_linhasB41 = contador74
            resultado_linhasC41 = contador75
            resultado_linhasD41 = contador76
            resultado_linhasE41 = contador77
            resultado_linhasF41 = contador78
            resultado_linhasB52 = contador79
            resultado_linhasC52 = contador80
            resultado_linhasD52 = contador81
            resultado_linhasE52 = contador82             
            resultado_linhasF52 = contador83
            resultado_linhasB40 = contador84
            resultado_linhasC40 = contador85
            resultado_linhasD40 = contador86
            resultado_linhasE40 = contador87
            resultado_linhasF40 = contador88
            resultado_linhasB51 = contador89
            resultado_linhasC51 = contador90
            resultado_linhasD51 = contador91
            resultado_linhasE51 = contador92
            resultado_linhasF51 = contador93
            resultado_linhasI42 = contador94
            resultado_linhasJ42 = contador95
            resultado_linhasK42 = contador96
            resultado_linhasL42 = contador97
            resultado_linhasM42 = contador98
            resultado_linhasI53 = contador99
            resultado_linhasJ53 = contador100
            resultado_linhasK53 = final
            resultado_linhasL53 = final1
            resultado_linhasM53 = final2
            resultado_linhasB42 = final3
            resultado_linhasC42 = final4
            resultado_linhasD42 = final5
            resultado_linhasE42 = final6
            resultado_linhasF42 = final7
            resultado_linhasB53 = final8
            resultado_linhasC53 = final9
            resultado_linhasD53 = final10
            resultado_linhasE53 = final11
            resultado_linhasF53 = final12
            resultado_linhasI43 = final13
            resultado_linhasJ43 = final14
            resultado_linhasK43 = final15
            resultado_linhasL43 = final16
            resultado_linhasM43 = final17
            resultado_linhasI54 = final18
            resultado_linhasJ54 = final19
            resultado_linhasK54 = final20
            resultado_linhasL54 = final21
            resultado_linhasM54 = final22
            resultado_linhasB43 = final23
            resultado_linhasC43 = final24  
            resultado_linhasD43 = final25
            resultado_linhasE43 = final26
            resultado_linhasF43 = final27
            resultado_linhasB54 = final28
            resultado_linhasC54 = final29
            resultado_linhasD54 = final30
            resultado_linhasE54 = final31
            resultado_linhasF54 = final32       
            resultado_linhasI44 = final33
            resultado_linhasJ44 = final34
            resultado_linhasK44 = final35
            resultado_linhasL44 = final36 
            resultado_linhasM44 = final37
            resultado_linhasI55 = final38
            resultado_linhasJ55 = final39
            resultado_linhasK55 = final40
            resultado_linhasL55 = final41
            resultado_linhasM55 = final42
            resultado_linhasB44 = final43
            resultado_linhasC44 = final44
            resultado_linhasD44 = final45
            resultado_linhasE44 = final46
            resultado_linhasF44 = final47
            resultado_linhasB55 = final48
            resultado_linhasC55 = final49
            resultado_linhasD55 = final50
            resultado_linhasE55 = final51
            resultado_linhasF55 = final52      
            resultado_linhasI45 = final53   
            resultado_linhasJ45 = final54     
            resultado_linhasK45 = final55      
            resultado_linhasL45 = final56   
            resultado_linhasM45 = final57   
            resultado_linhasI56 = final58
            resultado_linhasJ56 = final59
            resultado_linhasK56 = final60
            resultado_linhasL56 = final61  
            resultado_linhasM56 = final62
            resultado_linhasB45 = final63
            resultado_linhasC45 = final64
            resultado_linhasD45 = final65
            resultado_linhasE45 = final66
            resultado_linhasF45 = final67
            resultado_linhasB56 = final68
            resultado_linhasC56 = final69
            resultado_linhasD56 = final70
            resultado_linhasE56 = final71
            resultado_linhasF56 = final72
            resultado_linhasI46 = final73  
            resultado_linhasJ46 = final74
            resultado_linhasK46 = final75
            resultado_linhasL46 = final76
            resultado_linhasM46 = final77  
            resultado_linhasI57 = final78   
            resultado_linhasJ57 = final79
            resultado_linhasK57 = final80      
            resultado_linhasL57 = final81
            resultado_linhasM57 = final82
            resultado_linhasB46 = final83
            resultado_linhasC46 = final84
            resultado_linhasD46 = final85 
            resultado_linhasE46 = final86
            resultado_linhasF46 = final87
            resultado_linhasB57 = final88
            resultado_linhasC57 = final89    
            resultado_linhasD57 = final90  
            resultado_linhasE57 = final91
            resultado_linhasF57 = final92   
            resultado_linhasI47 = final93
            resultado_linhasJ47 = final94   
            resultado_linhasK47 = final95
            resultado_linhasL47 = final96     
            resultado_linhasM47 = final97    
            resultado_linhasI58 = final98
            resultado_linhasJ58 = final99   
            resultado_linhasJ58 = final_1
            resultado_linhasK58 = final_2 
            resultado_linhasL58 = final_3
            resultado_linhasM58 = final_4
            resultado_linhasB47 = final_5
            resultado_linhasC47 = final_6
            resultado_linhasD47 = final_7
            resultado_linhasE47 = final_8
            resultado_linhasF47 = final_9
            resultado_linhasB58 = final_10   
            resultado_linhasC58 = final_11   
            resultado_linhasD58 = final_12
            resultado_linhasE58 = final_13      
            resultado_linhasF58 = final_14  
            resultado_linhasB83 = contadorNovo1
            resultado_linhasC83 = contadorNovo2
            resultado_linhasD83 = contadorNovo3
            resultado_linhasE83 = contadorNovo4
            resultado_linhasF83 = contadorNovo5
            resultado_linhasB88 = contadorNovo6 
            resultado_linhasC88 = contadorNovo7  
            resultado_linhasD88 = contadorNovo8
            resultado_linhasE88 = contadorNovo9
            resultado_linhasF88 = contadorNovo10
            resultado_linhasB82 = contadorNovo11
            resultado_linhasC82 = contadorNovo12
            resultado_linhasD82 = contadorNovo13
            resultado_linhasE82 = contadorNovo14
            resultado_linhasF82 = contadorNovo15
            resultado_linhasB87 = contadorNovo16
            resultado_linhasC87 = contadorNovo17
            resultado_linhasD87 = contadorNovo18
            resultado_linhasE87 = contadorNovo19
            resultado_linhasF87 = contadorNovo20
            resultado_linhasB81 = contadorNovo21
            resultado_linhasC81 = contadorNovo22
            resultado_linhasD81 = contadorNovo23
            resultado_linhasE81 = contadorNovo24
            resultado_linhasF81 = contadorNovo25
            resultado_linhasB86 = contadorNovo26
            resultado_linhasC86 = contadorNovo27
            resultado_linhasD86 = contadorNovo28
            resultado_linhasE86 = contadorNovo29
            resultado_linhasF86 = contadorNovo30
            resultado_linhasB80 = contadorNovo31
            resultado_linhasC80 = contadorNovo32
            resultado_linhasD80 = contadorNovo33 
            resultado_linhasE80 = contadorNovo34
            resultado_linhasF80 = contadorNovo35
            resultado_linhasB85 = contadorNovo36
            resultado_linhasC85 = contadorNovo37
            resultado_linhasD85 = contadorNovo38
            resultado_linhasE85 = contadorNovo39 
            resultado_linhasF85 = contadorNovo40
            resultado_linhasB76 = contadorNovo41
            resultado_linhasC76 = contadorNovo42
            resultado_linhasD76 = contadorNovo43
            resultado_linhasE76 = contadorNovo44
            resultado_linhasF76 = contadorNovo45
            resultado_linhasB77 = contadorNovo46
            resultado_linhasC77 = contadorNovo47 
            resultado_linhasD77 = contadorNovo48
            resultado_linhasE77 = contadorNovo49 
            resultado_linhasF77 = contadorNovo50
            resultado_linhasB74 = contadorNovo51  
            resultado_linhasC74 = contadorNovo52
            resultado_linhasD74 = contadorNovo53  
            resultado_linhasE74 = contadorNovo54
            resultado_linhasF74 = contadorNovo55
            resultado_linhasB75 = contadorNovo56
            resultado_linhasC75 = contadorNovo57
            resultado_linhasD75 = contadorNovo58
            resultado_linhasE75 = contadorNovo59
            resultado_linhasF75 = contadorNovo60
            resultado_linhasI70 = contadorNovo61
            resultado_linhasJ70 = contadorNovo62
            resultado_linhasK70 = contadorNovo63
            resultado_linhasL70 = contadorNovo64
            resultado_linhasM70 = contadorNovo65
            resultado_linhasI71 = contadorNovo66
            resultado_linhasJ71 = contadorNovo67
            resultado_linhasK71 = contadorNovo68
            resultado_linhasL71 = contadorNovo69
            resultado_linhasM71 = contadorNovo70
            resultado_linhasI68 = contadorNovo71
            resultado_linhasJ68 = contadorNovo72
            resultado_linhasK68 = contadorNovo73
            resultado_linhasL68 = contadorNovo74
            resultado_linhasM68 = contadorNovo75
            resultado_linhasI69 = contadorNovo76
            resultado_linhasJ69 = contadorNovo77
            resultado_linhasK69 = contadorNovo78
            resultado_linhasL69 = contadorNovo79
            resultado_linhasM69 = contadorNovo80
            resultado_linhasB70 = contadorNovo81
            resultado_linhasC70 = contadorNovo82
            resultado_linhasD70 = contadorNovo83
            resultado_linhasE70 = contadorNovo84
            resultado_linhasF70 = contadorNovo85
            resultado_linhasB71 = contadorNovo86
            resultado_linhasC71 = contadorNovo87
            resultado_linhasD71 = contadorNovo88
            resultado_linhasE71 = contadorNovo89
            resultado_linhasF71 = contadorNovo90
            resultado_linhasB68 = contadorNovo91
            resultado_linhasC68 = contadorNovo92
            resultado_linhasD68 = contadorNovo93
            resultado_linhasE68 = contadorNovo94
            resultado_linhasF68 = contadorNovo95
            resultado_linhasB69 = contadorNovo96
            resultado_linhasC69 = contadorNovo97
            resultado_linhasD69 = contadorNovo98
            resultado_linhasE69 = contadorNovo99
            resultado_linhasF69 = contadorNovo100
            resultado_linhasB64 = contadorF1
            resultado_linhasC64 = contadorF2
            resultado_linhasD64 = contadorF3
            resultado_linhasE64 = contadorF4
            resultado_linhasF64 = contadorF5
            resultado_linhasB65 = contadorF6
            resultado_linhasC65 = contadorF7
            resultado_linhasD65 = contadorF8
            resultado_linhasE65 = contadorF9
            resultado_linhasF65 = contadorF10
            resultado_linhasB62 = contadorF11  
            resultado_linhasC62 = contadorF12  
            resultado_linhasD62 = contadorF13
            resultado_linhasE62 = contadorF14
            resultado_linhasF62 = contadorF15
            resultado_linhasB63 = contadorF16
            resultado_linhasC63 = contadorF17
            resultado_linhasD63 = contadorF18
            resultado_linhasE63 = contadorF19
            resultado_linhasF63 = contadorF20
            resultado_linhasI64 = contadorF21
            resultado_linhasJ64 = contadorF22
            resultado_linhasK64 = contadorF23
            resultado_linhasL64 = contadorF24
            resultado_linhasM64 = contadorF25
            resultado_linhasI65 = contadorF26
            resultado_linhasJ65 = contadorF27
            resultado_linhasK65 = contadorF28
            resultado_linhasL65 = contadorF29
            resultado_linhasM65 = contadorF30
            resultado_linhasI62 = contadorF31
            resultado_linhasJ62 = contadorF32
            resultado_linhasK62 = contadorF33
            resultado_linhasL62 = contadorF34
            resultado_linhasM62 = contadorF35
            resultado_linhasI63 = contadorF36
            resultado_linhasJ63 = contadorF37
            resultado_linhasK63 = contadorF38
            resultado_linhasL63 = contadorF39
            resultado_linhasM63 = contadorF40
            resultado_linhasI48 = contadorF41
            resultado_linhasJ48 = contadorF42
            resultado_linhasK48 = contadorF43
            resultado_linhasL48 = contadorF44
            resultado_linhasM48 = contadorF45
            resultado_linhasI59 = contadorF46
            resultado_linhasJ59 = contadorF47
            resultado_linhasK59 = contadorF48
            resultado_linhasL59 = contadorF49
            resultado_linhasM59 = contadorF50
            resultado_linhasB48 = contadorF51
            resultado_linhasC48 = contadorF52
            resultado_linhasD48 = contadorF53
            resultado_linhasE48 = contadorF54
            resultado_linhasF48 = contadorF55
            resultado_linhasB59 = contadorF56
            resultado_linhasC59 = contadorF57
            resultado_linhasD59 = contadorF58
            resultado_linhasE59 = contadorF59 
            resultado_linhasF59 = contadorF60      


#INSERINDO RESULTADOS NA PLANILHA APURAÇÃO 
        # if dir_2.endswith('.xlsx'):
        # dir_2 = "C:\\Users\\karina.mediani\\Desktop\\BRADOPERADORA\\Apuracao"
        # if dir_2.endswith('.xlsx'):
            wb1 = load_workbook(dir_2)
            ws1 = wb1.active
            # print(ws1)


            ws1['I10'] = resultado_linhasI10 #RESULTADO  I10
            ws1['J10'] = resultado_linhasJ10#RESULTADO  J10
            ws1['K10'] = resultado_linhasK10 #RESULTADO  K10
            ws1['L10'] = resultado_linhasL10 #RESULTADO  L10
            ws1['M10'] = resultado_linhasM10 #RESULTADO  M10
            ws1['I11'] = resultado_linhasI11 #RESULTADO  I11
            ws1['J11'] = resultado_linhasJ11 #RESULTADO  J11
            ws1['K11'] = resultado_linhasK11 #RESULTADO K11
            ws1['L11'] = resultado_linhasL11 #RESULTADO L11
            ws1['M11'] = resultado_linhasM11 #RESULTADO  M11
            ws1['I19'] = resultado_linhasI19 #RESULTADO  I19
            ws1['J19'] = resultado_linhasJ19 #RESULTADO J19
            ws1['K19'] = resultado_linhasK19 #RESULTADO K19
            ws1['L19'] = resultado_linhasL19 #RESULTADOL19
            ws1['M19'] = resultado_linhasM19 #RESULTADO M19
            ws1['B22'] = resultado_linhasB22 #RESULTADO B22
            ws1['C22'] = resultado_linhasC22 #RESULTADO C22
            ws1['D22'] = resultado_linhasD22 #RESULTADO D22
            ws1['E22'] = resultado_linhasE22 #RESULTADO E22
            ws1['F22'] = resultado_linhasF22 #RESULTADOF22
            ws1['B23'] = resultado_linhasB23 #RESULTADO B23
            ws1['C23'] = resultado_linhasC23 #RESULTADO C23
            ws1['D23'] = resultado_linhasD23 #RESULTADO D23
            ws1['E23'] = resultado_linhasE23 #RESULTADO E23
            ws1['F23'] = resultado_linhasF23 #RESULTADO F23
            ws1['B24'] = resultado_linhasB24 #RESULTADO B24
            ws1['C24'] = resultado_linhasC24 #RESULTADO C24
            ws1['D24'] = resultado_linhasD24 #RESULTADO D24
            ws1['E24'] = resultado_linhasE24 #RESULTADO E24
            ws1['F24'] = resultado_linhasF24 #RESULTADO F24
            ws1['B25'] = resultado_linhasB25 #RESULTADO B25
            ws1['C25'] = resultado_linhasC25 #RESULTADO C25
            ws1['D25'] = resultado_linhasD25 #RESULTADO D25
            ws1['E25'] = resultado_linhasE25 #RESULTADO E25
            ws1['F25'] = resultado_linhasF25 #RESULTADO F25
            ws1['B26'] = resultado_linhasB26 #RESULTADO B26
            ws1['C26'] = resultado_linhasC26 #RESULTADO C26
            ws1['D26'] = resultado_linhasD26 #RESULTADO D26
            ws1['E26'] = resultado_linhasE26 #RESULTADO E26
            ws1['F26'] = resultado_linhasF26 #RESULTADO F26
            ws1['B27'] = resultado_linhasB27 #RESULTADO B27
            ws1['C27'] = resultado_linhasC27 #RESULTADO C27
            ws1['D27'] = resultado_linhasD27 #RESULTADO D27        
            ws1['E27'] = resultado_linhasE27  #RESULTADO E27 
            ws1['F27'] = resultado_linhasF27 #RESULTADO F27 
            ws1['B28'] = resultado_linhasB28 #RESULTADO B28 
            ws1['C28'] = resultado_linhasC28 #RESULTADO C28 
            ws1['D28'] = resultado_linhasD28 #RESULTADO D28
            ws1['E28'] = resultado_linhasE28 #RESULTADOE28
            ws1['F28'] = resultado_linhasF28 #RESULTADO F28
            ws1['I23'] = resultado_linhasI23 #RESULTADO I23
            ws1['J23'] = resultado_linhasJ23 #RESULTADO J23
            ws1['K23'] = resultado_linhasK23 #RESULTADO K23
            ws1['L23'] = resultado_linhasL23 #RESULTADO L23
            ws1['M23'] = resultado_linhasM23 #RESULTADO M23
            ws1['I22'] = resultado_linhasI22 #RESULTADO I22
            ws1['J22'] = resultado_linhasJ22 #RESULTADO J22  
            ws1['K22'] = resultado_linhasK22 #RESULTADO K22
            ws1['L22'] = resultado_linhasL22 #RESULTADO L22
            ws1['M22'] = resultado_linhasM22 #RESULTADO M22
            ws1['I28'] = resultado_linhasI28 #RESULTADO I28
            ws1['J28'] = resultado_linhasJ28 #RESULTADO J28
            ws1['K28'] = resultado_linhasK28 #RESULTADO K28
            ws1['L28'] = resultado_linhasL28 #RESULTADP L28
            ws1['M28'] = resultado_linhasM28 #RESULTADO M28
            ws1['I26'] = resultado_linhasI26 #RESULTADO I26
            ws1['J26'] = resultado_linhasJ26 #RESULTADO J26
            ws1['K26'] = resultado_linhasK26 #RESULTADO K26
            ws1['L26'] = resultado_linhasL26 #RESULTADO L26
            ws1['M26'] = resultado_linhasM26 #RESULTADO M26
            ws1['I27'] = resultado_linhasI27 #RESULTADO I27
            ws1['J27'] = resultado_linhasJ27 #RESULTADO J27
            ws1['K27'] = resultado_linhasK27 #RESULTADO K27
            ws1['L27'] = resultado_linhasL27 #RESULTADO L27
            ws1['M27'] = resultado_linhasM27 #RESULTADO M27
            ws1['I25'] = resultado_linhasI25 #RESULTADO I25
            ws1['J25'] = resultado_linhasJ25 #RESULTADO J25
            ws1['K25'] = resultado_linhasK25 #RESULTADO K25
            ws1['L25'] = resultado_linhasL25 #RESULTADOL25
            ws1['M25'] = resultado_linhasM25 #RESULTADO M25
            ws1['I24'] = resultado_linhasI24 #RESULTADOI24
            ws1['J24'] = resultado_linhasJ24 #RESULTADO J24
            ws1['K24'] = resultado_linhasK24 #RESULTADO K24
            ws1['L24'] = resultado_linhasL24 #RESULTADO L24
            ws1['M24'] = resultado_linhasM24 #RESULTADO M24
            ws1['B32'] = resultado_linhasB32 #RESULTADO B32
            ws1['C32'] = resultado_linhasC32 #RESULTADO C32
            ws1['D32'] = resultado_linhasD32 #RESULTADO D32
            ws1['E32'] = resultado_linhasE32 #RESULTADO E32
            ws1['F32'] = resultado_linhasF32 #RESULTADO F32
            ws1['B31'] = resultado_linhasB31 #RESULTADO B31
            ws1['C31'] = resultado_linhasC31 #RESULTADO C31
            ws1['D31'] = resultado_linhasD31 #RESULTADO D31
            ws1['E31'] = resultado_linhasE31 #RESULTADO E31
            ws1['F31'] = resultado_linhasF31 #RESULTADO F31
            ws1['B37'] = resultado_linhasB37 #RESULTADO B37
            ws1['C37'] = resultado_linhasC37 #RESULTADO C37
            ws1['D37'] = resultado_linhasD37 #RESULTADO D37
            ws1['E37'] = resultado_linhasE37 #RESULTADO E37
            ws1['F37'] = resultado_linhasF37 #RESULTADO F37
            ws1['B35'] = resultado_linhasB35 #RESULTADO B35
            ws1['C35'] = resultado_linhasC35 #RESULTADO C35
            ws1['D35'] = resultado_linhasD35 #RESULTADO D35
            ws1['E35'] = resultado_linhasE35 #RESULTADO E35
            ws1['F35'] = resultado_linhasF35 #RESULTADO F35
            ws1['B36'] = resultado_linhasB36 #RESULTADO B36
            ws1['C36'] = resultado_linhasC36 #RESULTADO C36
            ws1['D36'] = resultado_linhasD36 #RESULTADO D36
            ws1['E36'] = resultado_linhasE36 #RESULTADO E36
            ws1['F36'] = resultado_linhasF36 #RESULTADO F36
            ws1['B34'] = resultado_linhasB34 #RESULTADO B34
            ws1['C34'] = resultado_linhasC34 #RESULTADO C34
            ws1['D34'] = resultado_linhasD34 #RESULTADO D34
            ws1['E34'] = resultado_linhasE34 #RESULTADO E34
            ws1['F34'] = resultado_linhasF34 #RESULTADO F34
            ws1['B33'] = resultado_linhasB33
            ws1['C33'] = resultado_linhasC33
            ws1['D33'] = resultado_linhasD33
            ws1['E33'] = resultado_linhasE33
            ws1['F33'] = resultado_linhasF33
            ws1['I32'] = resultado_linhasI32 
            ws1['J32'] = resultado_linhasJ32
            ws1['K32'] = resultado_linhasK32
            ws1['L32'] = resultado_linhasL32
            ws1['M32'] = resultado_linhasM32
            ws1['I31'] = resultado_linhasI31
            ws1['J31'] = resultado_linhasJ31
            ws1['K31'] = resultado_linhasK31
            ws1['L31'] = resultado_linhasL31
            ws1['M31'] = resultado_linhasM31
            ws1['I37'] = resultado_linhasI37
            ws1['J37'] = resultado_linhasJ37
            ws1['K37'] = resultado_linhasK37
            ws1['L37'] = resultado_linhasL37
            ws1['M37'] = resultado_linhasM37
            ws1['I34'] = resultado_linhasI34
            ws1['J34'] = resultado_linhasJ34
            ws1['K34'] = resultado_linhasK34
            ws1['L34'] = resultado_linhasL34
            ws1['M34'] = resultado_linhasM34
            ws1['I35'] = resultado_linhasI35
            ws1['J35'] = resultado_linhasJ35
            ws1['K35'] = resultado_linhasK35
            ws1['L35'] = resultado_linhasL35
            ws1['M35'] = resultado_linhasM35 
            ws1['I36'] = resultado_linhasI36 
            ws1['J36'] = resultado_linhasJ36
            ws1['K36'] = resultado_linhasK36
            ws1['L36'] = resultado_linhasL36
            ws1['M36'] = resultado_linhasM36
            ws1['I33'] = resultado_linhasI33       
            ws1['J33'] = resultado_linhasJ33
            ws1['K33'] = resultado_linhasK33
            ws1['L33'] = resultado_linhasL33 
            ws1['M33'] = resultado_linhasM33
            ws1['I40'] = resultado_linhasI40
            ws1['J40'] = resultado_linhasJ40
            ws1['K40'] = resultado_linhasK40  
            ws1['L40'] = resultado_linhasL40   
            ws1['M40'] = resultado_linhasM40
            ws1['I51'] = resultado_linhasI51
            ws1['J51'] = resultado_linhasJ51
            ws1['K51'] = resultado_linhasK51
            ws1['L51'] = resultado_linhasL51
            ws1['M51'] = resultado_linhasM51
            ws1['I41'] = resultado_linhasI41
            ws1['J41'] = resultado_linhasJ41
            ws1['K41'] = resultado_linhasK41
            ws1['L41'] = resultado_linhasL41
            ws1['M41'] = resultado_linhasM41
            ws1['I52'] = resultado_linhasI52
            ws1['J52'] = resultado_linhasJ52
            ws1['K52'] = resultado_linhasK52
            ws1['L52'] = resultado_linhasL52
            ws1['M52'] = resultado_linhasM52
            ws1['B41'] = resultado_linhasB41
            ws1['C41'] = resultado_linhasC41
            ws1['D41'] = resultado_linhasD41
            ws1['E41'] = resultado_linhasE41
            ws1['F41'] = resultado_linhasF41
            ws1['B52'] = resultado_linhasB52
            ws1['C52'] = resultado_linhasC52
            ws1['D52'] = resultado_linhasD52
            ws1['E52'] = resultado_linhasE52
            ws1['F52'] = resultado_linhasF52
            ws1['B40'] = resultado_linhasB40
            ws1['C40'] = resultado_linhasC40
            ws1['D40'] = resultado_linhasD40
            ws1['E40'] = resultado_linhasE40      
            ws1['F40'] = resultado_linhasF40    
            ws1['B51'] = resultado_linhasB51
            ws1['C51'] = resultado_linhasC51
            ws1['D51'] = resultado_linhasD51
            ws1['E51'] = resultado_linhasE51
            ws1['F51'] = resultado_linhasF51
            ws1['I42'] = resultado_linhasI42  
            ws1['J42'] = resultado_linhasJ42
            ws1['K42'] = resultado_linhasK42
            ws1['L42'] = resultado_linhasL42
            ws1['M42'] = resultado_linhasM42 
            ws1['I53'] = resultado_linhasI53
            ws1['J53'] = resultado_linhasJ53
            ws1['K53'] = resultado_linhasK53
            ws1['L53'] = resultado_linhasL53
            ws1['M53'] = resultado_linhasM53
            ws1['B42'] = resultado_linhasB42
            ws1['C42'] = resultado_linhasC42
            ws1['D42'] = resultado_linhasD42
            ws1['E42'] = resultado_linhasE42
            ws1['F42'] = resultado_linhasF42
            ws1['B53'] = resultado_linhasB53
            ws1['C53'] = resultado_linhasC53
            ws1['D53'] = resultado_linhasD53
            ws1['E53'] = resultado_linhasE53
            ws1['F53'] = resultado_linhasF53
            ws1['I43'] = resultado_linhasI43
            ws1['J43'] = resultado_linhasJ43
            ws1['K43'] = resultado_linhasK43
            ws1['L43'] = resultado_linhasL43
            ws1['M43'] = resultado_linhasM43
            ws1['I54'] = resultado_linhasI54
            ws1['J54'] = resultado_linhasJ54
            ws1['K54'] = resultado_linhasK54
            ws1['L54'] = resultado_linhasL54
            ws1['M54'] = resultado_linhasM54
            ws1['B43'] = resultado_linhasB43
            ws1['C43'] = resultado_linhasC43
            ws1['D43'] = resultado_linhasD43
            ws1['F43'] = resultado_linhasF43
            ws1['B54'] = resultado_linhasB54  
            ws1['C54'] = resultado_linhasC54      
            ws1['D54'] = resultado_linhasD54 
            ws1['E54'] = resultado_linhasE54
            ws1['F54'] = resultado_linhasF54
            ws1['I44'] = resultado_linhasI44           
            ws1['J44'] = resultado_linhasJ44    
            ws1['K44'] = resultado_linhasK44 
            ws1['L44'] = resultado_linhasL44
            ws1['M44'] = resultado_linhasM44     
            ws1['I55'] = resultado_linhasI55  
            ws1['J55'] = resultado_linhasJ55   
            ws1['K55'] = resultado_linhasK55  
            ws1['L55'] = resultado_linhasL55
            ws1['M55'] = resultado_linhasM55
            ws1['B44'] = resultado_linhasB44
            ws1['C44'] = resultado_linhasC44
            ws1['D44'] = resultado_linhasD44
            ws1['E44'] = resultado_linhasE44
            ws1['F44'] = resultado_linhasF44
            ws1['B55'] = resultado_linhasB55
            ws1['C55'] = resultado_linhasC55
            ws1['D55'] = resultado_linhasD55
            ws1['E55'] = resultado_linhasE55
            ws1['F55'] = resultado_linhasF55
            ws1['I45'] = resultado_linhasI45   
            ws1['J45'] = resultado_linhasJ45
            ws1['K45'] = resultado_linhasK45   
            ws1['L45'] = resultado_linhasL45
            ws1['M45'] = resultado_linhasM45
            ws1['I56'] = resultado_linhasI56
            ws1['J56'] = resultado_linhasJ56
            ws1['K56'] = resultado_linhasK56
            ws1['L56'] = resultado_linhasL56
            ws1['M56'] = resultado_linhasM56
            ws1['B45'] = resultado_linhasB45
            ws1['C45'] = resultado_linhasC45
            ws1['D45'] = resultado_linhasD45
            ws1['E45'] = resultado_linhasE45
            ws1['F45'] = resultado_linhasF45
            ws1['B56'] = resultado_linhasB56
            ws1['C56'] = resultado_linhasC56    
            ws1['D56'] = resultado_linhasD56
            ws1['E56'] = resultado_linhasE56
            ws1['F56'] = resultado_linhasF56
            ws1['I46'] = resultado_linhasI46
            ws1['J46'] = resultado_linhasJ46
            ws1['K46'] = resultado_linhasK46
            ws1['L46'] = resultado_linhasL46
            ws1['M46'] = resultado_linhasM46 
            ws1['I57'] = resultado_linhasI57
            ws1['J57'] = resultado_linhasJ57
            ws1['K57'] = resultado_linhasK57    
            ws1['L57'] = resultado_linhasL57
            ws1['M57'] = resultado_linhasM57        
            ws1['B46'] = resultado_linhasB46
            ws1['C46'] = resultado_linhasC46  
            ws1['D46'] = resultado_linhasD46
            ws1['E46'] = resultado_linhasE46 
            ws1['F46'] = resultado_linhasF46  
            ws1['B57'] = resultado_linhasB57
            ws1['C57'] = resultado_linhasC57  
            ws1['D57'] = resultado_linhasD57  
            ws1['E57'] = resultado_linhasE57
            ws1['F57'] = resultado_linhasF57 
            ws1['I47'] = resultado_linhasI47
            ws1['J47'] = resultado_linhasJ47   
            ws1['K47'] = resultado_linhasK47
            ws1['L47'] = resultado_linhasL47
            ws1['M47'] = resultado_linhasM47       
            ws1['I58'] = resultado_linhasI58
            ws1['J58'] = resultado_linhasJ58
            ws1['J58'] = resultado_linhasJ58
            ws1['K58'] = resultado_linhasK58
            ws1['L58'] = resultado_linhasL58
            ws1['M58'] = resultado_linhasM58
            ws1['B47'] = resultado_linhasB47   
            ws1['C47'] = resultado_linhasC47
            ws1['D47'] = resultado_linhasD47  
            ws1['E47'] = resultado_linhasE47
            ws1['F47'] = resultado_linhasF47   
            ws1['B58'] = resultado_linhasB58       
            ws1['C58'] = resultado_linhasC58      
            ws1['D58'] = resultado_linhasD58    
            ws1['E58'] = resultado_linhasE58      
            ws1['F58'] = resultado_linhasF58 

            ws1['B83'] = resultado_linhasB83
            ws1['C83'] = resultado_linhasC83
            ws1['D83'] = resultado_linhasD83
            ws1['E83'] = resultado_linhasE83
            ws1['F83'] = resultado_linhasF83
            ws1['B88'] = resultado_linhasB88
            ws1['C88'] = resultado_linhasC88
            ws1['D88'] = resultado_linhasD88
            ws1['E88'] = resultado_linhasE88
            ws1['F88'] = resultado_linhasF88
            ws1['B82'] = resultado_linhasB82
            ws1['C82'] = resultado_linhasC82
            ws1['D82'] = resultado_linhasD82
            ws1['E82'] = resultado_linhasE82
            ws1['F82'] = resultado_linhasF82
            ws1['B87'] = resultado_linhasB87
            ws1['C87'] = resultado_linhasC87
            ws1['D87'] = resultado_linhasD87
            ws1['E87'] = resultado_linhasE87
            ws1['F87'] = resultado_linhasF87
            ws1['B81'] = resultado_linhasB81
            ws1['C81'] = resultado_linhasC81
            ws1['D81'] = resultado_linhasD81
            ws1['E81'] = resultado_linhasE81
            ws1['F81'] = resultado_linhasF81
            ws1['B86'] = resultado_linhasB86
            ws1['C86'] = resultado_linhasC86
            ws1['D86'] = resultado_linhasD86
            ws1['E86'] = resultado_linhasE86
            ws1['F86'] = resultado_linhasF86
            ws1['B80'] = resultado_linhasB80
            ws1['C80'] = resultado_linhasC80
            ws1['D80'] = resultado_linhasD80
            ws1['E80'] = resultado_linhasE80
            ws1['F80'] = resultado_linhasF80
            ws1['B85'] = resultado_linhasB85
            ws1['C85'] = resultado_linhasC85
            ws1['D85'] = resultado_linhasD85
            ws1['E85'] = resultado_linhasE85
            ws1['F85'] = resultado_linhasF85
            ws1['B76'] = resultado_linhasB76
            ws1['C76'] = resultado_linhasC76
            ws1['D76'] = resultado_linhasD76
            ws1['E76'] = resultado_linhasE76
            ws1['F76'] = resultado_linhasF76
            ws1['B77'] = resultado_linhasB77
            ws1['C77'] = resultado_linhasC77
            ws1['D77'] = resultado_linhasD77
            ws1['E77'] = resultado_linhasE77
            ws1['F77'] = resultado_linhasF77
            ws1['B74'] = resultado_linhasB74
            ws1['C74'] = resultado_linhasC74
            ws1['D74'] = resultado_linhasD74
            ws1['E74'] = resultado_linhasE74
            ws1['F74'] = resultado_linhasF74
            ws1['B75'] = resultado_linhasB75
            ws1['C75'] = resultado_linhasC75
            ws1['D75'] = resultado_linhasD75
            ws1['E75'] = resultado_linhasE75
            ws1['F75'] = resultado_linhasF75
            ws1['I70'] = resultado_linhasI70
            ws1['J70'] = resultado_linhasJ70
            ws1['K70'] = resultado_linhasK70
            ws1['L70'] = resultado_linhasL70
            ws1['M70'] = resultado_linhasM70
            ws1['I71'] = resultado_linhasI71
            ws1['J71'] = resultado_linhasJ71
            ws1['K71'] = resultado_linhasK71
            ws1['L71'] = resultado_linhasL71
            ws1['M71'] = resultado_linhasM71
            ws1['I68'] = resultado_linhasI68
            ws1['J68'] = resultado_linhasJ68
            ws1['K68'] = resultado_linhasK68
            ws1['L68'] = resultado_linhasL68
            ws1['M68'] = resultado_linhasM68
            ws1['I69'] = resultado_linhasI69
            ws1['J69'] = resultado_linhasJ69
            ws1['K69'] = resultado_linhasK69
            ws1['L69'] = resultado_linhasL69
            ws1['M69'] = resultado_linhasM69
            ws1['B70'] = resultado_linhasB70
            ws1['C70'] = resultado_linhasC70
            ws1['D70'] = resultado_linhasD70
            ws1['E70'] = resultado_linhasE70
            ws1['F70'] = resultado_linhasF70
            ws1['B71'] = resultado_linhasB71
            ws1['C71'] = resultado_linhasC71
            ws1['D71'] = resultado_linhasD71
            ws1['E71'] = resultado_linhasE71
            ws1['F71'] = resultado_linhasF71
            ws1['B68'] = resultado_linhasB68
            ws1['C68'] = resultado_linhasC68
            ws1['D68'] = resultado_linhasD68
            ws1['E68'] = resultado_linhasE68
            ws1['F68'] = resultado_linhasF68
            ws1['B69'] = resultado_linhasB69
            ws1['C69'] = resultado_linhasC69
            ws1['D69'] = resultado_linhasD69
            ws1['E69'] = resultado_linhasE69
            ws1['F69'] = resultado_linhasF69
            ws1['B64'] = resultado_linhasB64
            ws1['C64'] = resultado_linhasC64
            ws1['D64'] = resultado_linhasD64
            ws1['E64'] = resultado_linhasE64
            ws1['F64'] = resultado_linhasF64
            ws1['B65'] = resultado_linhasB65
            ws1['C65'] = resultado_linhasC65
            ws1['D65'] = resultado_linhasD65
            ws1['E65'] = resultado_linhasE65
            ws1['F65'] = resultado_linhasF65
            ws1['B62'] = resultado_linhasB62
            ws1['C62'] = resultado_linhasC62
            ws1['D62'] = resultado_linhasD62
            ws1['E62'] = resultado_linhasE62
            ws1['F62'] = resultado_linhasF62
            ws1['B63'] = resultado_linhasB63
            ws1['C63'] = resultado_linhasC63
            ws1['D63'] = resultado_linhasD63
            ws1['E63'] = resultado_linhasE63
            ws1['F63'] = resultado_linhasF63
            ws1['I64'] = resultado_linhasI64
            ws1['J64'] = resultado_linhasJ64
            ws1['K64'] = resultado_linhasK64
            ws1['L64'] = resultado_linhasL64
            ws1['M64'] = resultado_linhasM64
            ws1['I65'] = resultado_linhasI65
            ws1['J65'] = resultado_linhasJ65
            ws1['K65'] = resultado_linhasK65
            ws1['L65'] = resultado_linhasL65
            ws1['M65'] = resultado_linhasM65
            ws1['I62'] = resultado_linhasI62
            ws1['J62'] = resultado_linhasJ62
            ws1['K62'] = resultado_linhasK62
            ws1['L62'] = resultado_linhasL62
            ws1['M62'] = resultado_linhasM62
            ws1['I63'] = resultado_linhasI63
            ws1['J63'] = resultado_linhasJ63
            ws1['K63'] = resultado_linhasK63
            ws1['L63'] = resultado_linhasL63
            ws1['M63'] = resultado_linhasM63
            ws1['I48'] = resultado_linhasI48
            ws1['J48'] = resultado_linhasJ48
            ws1['K48'] = resultado_linhasK48
            ws1['L48'] = resultado_linhasL48
            ws1['M48'] = resultado_linhasM48
            ws1['I59'] = resultado_linhasI59
            ws1['J59'] = resultado_linhasJ59
            ws1['K59'] = resultado_linhasK59
            ws1['L59'] = resultado_linhasL59
            ws1['M59'] = resultado_linhasM59
            ws1['B48'] = resultado_linhasB48
            ws1['C48'] = resultado_linhasC48
            ws1['D48'] = resultado_linhasD48
            ws1['E48'] = resultado_linhasE48
            ws1['F48'] = resultado_linhasF48
            ws1['B59'] = resultado_linhasB59
            ws1['C59'] = resultado_linhasC59
            ws1['D59'] = resultado_linhasD59
            ws1['E59'] = resultado_linhasE59
            ws1['F59'] = resultado_linhasF59    

            wb1.save(dir_2)
    except Exception as e:
        logging.error('| Ocorreu um erro: | 3')
        logging.exception(str(e))  