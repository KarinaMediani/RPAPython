import logging
import xml.etree.ElementTree as ET
from openpyxl import load_workbook

def bon_dep_menor_valid():
    try:
        tree = ET.parse("C:\\Users\\User\\Desktop\\206-BS-571\\diretório\\diretorio.xml")
        root2 = tree.getroot()
        for child2 in root2:
                for x2 in root2.findall(child2.tag):
                        confeop = x2.find('confeop').text
                        apcia = x2.find('apcia').text
        dir_2 = apcia

                
        wb = load_workbook(dir_2)
        ws = wb.active
        ##VALORES SOLICITADOS 
        valor1 = ws['N15'].value
        valor2 = ws['N11'].value
        if ((valor1 == 0 ) and (valor2 == 0)):
            ws['J85'] = 'BONIF. DEP. MENOR VALIDADO 0%'
        else:
            divisao = (valor1 /  valor2)
        # resultado = divisao
            conversao_para_percentual = "{:.0%}".format(divisao)

            ws['J85'] = 'BONIF. DEP. MENOR VALIDADO  ' + conversao_para_percentual

        wb.save(dir_2)
    except Exception as e:
                logging.error('| Ocorreu um erro: | 3')
                logging.exception(str(e))    