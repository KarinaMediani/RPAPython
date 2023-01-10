from openpyxl import load_workbook
import logging
import xml.etree.ElementTree as ET
def ansop_X_ProdutosAtivos():
    try:
        tree = ET.parse("C:\\Users\\User\\Desktop\\206-BS-571\\diretório\\diretorio.xml")
        root2 = tree.getroot()
        for child2 in root2:
                for x2 in root2.findall(child2.tag):
                        ansop = x2.find('ansop').text
                        proativosop = x2.find('proativosop').text
        #arquivo Planos Saúde ansop
        diretorio = ansop
        wb = load_workbook(diretorio)
        ws = wb.active

        ##arquivo produtos ativos###
        diretorio_2 = proativosop
        wb1 = load_workbook(diretorio_2)
        contador = 1

        ws1 = wb1['Situação do Plano']

        for linhas in ws1:
            coluna_A = str(linhas[0].value).upper().strip()
            for row in ws:
                if coluna_A == str(row[0].value):
                    ws1[f'B{contador}'] = str(row[7].value)
            contador +=1

        wb1.save(diretorio_2)
    except Exception as e:
                logging.error('| Ocorreu um erro: | 3')
                logging.exception(str(e))    