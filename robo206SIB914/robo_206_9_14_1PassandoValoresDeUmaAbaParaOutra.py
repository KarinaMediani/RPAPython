import logging
import xml.etree.ElementTree as ET
from openpyxl import load_workbook

def AbaProdutosAtivos():
    try:
        tree = ET.parse("C:\\Users\\User\\Desktop\\206-BS-571\\diretório\\diretorio.xml")
        root2 = tree.getroot()
        for child2 in root2:
                for x2 in root2.findall(child2.tag):
                        proativosop = x2.find('proativosop').text
                        apcia = x2.find('apcia').text
        #ABRINDO ARQUIVO PRODUTOS ATIVOS
        diretorio = proativosop
        wb = load_workbook(diretorio)
        ws = wb['Produtos Ativos']

        ws1 = wb['Situação do Plano']

        #PEGANDO VALORES DAS COLUNAS A e C ABA PRODUTOS ATIVOS
        coluna_c = []
        coluna_A = []
        for linha in ws.iter_rows(min_row=3):
            valores_A = str(linha[0].value)
            valores_C = str(linha[2].value)
            if valores_A == 'None':
                continue
            else:
                coluna_A.append(valores_A)
            
            if valores_C == 'None':
                continue
            else:
                coluna_c.append(valores_C)

        # print(coluna_c)
        #Colando itens das colunas A e C para SITUAÇÃO DO PLANO
        contador = 2
        for item in coluna_A: 
            ws1[f'A{contador}'] = item
            contador +=1

        contador = 2
        for item in coluna_c:
            ws1[f'C{contador}'] = item
            contador +=1
        ws1['A1'] = 'PRODUTO'
        ws1['B1'] = 'SITUAÇÃO DO PLANO'
        ws1['C1'] = 'VIDAS'
        wb.save(diretorio)
    except Exception as e:
                logging.error('| Ocorreu um erro: | 3')
                logging.exception(str(e))    