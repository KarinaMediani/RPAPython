from openpyxl import load_workbook
import logging
import xml.etree.ElementTree as ET
import os
def linhas_Vazias():
    try:
        tree = ET.parse("C:\\Users\\User\\Desktop\\206-BS-571\\diretório\\diretorio.xml")
        root2 = tree.getroot()
        for child2 in root2:
                for x2 in root2.findall(child2.tag):
                        proativosop = x2.find('proativosop').text
                        confeop = x2.find('confeop').text
        diretorio = proativosop
        diretorio2 = confeop
        lista_arquivo = os.listdir(diretorio2)

        vPlanilha = load_workbook(diretorio)
        vSeguradosSemProdutosAtivos = vPlanilha['Segurados sem Produtos Ativos']#colar

        for arquivo in lista_arquivo:
            # print(arquivo)
            if arquivo.startswith('ArqConf') and arquivo.endswith('.xlsx'):
                wb = load_workbook(diretorio2 + '\\' + arquivo)
                vAbaOrigem = wb['ORIGEM']#retira

                vListaDeLinhas = []
                
                for linhas in vAbaOrigem.iter_rows(min_row=2):
                    vColunaAI = str(linhas[34].value)
                    vColunaAO = str(linhas[40].value)
                    for item in linhas:#celula a celula
                        if "None" in (vColunaAO or vColunaAI):#coluna a ou d com none
                            for itens in linhas:
                                vListaDeLinhas.append(itens.value)
                            vSeguradosSemProdutosAtivos.append(vListaDeLinhas)
                            vListaDeLinhas = []
                        break
        vPlanilha.save(diretorio)
    except Exception as e:
                logging.error('| Ocorreu um erro: | 3')
                logging.exception(str(e))    