from openpyxl import load_workbook
import os
import logging
import xml.etree.ElementTree as ET
import time
def confXxml():
    try:
        tree = ET.parse("C:\\Users\\User\\Desktop\\206-BS-571\\diretório\\diretorio.xml")
        roo = tree.getroot()
        for child2 in roo:
                for x2 in roo.findall(child2.tag):
                        confeop = x2.find('confeop').text
                        diagop = x2.find('diagop').text
                        
        diretorio = confeop
        lista_arquivo = os.listdir(diretorio)

        tree = ET.parse(diagop + "\\Dependente maior não identificado.xml")
        root = tree.getroot()
        tree1 = ET.parse(diagop + "\\Dependente maior não validado.xml")
        root1 = tree1.getroot()
        tree2 = ET.parse(diagop + "\\Dependente menor não encontrado.xml")
        root2 = tree2.getroot()
        tree3 = ET.parse(diagop+ "\\Dependente menor não identificado.xml")
        root3 = tree3.getroot()
        tree4 = ET.parse(diagop + "\\Dependente menor não validado.xml")
        root4 = tree4.getroot()
        tree5 = ET.parse(diagop+"\\Registro de titular sem CPF.xml")
        root5 = tree5.getroot()
        tree6 = ET.parse(diagop+"\\Registros com CNS não preenchidos.xml")
        root6 = tree6.getroot()
        tree7 = ET.parse(diagop+"\\Registros com CPF não preenchidos.xml")
        root7 = tree7.getroot()
        tree8 = ET.parse(diagop+"\\Registros com data de Nascimento Divergente.xml")
        root8 = tree8.getroot()
        tree9 = ET.parse(diagop+"\\Registro de dependente maior sem CPF.xml")
        root9 = tree9.getroot()
        tree10 = ET.parse(diagop+"\\Registros envolvidos nas repetições de CNS.xml")
        root10 = tree10.getroot()
        tree11 = ET.parse(diagop+"\\Registros envolvidos nas repetições de CPF.xml")
        root11 = tree11.getroot()
        tree12 = ET.parse(diagop+"\\Titular não identificado.xml")
        root12 = tree12.getroot()

        dmNaoIndentificado = []
        dmNaoIndentificado1 = []
        dmNaoValidado = []
        dmNaoValidado1 = []
        dmNaoEncontrado = []
        dmNaoEncontrado1 = []
        dMenorNIdentificado = []
        dMenorNIdentificado1 = []
        dMenorNaoValidado = []
        dMenorNaoValidado1 = []
        rtSemCPF= []
        rtSemCPF1= []
        rCNSNaoPreenchidos = []
        rCNSNaoPreenchidos1 = []
        rCPFNaoPreenchidos = []
        rCPFNaoPreenchidos1 = []
        rdNascimentoD = []
        rdNascimentoD1 = []
        rdMaiorsCPF = []
        rdMaiorsCPF1 = []
        rerCNS =[]
        rerCNS1 =[]
        rerCPF =[]
        rerCPF1 =[]
        tNAOi =[]
        tNAOi1 =[]

        for child in root: #9.8.15
            for x in root.findall(child.tag+"/*"):
                dmNaoIndentificado.append(x.find('cco').text)
                dmNaoIndentificado1.append(x.find('codigoBeneficiario').text)
        for child1 in root1:#9.8.16
            for x1 in root1.findall(child1.tag+"/*"):
                dmNaoValidado.append(x1.find('cco').text)
                dmNaoValidado1.append(x1.find('codigoBeneficiario').text)
            #     #9.8.17
        for child2 in root2:
            for x2 in root2.findall(child2.tag+"/*"):
                dmNaoEncontrado.append(x2.find('cco').text)
                dmNaoEncontrado1.append(x2.find('codigoBeneficiario').text)

            #     #9.8.18
        for child3 in root3:
            for x3 in root3.findall(child3.tag+"/*"):
                dMenorNIdentificado.append(x3.find('cco').text)
                dMenorNIdentificado1.append(x3.find('codigoBeneficiario').text)
            #     #9.8.19
        for child4 in root4:
            for x4 in root4.findall(child4.tag+"/*"):
                dMenorNaoValidado.append(x4.find('cco').text)
                dMenorNaoValidado1.append(x4.find('codigoBeneficiario').text)
            #     #9.8.20
        for child5 in root5:
            for x5 in root5.findall(child5.tag+"/*"):
                rtSemCPF.append(x5.find('cco').text)
                rtSemCPF1.append(x5.find('codigoBeneficiario').text)

            #     #9.8.21
        for child6 in root6:
            for x6 in root6.findall(child6.tag+"/*"):
                rCNSNaoPreenchidos.append(x6.find('cco').text)
                rCNSNaoPreenchidos1.append(x6.find('codigoBeneficiario').text)
            #     #9.8.22
        for child7 in root7:
            for x7 in root7.findall(child7.tag+"/*"):
                rCPFNaoPreenchidos.append(x7.find('cco').text)
                rCPFNaoPreenchidos1.append(x7.find('codigoBeneficiario').text)
            #     #9.8.23
        for child8 in root8:
            for x8 in root8.findall(child8.tag+"/*"):
                rdNascimentoD.append(x8.find('cco').text)
                rdNascimentoD1.append(x8.find('codigoBeneficiario').text)
            #     #9.8.24
        for child9 in root9:
            for x9 in root9.findall(child9.tag+"/*"):
                rdMaiorsCPF.append(x9.find('cco').text)
                rdMaiorsCPF1.append(x9.find('codigoBeneficiario').text)
            #     #9.8.25
        for child10 in root10:
            for x10 in root10.findall(child10.tag+"/*"):
                rerCNS.append(x10.find('cco').text)
                rerCNS1.append(x10.find('codigoBeneficiario').text)
            #     #9.8.26
        for child11 in root11:
            for x11 in root11.findall(child11.tag+"/*"):
                rerCPF.append(x11.find('cco').text)
                rerCPF1.append(x11.find('codigoBeneficiario').text)
            #     #9.8.27
        for child12 in root12:
            for x12 in root12.findall(child12.tag+"/*"):
                tNAOi.append(x12.find('cco').text)
                tNAOi1.append(x12.find('codigoBeneficiario').text)

        # start_time = time.time()
        for arquivo in lista_arquivo:

            if arquivo.startswith('ArqConf') and arquivo.endswith('.xlsx'):
                wb = load_workbook(diretorio + '\\' + arquivo)
                # planilha = wb.active
                planilha = wb['ORIGEM']

                contador = 1
                for celula in planilha.iter_rows(min_row=2):
                    contador += 1
                    col_cco = str(celula[0].value)
                    col_cartao = str(celula[25].value)
                    # print(col_cco)
                    if ((col_cco in dmNaoIndentificado) and (col_cartao in dmNaoIndentificado1)): 
                        planilha[f'AR{contador}'] = '1'
                        # print("achou")
                    if ((col_cco in dmNaoValidado) and (col_cartao in dmNaoValidado1)):
                        planilha[f'AS{contador}'] = '1'
                    if ((col_cco in dmNaoEncontrado) and (col_cartao in dmNaoEncontrado1)):
                        planilha[f'AT{contador}'] = '1'
                    if ((col_cco in dMenorNIdentificado) and (col_cartao in dMenorNIdentificado1)):
                        planilha[f'AU{contador}'] = '1'
                    if ((col_cco in dMenorNaoValidado) and (col_cartao in dMenorNaoValidado1)):
                        planilha[f'AV{contador}'] = '1'   
                    if ((col_cco in rtSemCPF) and (col_cartao in rtSemCPF1)):
                        planilha[f'AW{contador}'] = '1'
                    if ((col_cco in rCNSNaoPreenchidos) and (col_cartao in rCNSNaoPreenchidos1)):
                        planilha[f'AX{contador}'] = '1'
                    if ((col_cco in rCPFNaoPreenchidos) and (col_cartao in rCPFNaoPreenchidos1)):
                        planilha[f'AY{contador}'] = '1'
                    if ((col_cco in rdNascimentoD) and (col_cartao in rdNascimentoD1)):
                        planilha[f'AZ{contador}'] = '1'
                    if ((col_cco in rdMaiorsCPF) and (col_cartao in rdMaiorsCPF1)):
                        planilha[f'BA{contador}'] = '1'
                    if ((col_cco in rerCNS) and (col_cartao in rerCNS1)):
                        planilha[f'BB{contador}'] = '1'
                    if ((col_cco in rerCPF) and (col_cartao in rerCPF1)):
                        planilha[f'BC{contador}'] = '1'
                    if ((col_cco in tNAOi) and (col_cartao in tNAOi1)):
                        planilha[f'BD{contador}'] = '1'    

                wb.save(diretorio + '\\' + arquivo)
    except Exception as e:
        logging.error('| Ocorreu um erro: | 3')
        logging.exception(str(e))    