from openpyxl import load_workbook
import os
import logging
import xml.etree.ElementTree as ET
import time
def conf():
    try:
        tree = ET.parse("C:\\Users\\User\\Desktop\\206-BS-571\\diretório\\diretorio.xml")
        root2 = tree.getroot()
        for child2 in root2:
                for x2 in root2.findall(child2.tag):
                        confeop = x2.find('confeop').text
        diretorio = confeop
        lista_arquivo = os.listdir(diretorio)

        for arquivo in lista_arquivo:
            if arquivo.startswith('ArqConf') and arquivo.endswith('.xlsx'):
                wb = load_workbook(diretorio + '\\' + arquivo)
                # planilha = wb.active
                planilha = wb['ORIGEM']
                contador = 1
                for celula in planilha.iter_rows(min_row=2):
                    contador += 1
                    #9.8.15
                    if celula[43].value == None: 
                        planilha[f'AR{contador}'] = 'VALIDO'
                    #9.8.16
                    if celula[44].value == None:
                        planilha[f'AS{contador}'] = 'VALIDO'
        #9.8.17
                    if celula[45].value == None:
                        planilha[f'AT{contador}'] = 'VALIDO'
        #9.8.18
                    if celula[46].value == None:
                        planilha[f'AU{contador}'] = 'VALIDO'
        #9.8.19
                    if celula[47].value == None:
                        planilha[f'AV{contador}'] = 'VALIDO'
        #9.8.20
                    if celula[48].value == None:
                        planilha[f'AW{contador}'] = 'VALIDO'
        #9.8.21
                    if celula[49].value == None:
                        planilha[f'AX{contador}'] = 'VALIDO'
        #9.8.22
                    if celula[50].value == None:
                        planilha[f'AY{contador}'] = 'VALIDO'
        #9.8.23
                    if celula[51].value == None:
                        planilha[f'AZ{contador}'] = 'VALIDO'
        #9.8.24
                    if celula[52].value == None:
                        planilha[f'BA{contador}'] = 'VALIDO'
        #9.8.25
                    if celula[53].value == None:
                        planilha[f'BB{contador}'] = 'VALIDO'
        #9.8.26
                    if celula[54].value == None:
                        planilha[f'BC{contador}'] = 'VALIDO'
        #9.8.27
                    if celula[55].value == None:
                        planilha[f'BD{contador}'] = 'VALIDO'

                wb.save(diretorio + '\\' + arquivo)
    except Exception as e:
        logging.error('| Ocorreu um erro: | 3')
        logging.exception(str(e))      