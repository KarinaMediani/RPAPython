from openpyxl import load_workbook
import os
import logging
import xml.etree.ElementTree as ET
#   4
def editarCabecalho():
    try:
        tree = ET.parse("C:\\Users\\User\\Desktop\\206-BS-571\\diretório\\diretorio.xml")
        root2 = tree.getroot()
        for child2 in root2:
                for x2 in root2.findall(child2.tag):
                        confeop = x2.find('confeop').text
        dir_ = confeop
        lista_dir = os.listdir(dir_)
        for arquivo in lista_dir:
            if arquivo.startswith('ArqConf') and arquivo.endswith('.xlsx'):
                wb = load_workbook(dir_+'\\'+arquivo)

                ws = wb['Sheet1']

                ws.insert_cols(5)
                ws.insert_cols(7)
                ws.insert_cols(10)
                ws.insert_cols(12)
                ws.insert_cols(13)
                ws.insert_cols(16)
                ws.insert_cols(27)
                ws.insert_cols(28)
                ws.insert_cols(29)
                ws['A1'] = 'CCO'
                ws['B1'] = 'STATUS'
                ws['C1'] = 'DT_ATUAL'
                ws['D1'] = 'CPF'
                ws['E1'] = 'NUM_CPF'
                ws['F1'] = 'NOME'
                ws['G1'] = 'ABREV_NOME'
                ws['H1'] = 'SEXO'
                ws['I1'] = 'DT_NASC'
                ws['J1'] = 'IDADE'
                ws['K1'] = 'MAE'
                ws['L1'] = 'ABREV_MAE'
                ws['M1'] = 'NOME_MAE'
                ws['N1'] = 'PIS'
                ws['O1'] = 'CNS'
                ws['P1'] = 'TIPO_CNS'
                ws['Q1'] = 'LOGRADOURO'
                ws['R1'] = 'NÚMERO'
                ws['S1'] = 'COMPLEMENTO'
                ws['T1'] = 'BAIRRO'
                ws['U1'] = 'IBGE'
                ws['V1'] = 'CEP'
                ws['W1'] = 'IND_RES'
                ws['X1'] = 'IBGE_RES'
                ws['Y1'] = 'TIPO_END'
                ws['Z1'] = 'CARTAO'
                ws['AA1'] = 'TITULARIDADE'
                ws['AB1'] = 'APOLICE'
                ws['AC1'] = 'RAMO'
                ws['AD1'] = 'VÍNCULO'
                ws['AE1'] = 'DT_CONTRAT'
                ws['AF1'] = 'DT_CANCEL'
                ws['AG1'] = 'DT_REATIV'
                ws['AH1'] = 'MOTIVO_CANCEL'
                ws['AI1'] = 'CD_PL_MS'
                ws['AJ1'] = 'IND_CPT'
                ws['AK1'] = 'IND_PROC_EXCL'
                ws['AL1'] = 'CNPJ'
                ws['AM1'] = 'CCO_TIT'
                ws['AN1'] = 'CEI'
                ws['AO1'] = 'PL_ANT_LEI'
                ws['AP1'] = 'PL_MS_PORT'
                ws['AQ1'] = 'CAEPF'
                ws.insert_cols(44)
                ws.insert_cols(45)
                ws.insert_cols(46)
                ws.insert_cols(47)
                ws.insert_cols(48)
                ws.insert_cols(49)
                ws.insert_cols(50)
                ws.insert_cols(51)
                ws.insert_cols(52)
                ws.insert_cols(53)
                ws.insert_cols(54)
                ws.insert_cols(55)
                ws.insert_cols(56)
                ws['AR1'] = 'DEPENDENTE MAIOR NÃO IDENTIFICADO'
                ws['AS1'] = 'DEPENDENTE MAIOR NÃO VALIDADO'
                ws['AT1'] = 'DEPENDENTE MENOR NÃO ENCONTRADO'
                ws['AU1'] = 'DEPENDENTE MENOR NÃO IDENTIFICADO'
                ws['AV1'] = 'DEPENDENTE MENOR NÃO VALIDADO'
                ws['AW1'] = 'REGISTRO DE TITULAR SEM CPF'
                ws['AX1'] = 'REGISTROS COM CNS NÃO PREENCHIDOS'
                ws['AY1'] = 'REGISTROS COM CPF NÃO PREENCHIDOS'
                ws['AZ1'] = 'REGISTROS COM DATA DE NASCIMENTO DIVERGENTE'
                ws['BA1'] = 'REGISTROS DE DEPENDENTE MAIOR SEM CPF'
                ws['BB1'] = 'REGISTROS ENVOLVIDOS NAS RESTRIÇÕES DE CNS'
                ws['BC1'] = 'REGISTROS ENVOLVIDOS NAS REPETIÇÕES DE CPF'
                ws['BD1'] = 'TITULAR NÃO IDENTIFICADO'



                wb.save(dir_+'\\'+arquivo )
    except Exception as e:
                logging.error('| Ocorreu um erro: | 3')
                logging.exception(str(e))    