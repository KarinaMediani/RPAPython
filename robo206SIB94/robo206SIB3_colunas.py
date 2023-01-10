import os
from openpyxl import load_workbook
import logging
import xml.etree.ElementTree as ET
import pandas as pd
### 3
def separandoColunas():
    try:
        tree = ET.parse("C:\\Users\\User\\Desktop\\206-BS-571\\diretório\\diretorio.xml")
        root2 = tree.getroot()
        for child2 in root2:
                for x2 in root2.findall(child2.tag):
                        confeop = x2.find('confeop').text
        diretorio = confeop
        lista_arquivos = os.listdir(diretorio)

        for arquivo in lista_arquivos:
            if arquivo.startswith('ArqConf') and arquivo.endswith('.xlsx'):

                wb = load_workbook(diretorio + '\\' + arquivo)
                ws = wb.active
                # ws.delete_rows(0)
                ws.delete_rows(1)
                contador = 0
                for coluna in ws:
                    contador += 1
                    linha = str(coluna[0].value)
                    ws[f'B{contador}'] = linha[0:12]
                    ws[f'C{contador}'] = linha[12:19]
                    ws[f'D{contador}'] = linha[19:29]
                    ws[f'E{contador}'] = linha[29:40]
                    ws[f'F{contador}'] = linha[40:110]
                    ws[f'G{contador}'] = linha[110:111]
                    ws[f'H{contador}'] = linha[111:121]
                    ws[f'I{contador}'] = linha[121:191]
                    ws[f'J{contador}'] = linha[191:202]
                    ws[f'K{contador}'] = linha[202:217]
                    ws[f'L{contador}'] = linha[217:267]
                    ws[f'M{contador}'] = linha[267:272]
                    ws[f'N{contador}'] = linha[272:287]
                    ws[f'O{contador}'] = linha[287:317]
                    ws[f'P{contador}'] = linha[317:323]
                    ws[f'Q{contador}'] = linha[323:331]
                    ws[f'R{contador}'] = linha[331:332]
                    ws[f'S{contador}'] = linha[332:338]
                    ws[f'T{contador}'] = linha[338:339]
                    ws[f'U{contador}'] = linha[339:369]
                    ws[f'V{contador}'] = linha[369:371]
                    ws[f'W{contador}'] = linha[371:381]
                    ws[f'X{contador}'] = linha[381:391]
                    ws[f'Y{contador}'] = linha[391:401]
                    ws[f'Z{contador}'] = linha[401:403]
                    ws[f'AA{contador}'] = linha[403:412]
                    ws[f'AB{contador}'] = linha[412:413]
                    ws[f'AC{contador}'] = linha[413:414]
                    ws[f'AD{contador}'] = linha[414:428]
                    ws[f'AE{contador}'] = linha[428:440]
                    ws[f'AF{contador}'] = linha[440:452]
                    ws[f'AG{contador}'] = linha[452:472]
                    ws[f'AH{contador}'] = linha[472:481]
                    ws[f'AI{contador}'] = linha[481:]
                wb.save(diretorio+'\\'+arquivo)
                wb = load_workbook(diretorio + '\\' + arquivo)
                ws = wb.active
                ws.delete_cols(1)
                wb.save(diretorio+'\\'+arquivo)
    except Exception as e:
                logging.error('| Ocorreu um erro: | 3')
                logging.exception(str(e))    