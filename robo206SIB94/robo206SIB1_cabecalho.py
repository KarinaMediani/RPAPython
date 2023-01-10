import os
from openpyxl import load_workbook
import logging
import xml.etree.ElementTree as ET
#   1
def parametro():
    try:
        tree = ET.parse("C:\\Users\\User\\Desktop\\206-BS-571\\diretório\\diretorio.xml")
        root2 = tree.getroot()
        for child2 in root2:
                for x2 in root2.findall(child2.tag):
                        confeop = x2.find('confeop').text
        diretorio = confeop
        lista_arquivo = os.listdir(diretorio)

        for arquivo in lista_arquivo:
            if arquivo.startswith('ArqConf') and arquivo.endswith('.xlsx'):
                wb = load_workbook(diretorio+'\\'+arquivo)
                ws = wb.active
                
                ws.insert_rows(0)

                ws['A1']= 'EXCLUIR'

                wb.save(diretorio+'\\'+arquivo)
    except Exception as e:
                logging.error('| Ocorreu um erro: | 3')
                logging.exception(str(e))    