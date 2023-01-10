from openpyxl import load_workbook
import os
import logging
import xml.etree.ElementTree as ET
import datetime
import locale
import time
import regex as re

 
def rpa206_975():
    try:
        # start_time = time.time()
        tree = ET.parse("C:\\Users\\User\\Desktop\\206-BS-571\\diretório\\diretorio.xml")
        root2 = tree.getroot()
        for child2 in root2:
                for x2 in root2.findall(child2.tag):
                        suop= x2.find('suop').text
                        confeop = x2.find('confeop').text
                        su05listaop= x2.find('su05listaop').text
        data_atual = datetime.datetime.now()
        data = data_atual.date()
        locale.setlocale(locale.LC_ALL, '')
        ano = data.strftime("%Y")
        mesAno = data.strftime("%m%Y")
        anoMes = data.strftime("%Y%m")
        #Abertura de arquivo SU05:
        dir_su05 =suop
        list_dir1 = os.listdir(dir_su05)
        lista874 = []
        lista875 = []
        lista876 = []
        lista878 = []
        for arquivo_txt in list_dir1:
            if arquivo_txt.startswith('SU05') and arquivo_txt.endswith('.txt'):
                ramo = str 
                file_text = open((dir_su05+'\\'+arquivo_txt),"r",encoding= "latin-1")
                # # file_text = open((dir_su05+'\\'+arquivo_txt),"r",encoding= "utf-8")
                arquivo = file_text.read()
                # arquivo = re.sub('[^A-z\s\d/.:_\\,+;@&*()~ÁÉÍÓÚ]?', '', arquivo)#tirando simbolos que o python nao aceita
                lista_dados = arquivo.splitlines()
                for i in range(0,len(lista_dados)):
                    lista_dados[i] = lista_dados[i].split(None)
                # print(lista_dados)
                for listas in lista_dados[:]:#retirando dados que nao ira ser utilizados para analise
                    if ((len(listas)<7) or (len(listas) == 17) ):
                        # print(len(listas))
                        lista_dados.remove(listas)
                    elif (("ESTIPULANTE" in listas )and("STATUS" in listas )and("SUC" in listas )and("CIA" in listas )and("AG.PRO" in listas )and("FATURAMENTO" in listas )and("VIGENCIA" in listas )and("QTE.CLI" in listas )):
                        lista_dados.remove(listas)
                # print(lista_dados)
                for li in lista_dados[:]:#pegando dados por ramo de uma lista ja tratada 
                    if (("874" in li[:][8]) or ("875" in li[:][8]) or ("876" in li[:][8]) or ("878" in li[:][8])):
                        ramo = li[:][8]#pegando os ramos
                    if ramo == "874":#adicionando dados para listas especificas de cada ramo
                        lista874.append(li[:][0])
                    elif ramo == "875":
                        lista875.append(li[:][0])
                    elif ramo == "876":
                        lista876.append(li[:][0])
                    elif ramo == "878":
                        lista878.append(li[:][0])
        # for l874 in lista874[:]:
        #     if (("CPD" in l874) or ('BRADESCO' in l874)):
        #         lista874.remove(l874)        
        # for l875 in lista875[:]:
        #     if (("CPD" in l875) or ('BRADESCO' in l875)):
        #         lista875.remove(l875)
        # for l876 in lista876[:]:
        #     if (("CPD" in l876) or ('BRADESCO' in l876)):
        #         lista876.remove(l876)
        # for l878 in lista878[:]:
        #     if (("CPD" in l878) or ('BRADESCO' in l878)):
        #         lista878.remove(l878)
        l874 = []
        l875 = []
        l876 = []
        l878 = []
        for item in lista874[:]:
            lista8744 = item.replace(".", "")
            l874.append(lista8744)
            # print(type(lista874))
        for item in lista875[:]:
            lista8755 = item.replace(".", "")
            l875.append(lista8755)
        for item in lista876[:]:
            lista8766 = item.replace(".", "")
            l876.append(lista8766)
        for item in lista878[:]:
            lista8788 = item.replace(".", "")
            l878.append(lista8788)
        # print(l874)
        # print(l875)
        # print(l876)
        # print(l878)

        #----------------------------------------------------------------------------------------------#
        excessões_su05 = su05listaop #Abrindo arquivo lista de excessões SU05
        wb1 = load_workbook(excessões_su05)# TODO passar caminho por completo até chegar ao arquivo 
        ws1 = wb1['SU05']
        excessao = []


        #     # lista_arquivo = os.listdir(dir_su05xlsx)

            #Abertura planilha CIA 
        dir_cia571 =  confeop
        lista_arquivocia = os.listdir(dir_cia571)
        for arquivo in lista_arquivocia:
            if arquivo.startswith('ArqConf') and arquivo.endswith('.xlsx'):#Abrindo arquivos conferencia
                wb2 = load_workbook(dir_cia571 + '\\' + arquivo)
                # ws2 = wb2.active
                ws2 = wb2['ORIGEM'] #abrindo aba do arquivo conferencia
                contador = 1  
                for item in ws2.iter_rows(min_row=2):            
                    apolice = str(item[27].value)#pegando coluna AB para analise
                    # print(apolice)
                    contador += 1
                                #Localização de contratos ###9.7.7          
                    # print(cel_excessao)
                    encontrado = False        
                                #COMPARAR VALORES DA COLUNA DA PLANILHA CIA COM SU05
                    if item[27].value == None: # se na coluna 27 do arquivo conferencia nao existir dados preencher a coluna AC com ramo 870
                        ws2[f'AC{contador}'] = '870'
                        continue
                    elif apolice in l874:# se na coluna 27 do arquivo conferencia o valor da celula existir dentro da lista874 preencher a coluna AC com ramo 874
                        ws2[f'AC{contador}'] = '874'
                        # print("ok")
                        continue
                    elif apolice in l875:# se na coluna 27 do arquivo conferencia o valor da celula existir dentro da lista874 preencher a coluna AC com ramo 875
                        ws2[f'AC{contador}'] = '875'
                        continue
                    elif apolice in l876:# se na coluna 27 do arquivo conferencia o valor da celula existir dentro da lista874 preencher a coluna AC com ramo 876
                        ws2[f'AC{contador}'] = '876'
                        continue
                    elif apolice in l878:# se na coluna 27 do arquivo conferencia o valor da celula existir dentro da lista874 preencher a coluna AC com ramo 878
                        ws2[f'AC{contador}'] = '878'
                        continue
                    for su05 in ws1:
                        if apolice == str(su05[0].value):# se o valor da coluna AB do arquivo conferencia exisitr na coluna A do arquivo SU05
                            encontrado = True
                            ws2[f'AC{contador}'] = str(su05[1].value)
                            break
                    if encontrado == False:
                        ws2[f'AC{contador}'] = 'NÃO ENCONTRADO'
                    else:
                        encontrado = False
                wb2.save(dir_cia571 + '//' + arquivo)
        # end_time = time.time()
        # print(end_time-start_time)

#             # for listinha in listas:
#                 # print(listinha)
#             # for c in 
#             #     novalista = list(range(listinha(0,5)))
                    
#             # elif lista_dados == '875':
#             #     lista875.append(lista_dados)
#             # elif lista_dados == '876':
#             #     lista876.append(lista_dados)
#             # elif lista_dados == '878':
#             #     lista878.append(lista_dados)
# # print(novalista)

#         #         wb = Workbook()
#         #         ws = wb.active

#         #         for row in lista_dados:
#         #             ws.append(row)

#         #         wb.save((dir_su05+'\\'+arquivo_txt)+ '.xlsx') 
#         # # end_time = time.time()
#         # # print(end_time-start_time)

    except Exception as e:
        logging.error('| Ocorreu um erro: | 3')
        logging.exception(str(e))