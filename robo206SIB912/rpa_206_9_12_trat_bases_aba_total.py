from openpyxl import load_workbook
import logging
import xml.etree.ElementTree as ET
def prodAtivos_X_Apuracao():
    try:
        tree = ET.parse("C:\\Users\\User\\Desktop\\206-BS-571\\diretório\\diretorio.xml")
        root2 = tree.getroot()
        for child2 in root2:
                for x2 in root2.findall(child2.tag):
                        proativosop = x2.find('proativosop').text
                        apcia = x2.find('apcia').text


        dir_apuracao = apcia

        diretorio = proativosop #PRODUTOS ATIVOS


        wb = load_workbook(diretorio)
        ws = wb['Total']
        #TODO - LISTA APURAÇÃO BRADESCO 


        wb1 = load_workbook(dir_apuracao)
        ws1 = wb1.active

        valor1 = ws1['B3'].value
        valor2 = ws1['B7'].value
        valor3 = ws1['G10'].value
        valor4 = ws1['N10'].value
        valor5 = ws1['B7'].value
        valor6 = ws1['G11'].value
        valor7 = ws1['N11'].value
        valor8 = ws1['B4'].value
        # print(valor1 , valor2 , valor3 , valor4)

        if (valor3 ==None) and (valor4 ==None):
            ws1['B5'] = '0'
            ws1['B8'] = '0'
            ws1['B12'] = '0'
        else:
            soma = (valor3 +  valor4) #G10 + N10

        if (valor6 ==None) and (valor7 ==None):
            ws1['B11'] = '0'
        else:    
            soma2 = (valor6 + valor7) #G11 + N11
        # print(soma)
        ws['B3'] = valor1 #verificado! testado??
        ws['B5'] = valor1 #verificado! testado??
        ws['B6'] = valor2 #verificado! testado ??
        # ws['B5'] = soma
        ws['B9'] = valor5 #verificado! testado ??
        ws['B11'] = soma2 #verificado! testado ??
        ws['B12'] = soma #verificado! testado ??
        ws['B14'] = valor8 #verificado! testado ??
        ws['B8'] = soma #verificado! testado ??

        wb.save(diretorio)
    except Exception as e:
                logging.error('| Ocorreu um erro: | 3')
                logging.exception(str(e))    



