import os
from openpyxl import load_workbook
import logging
from typing import Counter
import xml.etree.ElementTree as ET
from openpyxl.styles import Font
import time
def conf():
    try:
        tree = ET.parse("C:\\Users\\User\\Desktop\\206-BS-571\\diretório\\diretorio.xml")
        root2 = tree.getroot()
        for child2 in root2:
            for x2 in root2.findall(child2.tag):
                    confeop = x2.find('confeop').text
        lista_Cd_Pl_Ms = []
        # lista_Ant_Lei = []
        dir_apuracao = confeop
        lista_apuracao = os.listdir(dir_apuracao)
        # print(lista_apuracao)
        for arquivo in lista_apuracao:
            if arquivo.startswith('ArqConf') and arquivo.endswith('.xlsx'):
                wb = load_workbook(dir_apuracao + '\\' + arquivo)
                ws = wb['ORIGEM']
                ws1 = wb['PRODUTOS ATIVOS ']
            # #9.9.5 A 
            # #TABELA DINAMICA 


                for item in ws.iter_rows(min_row=2):
                    Cd_Pl_Ms = str(item[34].value).strip()   ## Criando dicionario com os valores da coluna 34
                    # print(Cd_Pl_Ms) 
                    if Cd_Pl_Ms == "":
                        continue
                    else:           
                        lista_Cd_Pl_Ms.append(Cd_Pl_Ms)
                    Ant_Lei = str(item[40].value).strip()     ## Criando dicionario com os valores da coluna 41 
                    if Ant_Lei == "":
                        continue
                    else:
                        lista_Cd_Pl_Ms.append(Ant_Lei)         
                    # lista_Cd_Pl_Ms.append(Ant_Lei)
                        # print(lista_Cd_Pl_Ms)
                duplicados = dict(Counter(lista_Cd_Pl_Ms))
                # print(duplicados)
                contador2 = 0

                for key, value in duplicados.items() :
                    contador2 += 1
                    ws1[f'A{contador2}'] = key
                    ws1[f'B{contador2}'] = value
                    ws1[f'A{contador2}'].font = Font(bold= True)


                soma_total = sum(duplicados.values())
                ws1['C1'] = soma_total   


                ws1.insert_rows(0)
                ws1['A1'] = 'PRODUTO'
                ws1['B1'] = 'VIDAS'
                ws1['C1'] = 'TOTAL DE VIDAS'

                
                wb.save(dir_apuracao + '\\' + arquivo)

    except Exception as e:
        logging.error('| Ocorreu um erro: | 3')
        logging.exception(str(e))  