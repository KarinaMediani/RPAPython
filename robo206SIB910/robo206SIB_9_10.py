import os
from openpyxl import load_workbook
import logging
import xml.etree.ElementTree as ET
import pandas as pd 
def conf():
    try:
        tree = ET.parse("C:\\Users\\User\\Desktop\\206-BS-571\\diretório\\diretorio.xml")
        root2 = tree.getroot()
        resultado = None
        for child2 in root2:
                for x2 in root2.findall(child2.tag):
                        confeop = x2.find('confeop').text
                        apcia = x2.find('apcia').text
        dir_1 = confeop
        list_dir1 = os.listdir(dir_1) ##diretorio arquivos principais###

        dir_2 = apcia

        wb = load_workbook(dir_2)
        ws = wb.active

        contador_vazias = 0
        contador_final1 = 0
        contador_final2 = 0
        contador_final3 = 0 
        contador_final4 = 0
        contador_final5 = 0
        contador_final6 = 0
        contador_final7 = 0
        contador_final8 = 0
        contador_final9 = 0
        contador_final10 = 0
        # lista_col_b = []
        # celula_b3 = None
        lista_col_A = []
        celula_B4 = None
        contador_colunaB = 0

        for arquivo in list_dir1:
            if arquivo.startswith('ArqConf') and arquivo.endswith('.xlsx'):
                # print(arquivo)
                contador = 0
                wb1 = load_workbook(dir_1 + '\\' + arquivo)
                ws1 = wb1['ORIGEM']
                ws2 = wb1['PRODUTOS ATIVOS ']

                for linhas in ws2.iter_rows(min_row=2):
                    colunaA = str(linhas[0].value)
                    lista_col_A.append(colunaA)
                    celula_B4 = (len(lista_col_A))

            #9.10.4 - subtração 
                valor1_ = ws['B1'].value
                if valor1_ == None:
                    valor1_ = 0
                # print(valor1)

                valor2_ = ws['B3'].value
                if valor2_ == None:
                    valor2_ = 0
                # print(valor2)
                
                
                for linha in ws.iter_rows(min_row= 2):  
                    contador += 1
                                
                    valorAI = ws1[f'AI{contador}'].value 
                    valorAO = ws1[f'AO{contador}'].value
                
                for linha in ws1.iter_rows(min_row= 2):   
                    valorJ = str(linha[9].value).upper().strip()
                    valorAC = str(linha[28].value).upper().strip()
                    valorAA = str(linha[26].value).upper().strip()
                    valorB = str(linha[1].value).upper().strip()
                    
                    if valorB == "ATIVO":
                        contador_colunaB = contador_colunaB + 1
        #             lista_col_b.append(contador_colunaB)
        # print(lista_col_b)

                    if (valorAI == None) and (valorAO == None):
                        contador_vazias = contador_vazias + 1
                    
                    contador += 1           

                    if ((valorJ == 'MAIOR')and (valorAC == '870')and(valorAA == 'TITULAR')):
                        contador_final1 = contador_final1 + 1
                    #resultado_linhas_B10 
                    
                    if ((valorJ == 'MAIOR')and (valorAC == '874')and(valorAA == 'TITULAR')):
                        contador_final2 = contador_final2 + 1
                    #resultado_linhasC10
                    
                    if ((valorJ == 'MAIOR')and (valorAC == '875')and(valorAA == 'TITULAR')):
                        contador_final3 = contador_final3 + 1
            #     #resultado_linhasD10

                    if ((valorJ == 'MAIOR')and (valorAC == '876')and(valorAA == 'TITULAR')):
                        contador_final4 = contador_final4 + 1
            #          #  resultado_linhasE10         
                
                    if ((valorJ == 'MAIOR')and (valorAC == '878')and(valorAA == 'TITULAR')):
                        contador_final5 = contador_final5 + 1
            #         #resultado_linhasF10
                    
                    if ((valorJ == 'MENOR')and (valorAC == '870')and(valorAA == 'TITULAR')):
                        contador_final6 = contador_final6 + 1
            #           # resultado_linhasB11         

                    if ((valorJ == 'MENOR')and (valorAC == '874')and(valorAA == 'TITULAR')):
                        contador_final7 = contador_final7 + 1
            #          #resultado_linhasC11           

                    if ((valorJ == 'MENOR')and (valorAC == '875')and(valorAA == 'TITULAR')):
                        contador_final8 = contador_final8 + 1
            #    #resultado_linhasD11  
                    if ((valorJ == 'MENOR')and (valorAC == '876')and(valorAA == 'TITULAR')):
                        contador_final9 = contador_final9 + 1
            #    #resultado_linhasE11  
                    if ((valorJ == 'MENOR')and (valorAC == '878')and(valorAA == 'TITULAR')):
                        contador_final10 = contador_final10 + 1
            #     ##resultado_linhas_F11
        # celula_b3 = sum(lista_col_b)

        subtracao = (valor1_ - valor2_)      
        resultado_total = contador_colunaB #resultado B3 apuração 
        # resultadoD = list_sumD #resultrado 10.3
        resultado_total1 = (celula_B4)
        resultado_vazias  = contador_vazias   #9.10.5
        resultado_linhas_B10 = contador_final1#resultado 9.10.6 B10
        resultado_linhasC10 = contador_final2 #RESULTADO 9.10.7 c10
        resultado_linhasD10 = contador_final3 #RESULTADO 9.10.8 D10
        resultado_linhasE10 = contador_final4 #RESULTADO 9.10.9 E10
        resultado_linhasF10 = contador_final5 #RESULTADO 9.10.10F10 
        resultado_linhasB11 = contador_final6 #RESULTADO 9.10.11 B11
        resultado_linhasC11 = contador_final7 #RESULTADO 9.10.12 C11
        resultado_linhasD11 = contador_final8 #RESULTADO 9.10.12 D11
        resultado_linhasE11 = contador_final9 #RESULTADO 9.10.13 E11
        resultado_linhasF11 = contador_final10 #RESULTADO 9.10.15

        ws['B3'] = resultado_total #produtos ativos #TODO OK
        ws['B4'] = resultado_total1 #resultado 9.10.3
        ws['B6'] = subtracao #resultado 9.10.4
        ws['B7'] = resultado_vazias #resultado 9.10.5
        ws['B10'] = resultado_linhas_B10 #B10 9.10.6
        ws['C10'] = resultado_linhasC10 #resultado 9.10.7 C10
        ws['D10'] = resultado_linhasD10 #RESULTADO 9.10.8 D10
        ws['E10'] = resultado_linhasE10 #RESULTADO 9.10 9 E10
        ws['F10'] = resultado_linhasF10 #RESULTADO 9.10.10 F10
        ws['B11'] = resultado_linhasB11 #RESULTADO 9.10.11 B11
        ws['C11'] = resultado_linhasC11 #RESULTADO 9.10.12 C11
        ws['D11'] = resultado_linhasD11#RESULTADO 9.10.13 D11
        ws['E11'] = resultado_linhasE11 #RESULTADO 9.10.14 E11
        ws['F11'] = resultado_linhasF11 #RESULTADO 9.10.15 F11


        wb.save(dir_2)



    except Exception as e:
        logging.error('| Ocorreu um erro: | 3')
        logging.exception(str(e))  