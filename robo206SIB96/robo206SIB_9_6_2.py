import os
from openpyxl import load_workbook

import time
import logging
import xml.etree.ElementTree as ET

def conf():
    try:
        tree = ET.parse("C:\\Users\\User\\Desktop\\206-BS-571\\diretório\\diretorio.xml")
        root2 = tree.getroot()
        for child2 in root2:
                for x2 in root2.findall(child2.tag):
                        confeop = x2.find('confeop').text
        diretorio = (confeop)
        lista_arquivo = os.listdir(diretorio)
        # start_time = time.time()

        a = 'A'
        b = 'B'
        c = 'C'
        d = 'D'
        f = 'F'
        g = 'G'
        h = 'H'
        i = 'I'
        j = 'J'
        k = 'K'
        l = 'L'
        m = 'M'
        n = 'N'
        o = 'O'
        p = 'P'
        q = 'Q'
        r = 'R'
        s = 'S'
        t = 'T'
        u = 'U'
        v = 'V'
        w = 'W'
        x = 'X'
        z = 'Z'
        conec = 'E'
        conec2 = 'Y'

        #9.6.15 e 6.16
        for arquivo in lista_arquivo:
            # print(arquivo)
            if arquivo.startswith('ArqConf') and arquivo.endswith('.xlsx'):
                wb = load_workbook(diretorio + '\\' + arquivo)

                ws = wb['ORIGEM']
                ## INSERÇÃO DE NOVAS COLUNAS 

                
                # ANALISE DA COLUNA K 
                contador = 1
                for celula in ws.iter_rows(min_row=2):
                    contador += 1
                #TODO FEITO!
                    item = str(celula[10].value).strip().upper().split()
                    s_nome = item[1:]
                    c_nome = item[:1]
                    if a in item:
                        ws[f'L{contador}'] = 'SIM'
                    elif b in item:
                        ws[f'L{contador}'] = 'SIM'
                    elif c in item:
                        ws[f'L{contador}'] = 'SIM'
                    elif d in item:
                        ws[f'L{contador}'] = 'SIM'
                    elif f in item:
                        ws[f'L{contador}'] = 'SIM'
                    elif g in item:
                        ws[f'L{contador}'] = 'SIM'
                    elif h in item:
                        ws[f'L{contador}'] = 'SIM'
                    elif i in item:
                        ws[f'L{contador}'] = 'SIM'
                    elif j in item:
                        ws[f'L{contador}'] = 'SIM'
                    elif k in item:
                        ws[f'L{contador}'] = 'SIM'
                    elif l in item:
                        ws[f'L{contador}'] = 'SIM'
                    elif m in item:
                        ws[f'L{contador}'] = 'SIM'
                    elif n in item:
                        ws[f'L{contador}'] = 'SIM'
                    elif o in item:
                        ws[f'L{contador}'] = 'SIM'
                    elif p in item:
                        ws[f'L{contador}'] = 'SIM'
                    elif q in item:
                        ws[f'L{contador}'] = 'SIM'
                    elif r in item:
                        ws[f'L{contador}'] = 'SIM'
                    elif s in item:
                        ws[f'L{contador}'] = 'SIM'
                    elif t in item:
                        ws[f'L{contador}'] = 'SIM'
                    elif u in item:
                        ws[f'L{contador}'] = 'SIM'
                    elif v in item:
                        ws[f'L{contador}'] = 'SIM'
                    elif w in item:
                        ws[f'L{contador}'] = 'SIM'
                    elif x in item:
                        ws[f'L{contador}'] = 'SIM'
                    elif z in item:
                        ws[f'L{contador}'] = 'SIM'           
                    elif a in c_nome:
                        ws[f'L{contador}'] = 'INVALIDO'
                    elif b in c_nome:
                        ws[f'L{contador}'] = 'INVALIDO'
                    elif c in c_nome:
                        ws[f'L{contador}'] = 'INVALIDO'
                    elif d in c_nome:
                        ws[f'L{contador}'] = 'INVALIDO'
                    elif f in c_nome:
                        ws[f'L{contador}'] = 'INVALIDO'
                    elif g in c_nome:
                        ws[f'L{contador}'] = 'INVALIDO'
                    elif h in c_nome:
                        ws[f'L{contador}'] = 'INVALIDO'
                    elif i in c_nome:
                        ws[f'L{contador}'] = 'INVALIDO'
                    elif j in c_nome:
                        ws[f'L{contador}'] = 'INVALIDO'
                    elif k in c_nome:
                        ws[f'L{contador}'] = 'INVALIDO'
                    elif l in c_nome:
                        ws[f'L{contador}'] = 'INVALIDO'
                    elif m in c_nome:
                        ws[f'L{contador}'] = 'INVALIDO'
                    elif n in c_nome:
                        ws[f'L{contador}'] = 'INVALIDO'
                    elif o in c_nome:
                        ws[f'L{contador}'] = 'INVALIDO'
                    elif p in c_nome:
                        ws[f'L{contador}'] = 'INVALIDO'
                    elif q in c_nome:
                        ws[f'L{contador}'] = 'INVALIDO'
                    elif r in c_nome:
                        ws[f'L{contador}'] = 'INVALIDO'
                    elif s in c_nome:
                        ws[f'L{contador}'] = 'INVALIDO'
                    elif t in c_nome:
                        ws[f'L{contador}'] = 'INVALIDO'
                    elif u in c_nome:
                        ws[f'L{contador}'] = 'INVALIDO'
                    elif v in c_nome:
                        ws[f'L{contador}'] = 'INVALIDO'
                    elif w in c_nome:
                        ws[f'L{contador}'] = 'INVALIDO'
                    elif x in c_nome:
                        ws[f'L{contador}'] = 'INVALIDO'
                    elif z in c_nome:
                        ws[f'L{contador}'] = 'INVALIDO'
                    elif conec2 in item:
                        ws[f'L{contador}'] = 'CONECTIVO'
                    elif conec in item:
                        ws[f'L{contador}'] = 'CONECTIVO'
                    else:
                        ws[f'L{contador}'] = 'NAO'
                        
                    vazio = str(celula[10].value).strip()
                    # print(vazio)
                    if vazio == "":
                        # print("vazio")
                        ws[f'M{contador}'] = "VAZIO"

                    
                    ## ANALISE CNS 
                    item1 = str(celula[14].value)
                    primeiro = "7"
                    segundo = "8"
                    vazio = ""
                    if item1.find(primeiro) == 0:
                        ws[f'P{contador}'] = 'DEFINITIVO'
                    elif item1.find(segundo) == 0:
                        ws[f'P{contador}'] = 'CNS TEMPORARIO'
                    elif item1.find(vazio) == 0:
                        ws[f'P{contador}'] = 'SEM CNS'
                    
                    
                    ## ANALISE TITULAR DEPENDENTE OU TITULAR     
                    item = str(celula[25].value).strip().upper()
                    zero1 = "0"

                    if len(item) == 15:
                        if ((item.rfind(zero1) == 13 and 14)) :
                            # print(item)
                            ws[f'AA{contador}'] = "TITULAR"
                        elif item.rfind(zero1) != 14 and 15:
                            ws[f'AA{contador}'] = "DEPENDENTE"
                        elif item[12:] == "000":
                            ws[f'AA{contador}'] = "TITULAR"
                            # print(item)
                    # #VERIFICAÇÃO NUMERO DO CARTÃO       
                    primeiro = "7"
                    segundo = "8"
                    terceiro = "9"
                    # print(item1)
                    if item.find(primeiro) == 0:
                        ws[f'AB{contador}'] = item[1:6]
                    elif item.find(segundo) == 0:
                        ws[f'AB{contador}'] = item[1:7]
                    elif item.find(terceiro) == 0:
                        ws[f'AB{contador}'] = item[1:5]
                    else:
                        ws[f'AB{contador}'] = None


                wb.save(diretorio + '\\' + arquivo)
        # end_time = time.time()
        # print(end_time-start_time)
    except Exception as e:
        logging.error('| Ocorreu um erro: | 3')
        logging.exception(str(e))  