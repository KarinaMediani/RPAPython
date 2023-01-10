import os
from openpyxl import load_workbook
import datetime
import time
import logging
import xml.etree.ElementTree as ET

def conf():
    try:
        tree = ET.parse("C:\\Users\\User\\Desktop\\206-BS-571\\diretório\\diretorio.xml")
        root2 = tree.getroot()
        for child2 in root2:
                for x2 in root2.findall(child2.tag):
                        confeop = x2.find('confeop').text

        diretorio = (confeop)

        lista_arquivo = os.listdir(diretorio)
        # start_time = time.time()
        for arquivo in lista_arquivo: 
            # print(arquivo)
            if arquivo.startswith('ArqConf') and arquivo.endswith('.xlsx'):
                data_atual = datetime.datetime.now()
                data = data_atual.date()
                data_br = data.strftime("%Y")
                ano_atual = float(data_br)
                wb = load_workbook(diretorio + '\\' + arquivo)
                ws = wb.active
                ws['A1'] = 'CCO'

                a = 'A'
                b = 'B'
                c = 'C'
                d = 'D'
                f = 'F' 
                g = 'G' 
                h = 'H'
                i = 'I'
                j = 'J'
                k = 'K'
                l = 'L'
                m = 'M'
                n = 'N'
                o = 'O'
                p = 'P' 
                q = 'Q'
                r = 'R'
                s = 'S'
                t = 'T'
                u = 'U'
                v = 'V'
                w = 'W'
                x = 'X'
                z = 'Z'
                conec = 'E'
                conec2 = 'Y'
                # #VERIFICANDO VALOR DA COLUNA NUM_CPF COLUNA D
                contador = 1
                for row in ws.iter_rows(min_row=2):

                    contador += 1
                    colunaD= str(row[3].value).strip()
                    #TODO FEITO!
                    if colunaD == "":
                        ws[f'E{contador}'] = 'VAZIO'
                        # print("ok")
                    else:
                        ws[f'E{contador}'] = 'VALIDO'
                


                        ## VERIFICANDO VALOR DA COLUNA NUM_CPF 
        #TODO FEITO!
                    item = str(row[5].value).strip().upper().split()
                    # print(item)
                    if a in item:
                        ws[f'G{contador}'] = 'SIM'
                    elif b in item:
                        ws[f'G{contador}'] = 'SIM'
                    elif c in item:
                        ws[f'G{contador}'] = 'SIM'
                    elif d in item:
                        ws[f'G{contador}'] = 'SIM'
                    elif f in item:
                        ws[f'G{contador}'] = 'SIM'
                    elif g in item:
                        ws[f'G{contador}'] = 'SIM'
                    elif h in item:
                        ws[f'G{contador}'] = 'SIM'
                    elif i in item:
                        ws[f'G{contador}'] = 'SIM'
                    elif j in item:
                        ws[f'G{contador}'] = 'SIM'
                    elif k in item:
                        ws[f'G{contador}'] = 'SIM'
                    elif l in item:
                        ws[f'G{contador}'] = 'SIM'
                    elif m in item:
                        ws[f'G{contador}'] = 'SIM'
                    elif n in item:
                        ws[f'G{contador}'] = 'SIM'
                    elif o in item:
                        ws[f'G{contador}'] = 'SIM'
                    elif p in item:
                        ws[f'G{contador}'] = 'SIM'
                    elif q in item:
                        ws[f'G{contador}'] = 'SIM'
                    elif r in item:
                        ws[f'G{contador}'] = 'SIM'
                    elif s in item:
                        ws[f'G{contador}'] = 'SIM'
                    elif t in item:
                        ws[f'G{contador}'] = 'SIM'
                    elif u in item:
                        ws[f'G{contador}'] = 'SIM'
                    elif v in item:
                        ws[f'G{contador}'] = 'SIM'
                    elif w in item:
                        ws[f'G{contador}'] = 'SIM'
                    elif x in item:
                        ws[f'G{contador}'] = 'SIM'
                    elif z in item:
                        ws[f'G{contador}'] = 'SIM'
                    elif conec2 in item:
                        ws[f'G{contador}'] = 'CONECTIVO'
                    elif conec in item:
                        ws[f'G{contador}'] = 'CONECTIVO'
                    else:
                        ws[f'G{contador}'] = 'NAO'
                        
                        
                            
                        ## VERIFICANDO VALOR MAIOR OU MENOR DE IDADE 
                    i = str(row[8].value)
                    # print((i))
                    ano_nas = (i[:4])
                    # print(len(ano_nas))
                    idade = float(ano_nas)
                    # print((idade))
                    resultado = ano_atual - idade
                    # print(resultado)
                    if resultado >= 18:
                            ws[f'J{contador}'] = 'MAIOR'
                    else:
                            ws[f'J{contador}'] = 'MENOR'
                ws = wb['Sheet1']
                ws.title = 'ORIGEM'        
                ws = wb.create_sheet('PRODUTOS ATIVOS ')
                ws = wb.create_sheet('TOTAL') 
                wb.save(diretorio + '\\' + arquivo)
    # #     # end_time = time.time()
    # #     # print(end_time-start_time)
    except Exception as e:
            logging.error('| Ocorreu um erro: | 3')
            logging.exception(str(e))    