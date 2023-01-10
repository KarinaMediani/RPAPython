import os
from openpyxl import load_workbook
import logging
import xml.etree.ElementTree as ET

def conf():
    try:
        tree = ET.parse("C:\\Users\\User\\Desktop\\206-BS-571\\diretório\\diretorio.xml")
        root2 = tree.getroot()
        for child2 in root2:
                for x2 in root2.findall(child2.tag):
                        confeop = x2.find('confeop').text
                        nomesop = x2.find('nomesop').text
        diretorio = (confeop)
        lista_arquivo = os.listdir(diretorio)
        # start_time = time.time()

        a = 'A'
        b = 'B'
        c = 'C'
        d = 'D'
        f = 'F'
        g = 'G'
        h = 'H'
        i = 'I'
        j = 'J'
        k = 'K'
        l = 'L'
        m = 'M'
        n = 'N'
        o = 'O'
        p = 'P'
        q = 'Q'
        r = 'R'
        s = 'S'
        t = 'T'
        u = 'U'
        v = 'V'
        w = 'W'
        x = 'X'
        z = 'Z'
        e = 'E'
        y = 'Y'
        diretorio_nomes_m = (nomesop)
        wb1 = load_workbook(diretorio_nomes_m)
        ws1 = wb1['Nomes']
        planilha_nomes = []
        for nome in ws1:
            planilha_nomes.append(str(nome[0].value).upper().strip())

        for arquivo in lista_arquivo:
                    # print(arquivo)
            if arquivo.startswith('ArqConf') and arquivo.endswith('.xlsx'):
                wb = load_workbook(diretorio + '\\' + arquivo)
                # print(wb)
                ws = wb['ORIGEM']
            
                #9.6.20/.21/.22
                contador = 1
                for celula in ws.iter_rows(min_row=2):
                    contador += 1
                    ## VERIFICAÇÃO COLUNA M 
                    item = str(celula[10].value).upper().strip().split()
                    # print(item)
                    item2 = str(celula[5].value).upper().strip().split()
                    if item == []:
                        continue
                    else:
                        coluna_k = str(item[0])#TODO VERIFICAR COLUNA K (21/11)
                    # print(coluna_k)
                    abrv = str(celula[10].value).split()
                    primeiro_nome = abrv[:1]
                    if celula[10].value == None:
                        ws[f'M{contador}'] = 'VAZIO'
                    elif celula[10].value == 'NULL':
                        ws[f'M{contador}'] = 'INVALIDO'
                    elif celula[10].value == 'NAO CONSTA EM REGISTRO':
                        ws[f'M{contador}'] = 'INVALIDO'
                    elif item == item2:
                        ws[f'M{contador}'] = 'INVALIDO'
                    elif coluna_k in planilha_nomes:
                        ws[f'M{contador}'] = 'INVALIDO'#TODO VERIFICAR COLUNA K (21/11)
                    elif a in primeiro_nome:
                        ws[f'M{contador}'] = 'INVALIDO'
                    elif b in primeiro_nome:
                        ws[f'M{contador}'] = 'INVALIDO'
                    elif c in primeiro_nome:
                        ws[f'M{contador}'] = 'INVALIDO'
                    elif d in primeiro_nome:
                        ws[f'M{contador}'] = 'INVALIDO'
                    elif f in primeiro_nome:
                        ws[f'M{contador}'] = 'INVALIDO'
                    elif g in primeiro_nome:
                        ws[f'M{contador}'] = 'INVALIDO'
                    elif h in primeiro_nome:
                        ws[f'M{contador}'] = 'INVALIDO'
                    elif i in primeiro_nome:
                        ws[f'M{contador}'] = 'INVALIDO'
                    elif j in primeiro_nome:
                        ws[f'M{contador}'] = 'INVALIDO'
                    elif k in primeiro_nome:
                        ws[f'M{contador}'] = 'INVALIDO'
                    elif l in primeiro_nome:
                        ws[f'M{contador}'] = 'INVALIDO'
                    elif m in primeiro_nome:
                        ws[f'M{contador}'] = 'INVALIDO'
                    elif n in primeiro_nome:
                        ws[f'M{contador}'] = 'INVALIDO'
                    elif o in primeiro_nome:
                        ws[f'M{contador}'] = 'INVALIDO'
                    elif p in primeiro_nome:
                        ws[f'M{contador}'] = 'INVALIDO'
                    elif q in primeiro_nome:
                        ws[f'M{contador}'] = 'INVALIDO'
                    elif r in primeiro_nome:
                        ws[f'M{contador}'] = 'INVALIDO'
                    elif s in primeiro_nome:
                        ws[f'M{contador}'] = 'INVALIDO'
                    elif t in primeiro_nome:
                        ws[f'M{contador}'] = 'INVALIDO'
                    elif u in primeiro_nome:
                        ws[f'M{contador}'] = 'INVALIDO'
                    elif v in primeiro_nome:
                        ws[f'M{contador}'] = 'INVALIDO'
                    elif w in primeiro_nome:
                        ws[f'M{contador}'] = 'INVALIDO'
                    elif x in primeiro_nome:
                        ws[f'M{contador}'] = 'INVALIDO'
                    elif z in primeiro_nome:
                        ws[f'M{contador}'] = 'INVALIDO'
                    elif e in primeiro_nome:
                        ws[f'M{contador}'] = 'INVALIDO'
                    elif y in primeiro_nome:
                        ws[f'M{contador}'] = 'INVALIDO'
                    else:
                        ws[f'M{contador}'] = 'VALIDO'
                        

                wb.save(diretorio + '\\' + arquivo)
    # #     # end_time = time.time()
    # #     # print(end_time-start_time)
    except Exception as e:
        logging.error('| Ocorreu um erro: | 3')
        logging.exception(str(e))  
        