from openpyxl import load_workbook
import logging
import xml.etree.ElementTree as ET
def ansop_X_Produtos():
    try:
        tree = ET.parse("C:\\Users\\User\\Desktop\\206-BS-571\\diretório\\diretorio.xml")
        root2 = tree.getroot()
        for child2 in root2:
                for x2 in root2.findall(child2.tag):
                        proativosop = x2.find('proativosop').text
                        ansop = x2.find('ansop').text
                        
        #arquivo Planos Saúde ansop
        diretorio = ansop
        wb = load_workbook(diretorio)
        ws = wb.active

        ##arquivo produtos ativos###
        diretorio_2 = proativosop
        wb1 = load_workbook(diretorio_2)
        contador = 1

        ws1 = wb1['Produtos Ativos']
        for linhas in ws1:
            coluna_A = str(linhas[0].value).upper().strip()
            for row in ws:
                if coluna_A == str(row[0].value):
                    ws1[f'B{contador}'] = str(row[4].value)
            contador +=1

        wb1.save(diretorio_2)
    except Exception as e:
                logging.error('| Ocorreu um erro: | 3')
                logging.exception(str(e))    