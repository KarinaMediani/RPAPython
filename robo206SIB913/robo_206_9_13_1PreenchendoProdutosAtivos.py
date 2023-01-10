from openpyxl import load_workbook
import logging
import xml.etree.ElementTree as ET
import os

def Confe_x_Produtos():
    try:
        tree = ET.parse("C:\\Users\\User\\Desktop\\206-BS-571\\diretório\\diretorio.xml")
        root2 = tree.getroot()
        for child2 in root2:
                for x2 in root2.findall(child2.tag):
                        confeop = x2.find('confeop').text
                        proativosop = x2.find('proativosop').text
                        


        diretorio = (confeop)
        lista_arquivo = os.listdir(diretorio)
        listaColA = []
        listaColC = []
        lista_A = None
        lista_C = None
        for arquivo in lista_arquivo: 
            if arquivo.startswith('ArqConf') and arquivo.endswith('.xlsx'):
                wb = load_workbook(diretorio + '\\' + arquivo)
                ws = wb['PRODUTOS ATIVOS ']
                #Acessando valores das colunas A e C da planilha Conferencia 
                for celula in ws.iter_rows(min_row=2):
                    # print(celula[0].value)
                    col_A = str(celula[0].value).strip()
                    if col_A == "None":
                        continue
                    else:
                        listaColA.append(col_A)
                    # print(listaColA)
                    if celula[1].value == None:
                        continue
                    else:
                        col_C = float(celula[1].value)
                    # listaColA.append(col_A)
                        listaColC.append(col_C)
        # ws.delete_rows(0)


        ##arquivo produtos ativos###
        dir_2 = proativosop
        wb1 = load_workbook(dir_2)
        ws1 = wb1['Produtos Ativos']



        # print(soma_C)
                    
        lista_A = listaColA
        # print(lista_A)
        lista_C = listaColC

        #Colando itens da coluna A para Produtos ativos
        contador = 3
        for item in lista_A: 
            ws1[f'A{contador}'] = item
            contador +=1

        #Colando itens da coluna C para Produtos ativos
        contador = 3
        for item in lista_C:
            ws1[f'C{contador}'] = item
            contador +=1


        coluna_c = []
        coluna_a = []
        for itens in ws1.iter_rows(min_row=3):
            if itens[2].value == None:
                continue
            else: 
                colunaC = float(itens[2].value)
                coluna_c.append(colunaC)
            colunaA = str(itens[0].value)
            coluna_a.append(colunaA)
            produtos = len(coluna_a)
            vidas = sum(coluna_c)
        # print(vidas)
        ws1['A2'] = produtos #adicionando soma da coluna A
        ws1['C2'] = vidas #adicionando soma da coluna c

        wb1.save(dir_2)
    except Exception as e:
                logging.error('| Ocorreu um erro: | 3')
                logging.exception(str(e))    





                    