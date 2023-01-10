from typing import Counter
from openpyxl import load_workbook
from openpyxl.styles import Font,Border,Side
from operator import itemgetter
import logging
import xml.etree.ElementTree as ET
def ProdutosAtivos():
    try:
        tree = ET.parse("C:\\Users\\User\\Desktop\\206-BS-571\\diretório\\diretorio.xml")
        root2 = tree.getroot()
        for child2 in root2:
                for x2 in root2.findall(child2.tag):
                        proativosop = x2.find('proativosop').text
                        apcia = x2.find('apcia').text
        #ABRINDO ARQUIVO
        diretorio_2 = proativosop
        wb1 = load_workbook(diretorio_2)
        ws1 = wb1['Produtos Ativos']
        seg = []#LISTA COM OS VALORES DA COLUNA B
        am = []
        ho = []
        amb = []
        ams = []
        re = []
        hob=[] 
        amso= []
        ambulatorial = []
        soma_AM =None
        soma_ho1 = None 
        soma_1  = None
        soma_2  = None
        soma_re = None 
        soma_hob = None
        soma_3  = None
        soma_ambulatorial = None
        total = None
        # ws1.delete_rows(0) #DELETANDO CABEÇALHO
        for row in ws1.iter_rows(min_row=3):
            contador = 3 
            segm = str(row[1].value)
            # print(segm)
            if "Ambulatorial + Hospitalar com obstetrícia" in segm:
                am.append(row[2].value)
                soma_AM = sum(am)
            if soma_AM == None:
                soma_AM = 0
                # print(am)
            if "Hospitalar com obstetrícia" in str(row[1].value).strip():
                if not "Ambulatorial + Hospitalar com obstetrícia" in str(row[1].value).strip():
                    ho.append(row[2].value)
                    soma_ho1 = sum(ho)
                if soma_ho1 == None:
                    soma_ho1 = 0

            if "AMBULATORIAL + HOSPITALAR COM OBSTETRÍCIA" in segm:
                amb.append(row[2].value)
                soma_1 = sum(amb)
            if not "AMBULATORIAL + HOSPITALAR COM OBSTETRÍCIA" in segm:
                soma_1 = 0
        # print(soma_ho)
        # print(soma_1)
            if "Ambulatorial + Hospitalar sem obstetrícia" in segm:
                ams.append(row[2].value)
                soma_2 = sum(ams)
            if soma_2 == None:
                soma_2 = 0
        # print(len(ams))
        # print(soma_2)
            if "Referência" in segm:
                re.append(row[2].value)
                soma_re = sum(re)
            if soma_re == None:
                soma_re = 0
        # print(len(re))
        # print(soma_re)
            if "Hospitalar sem obstetrícia" in segm:
                if not "Ambulatorial + Hospitalar sem obstetrícia" in str(row[1].value).strip():
                    hob.append(row[2].value)
                    soma_hob = sum(hob)
                if soma_hob == None:
                    soma_hob = 0
        # print(len(hob))
        # print(soma_hob)
            if "AMBULATORIAL + HOSPITALAR SEM OBSTETRÍCIA" in segm:
                amso.append(row[2].value)
                soma_3 = sum(amso)
            if not "AMBULATORIAL + HOSPITALAR SEM OBSTETRÍCIA" in segm:
                soma_3 = 0
        # print(len(amso))
        # print(soma_3)
            
            if  "Ambulatorial" in segm:
                if not "Ambulatorial + Hospitalar sem obstetrícia" in str(row[1].value).strip():
                    if not "Ambulatorial + Hospitalar com obstetrícia" in str(row[1].value).strip():
                        ambulatorial.append(row[2].value)
                        soma_ambulatorial = sum(ambulatorial)
            if soma_ambulatorial == None:
                soma_ambulatorial = 0
        # print(soma_ambulatorial)
        # print(len(ambulatorial))
        # print(soma_ambulatorial)
        total = (soma_AM + soma_ho1 + soma_1 + soma_2 + soma_re + soma_hob + soma_3 + soma_ambulatorial)

        ws1["E3"] = "Ambulatorial + Hospitalar com obstetrícia"
        ws1["E4"] = "Hospitalar com obstetrícia"
        ws1["E5"] = "AMBULATORIAL + HOSPITALAR COM OBSTETRÍCIA"
        ws1["E6"] = "Ambulatorial + Hospitalar sem obstetrícia"
        ws1["E7"] = "Referência"
        ws1["E8"] = "Hospitalar sem obstetrícia"
        ws1["E9"] = "AMBULATORIAL + HOSPITALAR SEM OBSTETRÍCIA"
        ws1["E10"] = "Ambulatorial"
        ws1["E11"] = "TOTAL"
        ws1["F3"] = soma_AM
        ws1["F4"] = soma_ho1
        ws1["F5"] = soma_1
        ws1["F6"] = soma_2
        ws1["F7"] = soma_re
        ws1["F8"] = soma_hob
        ws1["F9"] = soma_3
        ws1["F10"] = soma_ambulatorial
        ws1["F11"] = total
        ws1[f'E3'].font = Font(bold= True) #NEGRITO
        ws1[f'F3'].font = Font(bold= True) #NEGRITO
        ws1[f'E4'].font = Font(bold= True) #NEGRITO
        ws1[f'F4'].font = Font(bold= True) #NEGRITO
        ws1[f'E5'].font = Font(bold= True) #NEGRITO
        ws1[f'F5'].font = Font(bold= True) #NEGRITO
        ws1[f'E6'].font = Font(bold= True) #NEGRITO
        ws1[f'F6'].font = Font(bold= True) #NEGRITO
        ws1[f'E7'].font = Font(bold= True) #NEGRITO
        ws1[f'F7'].font = Font(bold= True) #NEGRITO
        ws1[f'E8'].font = Font(bold= True) #NEGRITO
        ws1[f'F8'].font = Font(bold= True) #NEGRITO
        ws1[f'E9'].font = Font(bold= True) #NEGRITO
        ws1[f'F9'].font = Font(bold= True) #NEGRITO
        ws1[f'E10'].font = Font(bold= True) #NEGRITO
        ws1[f'F10'].font = Font(bold= True) #NEGRITO
        ws1[f'E11'].font = Font(bold= True) #NEGRITO
        ws1[f'F11'].font = Font(bold= True) #NEGRIT4


        cel_E3 = ws1['E3']
        cel_E4 = ws1['E4']
        cel_E5 = ws1['E5']
        cel_E6 = ws1['E6']
        cel_E7 = ws1['E7']
        cel_E8 = ws1['E8']
        cel_E9 = ws1['E9']
        cel_E10 = ws1['E10']
        cel_E11 = ws1['E11']

        cel_F3 = ws1['F3']
        cel_F4 = ws1['F4']
        cel_F5 = ws1['F5']
        cel_F6 = ws1['F6']
        cel_F7 = ws1['F7']
        cel_F8 = ws1['F8']
        cel_F9 = ws1['F9']
        cel_F10 = ws1['F10']
        cel_F11 = ws1['F11']

        thin = Side(border_style="thin", color="000000")
        cel_E3.border = Border(top= thin, left= thin, right= thin, bottom= thin)#BORDAS
        cel_E4.border = Border(top= thin, left= thin, right= thin, bottom= thin)#BORDAS
        cel_E5.border = Border(top= thin, left= thin, right= thin, bottom= thin)#BORDAS
        cel_E6.border = Border(top= thin, left= thin, right= thin, bottom= thin)#BORDAS
        cel_E7.border = Border(top= thin, left= thin, right= thin, bottom= thin)#BORDAS
        cel_E8.border = Border(top= thin, left= thin, right= thin, bottom= thin)#BORDAS
        cel_E9.border = Border(top= thin, left= thin, right= thin, bottom= thin)#BORDAS
        cel_E10.border = Border(top= thin, left= thin, right= thin, bottom= thin)#BORDAS
        cel_E11.border = Border(top= thin, left= thin, right= thin, bottom= thin)#BORDAS
        cel_F3.border = Border(top= thin, left= thin, right= thin, bottom= thin)#BORDAS
        cel_F4.border = Border(top= thin, left= thin, right= thin, bottom= thin)#BORDAS
        cel_F5.border = Border(top= thin, left= thin, right= thin, bottom= thin)#BORDAS
        cel_F6.border = Border(top= thin, left= thin, right= thin, bottom= thin)#BORDAS
        cel_F7.border = Border(top= thin, left= thin, right= thin, bottom= thin)#BORDAS
        cel_F8.border = Border(top= thin, left= thin, right= thin, bottom= thin)#BORDAS
        cel_F9.border = Border(top= thin, left= thin, right= thin, bottom= thin)#BORDAS
        cel_F10.border = Border(top= thin, left= thin, right= thin, bottom= thin)#BORDAS
        cel_F11.border = Border(top= thin, left= thin, right= thin, bottom= thin)#BORDAS



        wb1.save(diretorio_2)
    except Exception as e:
                logging.error('| Ocorreu um erro: | 3')
                logging.exception(str(e))    