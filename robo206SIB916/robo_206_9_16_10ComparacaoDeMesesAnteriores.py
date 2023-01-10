import logging
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.styles import Border,Side,PatternFill
import numpy as np
def meses_Anteriores2():
    try:
        tree = ET.parse("C:\\Users\\User\\Desktop\\206-BS-571\\diretório\\diretorio.xml")
        root2 = tree.getroot()
        for child2 in root2:
                for x2 in root2.findall(child2.tag):
                        confeop = x2.find('confeop').text
                        proativosop = x2.find('proativosop').text
        diretorio = proativosop
        wb = load_workbook(diretorio)
        ws = wb['Comparação Meses Anteriores']
        ws2 = wb['Situação do Plano']#COPIANDO VALORES NOVOS ABA SITUAÇÃO DO PLANO

        # PEGANDO NOVOS VALORES
        nao_duplicados = []
        for linhas in ws.iter_rows(min_row=5):
            cola= str(linhas[0].value).strip()
            cole = str(linhas[4].value).strip()
            if cola == cole:
                continue
            else:
                nao_duplicados.append(cola)

        #pegando valores da coluna C ABA Situação do Plano e adicionando na coluna B ABA 'Comparação Meses Anteriores'
        l = []
        contador = 5
        for linhas in ws.iter_rows(min_row=5):
            coluna_A = str(linhas[0].value).upper().strip()
            l.append(coluna_A)
            for row in ws2.iter_rows(min_row=2):
                if coluna_A == str(row[0].value):
                    ws[f'B{contador}'] = str(row[2].value)
            contador +=1

        contador = 4
        for linhas in ws.iter_rows(min_row=4):   
            colunaB = str(linhas[1].value).upper().strip()
            # print(colunaB)
            if colunaB == "NONE":
                ws[f'B{contador}'] = 0
            contador += 1



        contador2 = 5
        # PEGANDO VALORES DA COLUNA B PARA REALIZAR SOMA TOTAL
        total_vidas = []
        for celula in ws.iter_rows(min_row=4):
            if celula[1].value == None:
                continue
            else:
                col_B = float(celula[1].value)
            total_vidas.append(col_B)
        soma_vidas = sum(total_vidas)#somando total de valores da coluna da coluna B

        ws['B4'] = soma_vidas

        col_b = []
        col_f = []

        #pegando valores da coluna B e F
        for colunas in ws.iter_rows(min_row=4):
            if colunas[1].value == None:
                continue
            else:
                col_b.append(int(colunas[1].value))
            if colunas[5].value == None:
                continue
            else:
                col_f.append(int(colunas[5].value))
            
        #Realizando subtração da coluna b e f
        cb = np.array(col_b)
        cf = np.array(col_f)
        resultado = (cb-cf)
        contador = 4
        for item in resultado:
            ws[f'C{contador}'] = item
            contador += 1





        #PEGANDO AS COLUNAS PARA SEREM PREENCHIDAS COM NOVOS VALORES
        thin = Side(border_style="thin", color="000000")
        for column in range(5,ws.max_column,4):#pegando a quinta coluna e andando de 4 em 4 colunas
            if (ws.cell(row= 1 , column= column).value) != None:#travando a leitura nas colunas especificas
                contador = 0
                coluna_ref = (ws.cell(row = 1, column= column).column_letter)
                coluna_vida = (ws.cell(row = 3, column= column + 1).column_letter)
                coluna_diferenca = (ws.cell(row = 3, column= column + 2).column_letter)
                for col in ws.iter_cols(min_col= column, max_col= column):#percorrendo as linhas nas colunas especificas
                    for celula in col:#lendo as linhas
                        if celula.value != None:
                            contador += 1

        #inserindo novos valores nas colunas E em diante
                contador += 1
                for item in nao_duplicados:
                    ws[f'{coluna_ref}{contador}'] = item
                    ws[f'{coluna_vida}{contador}'] = '0'
                    ws[f'{coluna_diferenca}{contador}'] = '0'
                    ws[f'{coluna_ref}{contador}'].border = Border(top= thin, left= thin, right= thin, bottom= thin)#BORDAS
                    ws[f'{coluna_vida}{contador}'].border = Border(top= thin, left= thin, right= thin, bottom= thin)#BORDAS
                    ws[f'{coluna_diferenca}{contador}'].fill = PatternFill("solid", start_color="87CEEB")#cor de fundo
                    ws[f'{coluna_diferenca}{contador}'].border = Border(top= thin, left= thin, right= thin, bottom= thin)#BORDAS
                    contador += 1


        wb.save(diretorio)
    except Exception as e:
                logging.error(' | Ocorreu um erro: | 3 | '+ str(e))
                logging.exception(str(e)) 