import datetime
from openpyxl import load_workbook
import locale
from openpyxl.styles import Border,Side,PatternFill
import logging
import xml.etree.ElementTree as ET
def meses_Anteriores():
    try:
        tree = ET.parse("C:\\Users\\User\\Desktop\\206-BS-571\\diretório\\diretorio.xml")
        root2 = tree.getroot()
        for child2 in root2:
                for x2 in root2.findall(child2.tag):
                        confeop = x2.find('confeop').text
                        proativosop = x2.find('proativosop').text
        diretorio = proativosop
        wb = load_workbook(diretorio)

        nao_duplicados = []
        coluna_A =[]
        redFill = PatternFill(bgColor= '1e90ff')
        ws = wb['Comparação Meses Anteriores']
        ws.insert_cols(0)
        ws.insert_cols(0)
        ws.insert_cols(0)
        ws.insert_cols(0)


        for linha in ws.iter_rows(min_row=5):
            col_a = str(linha[4].value)
            if col_a == "None":
                continue
            else:
            # print(col_a)
                coluna_A.append(col_a)
            ws['B3'] = 'VIDAS'
            ws['C3'] = 'DIFERENÇA'
        # print(coluna_A)
        ws2 = wb['Situação do Plano']#COPIANDO VALORES NOVOS ABA SITUAÇÃO DO PLANO 
        novosValores = []#lista com novos valores para as outras colunas
        for row in ws2.iter_rows(min_row=2):
            rows = str(row[0].value)
            # print(rows)
            if rows == "None":
                continue
            else:
                novosValores.append(rows)
        # del novosValores[0]
        # print(coluna_A)
        for itens in novosValores:#adicionando todos os valores em uma so lista
            coluna_A.append(itens)

        # print(coluna_A)
        # print(len(coluna_A))

        for i in coluna_A:#retirando os valores duplicados
            if i not in nao_duplicados:
                nao_duplicados.append(i)
        # print(len(nao_duplicados)  )      
        #COLOCANDO VALORES NA COLUNA A 
        contador = 5
        for item in nao_duplicados: 
            ws[f'A{contador}'] = item
            contador +=1


        #RENOMEANDO COM MES E ANO ATUAL
        data_atual = datetime.datetime.now()
        data = data_atual.date()
        locale.setlocale(locale.LC_ALL, '')
        data_br = data.strftime("%b/%Y")
        ws['A1'] = data_br

        ws1 = wb['Total']
        for linha in ws1:#copiado valor da celula B14 ABA TOTAL 
            colu_b = ws1['B14'].value
            valores = str(colu_b)

        ws['A2'] = valores + " PRODUTOS / 0 SEG. SEM PROD."
        ws['A3'] = "PRODUTO"
        ws['A4'] = "TOTAL"




        # thin = Side(border_style="thin", color="000000")
        # contadore = 3
        # for linha in ws:
        #     if linha[0].value != None:
        #         ws[f"C{contadore}"].fill = PatternFill("solid", start_color="87CEEB")#cor de fundo
        #         ws[f"A{contadore}"].border = Border(top= thin, left= thin, right= thin, bottom= thin)#BORDAS
        #         ws[f"B{contadore}"].border = Border(top= thin, left= thin, right= thin, bottom= thin)#BORDAS
        #         ws[f"C{contadore}"].border = Border(top= thin, left= thin, right= thin, bottom= thin)#BORDAS
        #     contadore += 1

        wb.save(diretorio)
    except Exception as e:
                logging.error('| Ocorreu um erro: | 3')
                logging.exception(str(e))    