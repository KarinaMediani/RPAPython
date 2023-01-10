import logging
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
import os
#2
def exlusao():
    try:
        tree = ET.parse("C:\\Users\\User\\Desktop\\206-BS-571\\diretório\\diretorio.xml")
        root2 = tree.getroot()
        for child2 in root2:
                for x2 in root2.findall(child2.tag):
                        confeop = x2.find('confeop').text
        diretorio =confeop
        lista_arquivo = os.listdir(diretorio)

        for arquivo in lista_arquivo:
            if arquivo.startswith('ArqConf') and arquivo.endswith('.xlsx'):
                wb = load_workbook(diretorio+'\\'+arquivo)
                ws = wb.active
                contador = 0
                for item in ws:
                    contador += 1
                    linha = str(item[0].value)
                    status__ = linha[12:19]
                    # print(status__)
                    if status__ == 'INATIVO':
                        # print(status__)
                        ws[f'A{contador}'] = ''

                wb.save(diretorio+'\\'+arquivo)
    except Exception as e:
                logging.error('| Ocorreu um erro: | 3')
                logging.exception(str(e))    

