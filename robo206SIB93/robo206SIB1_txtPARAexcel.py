import logging
import xml.etree.ElementTree as ET
import os
from openpyxl import Workbook
from openpyxl import load_workbook
#ja é o novo 1
def para_excel():
    try:
        tree = ET.parse("C:\\Users\\User\\Desktop\\206-BS-571\\diretório\\diretorio.xml")
        root2 = tree.getroot()
        for child2 in root2:
                for x2 in root2.findall(child2.tag):
                        confeop = x2.find('confeop').text
                        apcia = x2.find('apcia').text
        dir_1 = confeop
        list_dir1 = os.listdir(dir_1)
        arq = []
        # print(len(list_dir1))
        for arquivo_txt in list_dir1:
            if arquivo_txt.startswith("ArqConf") and arquivo_txt.endswith('.TXT'):
                arq.append(arquivo_txt)
        qnd_arq = (len(arq))
        li = []
        total_as = None
        for arquivo_txt in list_dir1:
            if arquivo_txt.startswith("ArqConf") and arquivo_txt.endswith('.TXT'):
                # arq = list(arquivo_txt)
            # if arquivo_txt.startswith("ArqConf") and arquivo_txt.endswith('.txt'):
                
                file_text = open((dir_1+'\\'+arquivo_txt),"r",encoding='latin-1')
                
                arquivo = file_text.read()
                    
                lista_dados = arquivo.splitlines()
                l = len(lista_dados)
                li.append(l)
                s = sum(li)
                total_as = (s - qnd_arq)



                for i in range(0,len(lista_dados)):
                    # nome_arquivo = arquivo[:arquivo.index('.')]
                    lista_dados[i] = lista_dados[i].split(',')

                wb = Workbook()
                ws = wb.active

                for row in lista_dados:
                    ws.append(row)
                wb.save((dir_1+'\\'+arquivo_txt)+ '.xlsx')
        dir_2 = apcia
        wb1= load_workbook(dir_2)
        ws1 = wb1.active
        ws1["B1"] = total_as
        wb1.save(dir_2)
    except Exception as e:
                logging.error('| Ocorreu um erro: | 3')
                logging.exception(str(e))    