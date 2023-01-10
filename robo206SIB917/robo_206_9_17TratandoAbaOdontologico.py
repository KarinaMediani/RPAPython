import logging
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
import os
def odonto():
    try:
        tree = ET.parse("C:\\Users\\User\\Desktop\\206-BS-571\\diretório\\diretorio.xml")
        root2 = tree.getroot()
        for child2 in root2:
                for x2 in root2.findall(child2.tag):
                        confeop = x2.find('confeop').text
                        proativosop = x2.find('proativosop').text
        diretorio = proativosop
        wb = load_workbook(diretorio)
        ws = wb['Produtos Ativos']#primeiro parametro 
        vAbaOdonto = wb['Produto Odontológico']#planilha onde os dados iram ser adicionados

        vListaColunaA = []
        for vLinhasProdutosAtivos in ws:#PLANILHA ABA PLANILHA1
            vColunaB = str(vLinhasProdutosAtivos[1].value).upper().strip()#pegando valores da coluna B
            vColunaA = str(vLinhasProdutosAtivos[0].value).upper().strip()#pegando valores da coluna A
            if vColunaB == 'ODONTOLOGICA':
                vListaColunaA.append(vColunaA)  

                
        diretorio2 = (confeop)
        lista_arquivo = os.listdir(diretorio2)

        for arquivo in lista_arquivo: 
                    # print(arquivo)
            if arquivo.startswith('ArqConf') and arquivo.endswith('.xlsx'):
                wb2 = load_workbook(diretorio2 + '\\' + arquivo)
                ws2 = wb2['ORIGEM']#Planilha utilizada para fazer as comparações e extrair dados
                vListaLinhaInteira = []
                for vLinhasOrigem in ws2:#PLANILHA ORIGEM
                    vColunaAI = str(vLinhasOrigem[34].value).upper().strip()#pegando valores da coluna AI
                    vColunaAO = str(vLinhasOrigem[40].value).upper().strip()#pegando valores da coluna AO
                    for itens in vLinhasOrigem:#PLANILHA ORIGEM
                        if (vColunaAO in vListaColunaA) or (vColunaAI in vListaColunaA):
                            for itens in vLinhasOrigem:#itens da PLANILHA ORIGEM a serem transopportador para a planilha Produtos Odontologicos
                                vListaLinhaInteira.append(itens.value)#lista de linhas a serem transopportadas
                            vAbaOdonto.append(vListaLinhaInteira)#escrevendo as linhas na planilha Produtos Odontologicos
                            vListaLinhaInteira = []
                        break
        wb.save(diretorio)
    except Exception as e:
                logging.error('| Ocorreu um erro: | 3')
                logging.exception(str(e))    